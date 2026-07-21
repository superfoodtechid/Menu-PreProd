#!/usr/bin/env python3
"""
=================================================================
  SUPERFOOD TECH — Unified Menu & Modifier Extractor Pipeline
  Interactive CLI for Shopee, Grab & GoFood
=================================================================
"""

import os
import sys
import time
import re
import shutil
import glob
from datetime import datetime

# Add parent directory of menu_core to sys.path so imports work
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from menu_core.sheets import get_outlets_for_applicator
from shopee.core.pull import extract_shopee_menu
from menu_core.grab import extract_grab_menu
from menu_core.gofood import extract_gofood_menu

import openpyxl
import pandas as pd
from upload_drive import upload_combined_to_drive

FILE_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORTS_DIR = os.path.join(FILE_DIR, "data", "exports")

def combine_c5(excel_paths, output_path):
    all_items = []
    all_mods = []
    for f in excel_paths:
        if os.path.exists(f):
            try:
                df_item = pd.read_excel(f, sheet_name='Item')
                df_mod = pd.read_excel(f, sheet_name='Modifier')
                all_items.append(df_item)
                all_mods.append(df_mod)
            except Exception as e:
                print(f"  \033[91mError reading {f} for combine: {e}\033[0m")
                
    if not all_items:
        return False
        
    df_combined_items = pd.concat(all_items, ignore_index=True)
    df_combined_mods = pd.concat(all_mods, ignore_index=True)
    
    template_path = os.path.join(FILE_DIR, "O. C5 Template.xlsx")
    try:
        wb = openpyxl.load_workbook(template_path)
        sheet_item = wb['Item']
        if sheet_item.max_row > 1:
            sheet_item.delete_rows(2, sheet_item.max_row - 1)
            
        headers_item = {cell.value: cell.column for cell in sheet_item[1]}
        for r_idx, row in df_combined_items.iterrows():
            for col_name, val in row.items():
                if col_name in headers_item:
                    cell = sheet_item.cell(row=r_idx + 2, column=headers_item[col_name], value=val)
                    if '%' in str(col_name):
                        cell.number_format = '0%'
                    
        sheet_mod = wb['Modifier']
        if sheet_mod.max_row > 1:
            sheet_mod.delete_rows(2, sheet_mod.max_row - 1)
            
        headers_mod = {cell.value: cell.column for cell in sheet_mod[1]}
        for r_idx, row in df_combined_mods.iterrows():
            for col_name, val in row.items():
                if col_name in headers_mod:
                    cell = sheet_mod.cell(row=r_idx + 2, column=headers_mod[col_name], value=val)
                    if '%' in str(col_name):
                        cell.number_format = '0%'
                    
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"  \033[91mFailed to write combined C5 to template: {e}\033[0m")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_combined_items.to_excel(writer, sheet_name='Item', index=False)
            df_combined_mods.to_excel(writer, sheet_name='Modifier', index=False)
        return True

def check_outlet_processed(applicator, o, exports_dir=EXPORTS_DIR):
    raw_outlet = o.get('nama_outlet') or o.get('nama_resto_final') or o.get('merchant_name') or 'unknown'
    raw_brand = o.get('brand') or ''
    
    def clean_filename_part(s):
        return "".join(c for c in s if c.isalnum() or c in (' ', '_', '-')).strip()
        
    clean_outlet_filename = clean_filename_part(raw_outlet)
    excel_filename = f"O.C5 {clean_outlet_filename}.xlsx"
        
    clean_outlet_folder = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
    clean_outlet_folder = re.sub(r'\s+', ' ', clean_outlet_folder).lower()
    
    excel_path = os.path.join(exports_dir, applicator, clean_outlet_folder, excel_filename)
    return os.path.exists(excel_path)

RESET   = "\033[0m"
BOLD    = "\033[1m"
GREEN   = "\033[92m"
CYAN    = "\033[96m"
YELLOW  = "\033[93m"
RED     = "\033[91m"
MAGENTA = "\033[95m"
DIM     = "\033[2m"

def banner():
    print(f"\033[90m=================================================================\033[0m")
    print(f"  {BOLD}{CYAN}      SUPERFOOD TECH — MENU & MODIFIER EXTRACTOR PIPELINE{RESET}")
    print(f"\033[90m=================================================================\033[0m")
    print()

def clear_all_caches():
    print(f"\n  {YELLOW}[*] Membersihkan cache master spreadsheet...{RESET}")
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Master Merchants Cache (Google Sheets)
    master_cache = os.path.join(base_dir, "master_merchants_cache.csv")
    if os.path.exists(master_cache):
        try:
            os.remove(master_cache)
            print(f"    {DIM}- Terhapus: master_merchants_cache.csv{RESET}")
        except Exception as e:
            print(f"    {RED}- Gagal menghapus master_merchants_cache.csv: {e}{RESET}")
                
    print(f"  {GREEN}✔ Cache data spreadsheet berhasil dibersihkan! (Session & Profile tetap aman){RESET}")
    time.sleep(2)


def interactive_menu():
    state = "applicator"
    applicator = None
    outlets = []
    selected_outlet = None
    all_shopee = []
    all_grab = []
    all_gofood = []
    
    while True:
        if state == "applicator":
            os.system('cls' if os.name == 'nt' else 'clear')
            banner()
            print(f"  {BOLD}Pilih Aplikator/Platform:{RESET}")
            print(f"    {MAGENTA}[1]{RESET} ShopeeFood")
            print(f"    {GREEN}[2]{RESET} GrabFood")
            print(f"    {CYAN}[3]{RESET} GoFood")
            print(f"    {YELLOW}[4]{RESET} Semua (Jadikan 1 C5)")
            print(f"    {DIM}[5]{RESET} Clear Cache (Reset Sesi)")
            print(f"    {CYAN}[6]{RESET} Perbaiki Login Shopee (Manual OTP)")
            print(f"    {RED}[7]{RESET} Keluar")
            print()
            
            choice = input(f"  {BOLD}Pilihan (1/2/3/4/5/6/7):{RESET} ").strip()
            if choice == "7":
                print("  Keluar.")
                sys.exit(0)
            elif choice == "6":
                print(f"\n  {CYAN}=== PERBAIKI LOGIN SHOPEE ==={RESET}")
                uname = input(f"  {BOLD}Username Shopee:{RESET} ").strip()
                upass = input(f"  {BOLD}Password Shopee:{RESET} ").strip()
                if uname and upass:
                    print(f"  [*] Membuka browser Chrome (headless=False) untuk login manual...")
                    # Set up session file path for shopee
                    automation_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "src", "shopee-omzet-automation")
                    if automation_dir not in sys.path:
                        sys.path.insert(0, automation_dir)
                    from core import browser as shopee_browser
                    
                    session_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "shopee", "data", "session.json")
                    shopee_browser.set_session_file(session_file)
                    
                    session_data = shopee_browser.get_session(
                        username=uname, 
                        password=upass, 
                        headless=False, 
                        close_browser=True, 
                        interactive=True,
                        allow_otp=True
                    )
                    if session_data and "shopee_tob_token" in session_data:
                        print(f"  {GREEN}✔ Login berhasil diperbaiki dan session tersimpan!{RESET}\n")
                    else:
                        print(f"  {RED}✘ Gagal memperbaiki login.{RESET}\n")
                else:
                    print(f"  {RED}✘ Username dan Password tidak boleh kosong!{RESET}\n")
                state = "applicator"
            elif choice == "5":
                clear_all_caches()
                state = "applicator"
            elif choice == "1":
                applicator = "shopee"
                state = "load_outlets"
            elif choice == "2":
                applicator = "grab"
                state = "load_outlets"
            elif choice == "3":
                applicator = "gofood"
                state = "load_outlets"
            elif choice == "4":
                applicator = "all"
                state = "load_outlets"
            else:
                print(f"  {RED}Pilihan tidak valid.{RESET}")
                time.sleep(1)
                
        elif state == "load_outlets":
            print(f"\n  [*] Mengunduh & memuat daftar outlet untuk {applicator.upper()}...")
            try:
                if applicator == "all":
                    all_shopee = get_outlets_for_applicator("shopee")
                    all_grab = get_outlets_for_applicator("grab")
                    all_gofood = get_outlets_for_applicator("gofood")
                    
                    seen = set()
                    outlets = []
                    for o_list in [all_shopee, all_grab, all_gofood]:
                        for o in o_list:
                            n_outlet = str(o.get('nama_outlet') or '').strip().lower()
                            ident = f"{n_outlet}"
                            if ident and ident not in seen:
                                seen.add(ident)
                                outlets.append(o)
                                
                    # Store mapping for later use
                    # To keep it simple, we just return the master list of outlets, and inside main() we will fetch again or match
                else:
                    outlets = get_outlets_for_applicator(applicator)
                    
                if not outlets:
                    print(f"  {RED}[ERROR] Tidak ada outlet live yang ditemukan.{RESET}")
                    time.sleep(2)
                    state = "applicator"
                else:
                    state = "select_outlet"
            except Exception as e:
                print(f"  {RED}[ERROR] Gagal memuat daftar outlet: {e}{RESET}")
                time.sleep(3)
                state = "applicator"
                
        elif state == "select_outlet":
            os.system('cls' if os.name == 'nt' else 'clear')
            banner()
            
            # Get unique Nama Outlet (nama_outlet) values
            unique_outlets = sorted(list(set(o['nama_outlet'] for o in outlets if o['nama_outlet'])))
            
            print(f"  {BOLD}Pilih Nama Outlet {applicator.upper()}:{RESET}")
            print(f"    {GREEN}[all]{RESET} Jalankan semua outlet dan cabang")
            print(f"    {GREEN}[new]{RESET} Jalankan HANYA outlet/cabang yang belum ditarik")
            for idx, name in enumerate(unique_outlets):
                print(f"    {GREEN}[{idx + 1:3d}]{RESET} {name}")
                
            print(f"    {YELLOW}[b  ]{RESET} Kembali ke pemilihan aplikator")
            print()
            
            choice = input(f"  {BOLD}Pilih nomor outlet (atau 'all'/'new'/'b'):{RESET} ").strip()
            if choice.lower() == 'b':
                state = "applicator"
            elif choice.lower() == 'all':
                selected_outlet = outlets
                state = "confirm_all"
            elif choice.lower() == 'new':
                # Filter outlets to keep only those that have not been run
                filtered_outlets = []
                for o in outlets:
                    if not check_outlet_processed(applicator, o):
                        filtered_outlets.append(o)
                
                if not filtered_outlets:
                    print(f"\n  {GREEN}Semua outlet sudah berhasil ditarik sebelumnya!{RESET}")
                    time.sleep(3)
                else:
                    selected_outlet = filtered_outlets
                    state = "confirm_all"
            else:
                try:
                    idx = int(choice) - 1
                    if 0 <= idx < len(unique_outlets):
                        target_parent = unique_outlets[idx]
                        # Find all branches under this parent
                        matching_branches = [o for o in outlets if o['nama_outlet'] == target_parent]
                        
                        if len(matching_branches) == 1:
                            selected_outlet = matching_branches[0]
                            state = "confirm"
                        else:
                            parent_name = target_parent
                            branches = matching_branches
                            state = "select_branch"
                    else:
                        print(f"  {RED}Nomor outlet di luar jangkauan.{RESET}")
                        time.sleep(1)
                except ValueError:
                    print(f"  {RED}Pilihan tidak valid.{RESET}")
                    time.sleep(1)
                    
        elif state == "select_branch":
            os.system('cls' if os.name == 'nt' else 'clear')
            banner()
            print(f"  {BOLD}Pilih Cabang untuk '{parent_name}':{RESET}")
            print(f"    {GREEN}[all]{RESET} Jalankan semua cabang untuk outlet ini")
            print(f"    {GREEN}[new]{RESET} Jalankan HANYA cabang yang belum ditarik")
            
            for idx, b in enumerate(branches):
                branch_name = b['brand'] or b['nama_resto_final'] or b['merchant_name']
                print(f"    {GREEN}[{idx + 1:3d}]{RESET} {branch_name} (ID: {b['store_id']})")
                
            print(f"    {YELLOW}[b  ]{RESET} Kembali ke pemilihan outlet")
            print()
            
            choice = input(f"  {BOLD}Pilih nomor cabang (atau 'all'/'new'/'b'):{RESET} ").strip()
            if choice.lower() == 'b':
                state = "select_outlet"
            elif choice.lower() == 'all':
                selected_outlet = branches
                state = "confirm_all"
            elif choice.lower() == 'new':
                # Filter branches to keep only those that have not been run
                filtered_branches = []
                for o in branches:
                    if not check_outlet_processed(applicator, o):
                        filtered_branches.append(o)
                
                if not filtered_branches:
                    print(f"\n  {GREEN}Semua cabang untuk outlet ini sudah berhasil ditarik sebelumnya!{RESET}")
                    time.sleep(3)
                else:
                    selected_outlet = filtered_branches
                    state = "confirm_all"
            else:
                try:
                    idx = int(choice) - 1
                    if 0 <= idx < len(branches):
                        selected_outlet = branches[idx]
                        state = "confirm"
                    else:
                        print(f"  {RED}Nomor cabang di luar jangkauan.{RESET}")
                        time.sleep(1)
                except ValueError:
                    print(f"  {RED}Pilihan tidak valid.{RESET}")
                    time.sleep(1)
                    
        elif state == "confirm":
            os.system('cls' if os.name == 'nt' else 'clear')
            banner()
            print(f"  {CYAN}{'─'*60}{RESET}")
            print(f"  Aplikator : {BOLD}{applicator.upper()}{RESET}")
            name_to_show = selected_outlet['brand'] or selected_outlet['nama_resto_final'] or selected_outlet['nama_outlet']
            print(f"  Outlet    : {BOLD}{name_to_show}{RESET}")
            
            if applicator == "all":
                target_outlet = str(selected_outlet.get('nama_outlet') or '').strip().lower()
                s_ids = [str(o.get('store_id')) for o in all_shopee if str(o.get('nama_outlet') or '').strip().lower() == target_outlet and o.get('store_id')]
                g_ids = [str(o.get('store_id')) for o in all_grab if str(o.get('nama_outlet') or '').strip().lower() == target_outlet and o.get('store_id')]
                gf_ids = [str(o.get('store_id')) for o in all_gofood if str(o.get('nama_outlet') or '').strip().lower() == target_outlet and o.get('store_id')]
                
                print(f"  Store IDs :")
                if s_ids:
                    print(f"    - ShopeeFood ({len(s_ids)}) = {', '.join(s_ids)}")
                if g_ids:
                    print(f"    - GrabFood ({len(g_ids)})   = {', '.join(g_ids)}")
                if gf_ids:
                    print(f"    - GoFood ({len(gf_ids)})     = {', '.join(gf_ids)}")
            else:
                print(f"  Store ID  : {BOLD}{selected_outlet['store_id']}{RESET}")
                
            print(f"  {CYAN}{'─'*60}{RESET}")
            print()
            print(f"  {BOLD}Konfirmasi tindakan:{RESET}")
            print(f"    {GREEN}[1]{RESET} Lanjutkan Tarik Menu")
            print(f"    {YELLOW}[2]{RESET} Kembali ke daftar outlet")
            print(f"    {RED}[3]{RESET} Batal dan Keluar")
            print()
            
            choice = input(f"  {BOLD}Pilihan (1/2/3):{RESET} ").strip()
            if choice == "1":
                break
            elif choice == "2":
                state = "select_outlet"
            elif choice == "3":
                print("  Dibatalkan.")
                sys.exit(0)
            else:
                print(f"  {RED}Pilihan tidak valid.{RESET}")
                time.sleep(1)
                
        elif state == "confirm_all":
            os.system('cls' if os.name == 'nt' else 'clear')
            banner()
            
            # Count unprocessed outlets
            unprocessed_count = 0
            for o in selected_outlet:
                if not check_outlet_processed(applicator, o):
                    unprocessed_count += 1
            
            print(f"  {CYAN}{'─'*60}{RESET}")
            print(f"  Aplikator : {BOLD}{applicator.upper()}{RESET}")
            print(f"  Mode      : {BOLD}{YELLOW}BATCH RUN (Massal){RESET}")
            print(f"  Total     : {BOLD}{len(selected_outlet)} outlet/cabang{RESET}")
            print(f"  Belum Run : {BOLD}{GREEN}{unprocessed_count} outlet/cabang{RESET}")
            print(f"  Sudah Run : {BOLD}{len(selected_outlet) - unprocessed_count} outlet/cabang (Skipped jika pilih [2]){RESET}")
            print(f"  Jeda      : {BOLD}Setiap 10 outlet akan dijeda 1 menit{RESET}")
            print(f"  {CYAN}{'─'*60}{RESET}")
            print()
            print(f"  {BOLD}Konfirmasi tindakan:{RESET}")
            print(f"    {GREEN}[1]{RESET} Lanjutkan Jalankan SEMUA (Overwrite)")
            print(f"    {GREEN}[2]{RESET} Lanjutkan Jalankan HANYA yang Belum Selesai ({unprocessed_count} outlet)")
            print(f"    {YELLOW}[3]{RESET} Kembali ke daftar outlet")
            print(f"    {RED}[4]{RESET} Batal dan Keluar")
            print()
            
            choice = input(f"  {BOLD}Pilihan (1/2/3/4):{RESET} ").strip()
            if choice == "1":
                break
            elif choice == "2":
                filtered_outlets = []
                for o in selected_outlet:
                    if not check_outlet_processed(applicator, o):
                        filtered_outlets.append(o)
                
                if not filtered_outlets:
                    print(f"\n  {GREEN}Semua outlet dalam batch ini sudah berhasil ditarik sebelumnya!{RESET}")
                    time.sleep(3)
                    state = "select_outlet"
                else:
                    selected_outlet = filtered_outlets
                    break
            elif choice == "3":
                state = "select_outlet"
            elif choice == "4":
                print("  Dibatalkan.")
                sys.exit(0)
            else:
                print(f"  {RED}Pilihan tidak valid.{RESET}")
                time.sleep(1)
                
    return applicator, selected_outlet

def main():
    try:
        applicator, outlet = interactive_menu()
    except KeyboardInterrupt:
        print("\n  Dibatalkan oleh pengguna.")
        sys.exit(0)
        
    import re
    
    outlets_to_process = outlet if isinstance(outlet, list) else [outlet]
    total_outlets = len(outlets_to_process)
    
    print(f"\n{CYAN}=== MEMULAI PENARIKAN MENU ({total_outlets} OUTLET) ==={RESET}")
    print(f"[*] Waktu mulai: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    success_count = 0
    fail_count = 0
    
    all_shopee = []
    all_grab = []
    all_gofood = []
    if applicator == "all":
        all_shopee = get_outlets_for_applicator("shopee")
        all_grab = get_outlets_for_applicator("grab")
        all_gofood = get_outlets_for_applicator("gofood")
        
    for idx, o in enumerate(outlets_to_process):
        raw_outlet = o.get('nama_outlet') or o.get('nama_resto_final') or o.get('merchant_name') or 'unknown'
        clean_outlet = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
        clean_outlet = re.sub(r'\s+', ' ', clean_outlet).lower()
        
        name_to_show = o['brand'] or o['nama_resto_final'] or o['nama_outlet']
        print(f"\n{BOLD}[{idx + 1}/{total_outlets}] Memproses: {name_to_show}{RESET}")
        
        if applicator == "all":
            excel_paths = []
            
            def find_all_o(olist):
                target_outlet = str(o.get('nama_outlet') or '').strip().lower()
                
                results = []
                for ro in olist:
                    ro_outlet = str(ro.get('nama_outlet') or '').strip().lower()
                    
                    if ro_outlet == target_outlet and ro_outlet != "":
                        results.append(ro)
                return results
                
            o_s_list = find_all_o(all_shopee)
            o_g_list = find_all_o(all_grab)
            o_gf_list = find_all_o(all_gofood)
            
            if o_gf_list:
                for idx_item, o_gf in enumerate(o_gf_list):
                    output_dir_gf = os.path.join(EXPORTS_DIR, "gofood", clean_outlet, f"gofood_{idx_item}")
                    os.makedirs(output_dir_gf, exist_ok=True)
                    for attempt in range(3):
                        if attempt > 0: print(f"  [*] Mengulang GoFood (ID: {o_gf.get('store_id', '-')}) - Percobaan ke-{attempt+1}...")
                        else: print(f"  [*] Menjalankan GoFood (ID: {o_gf.get('store_id', '-')})...")
                        try:
                            s, r = extract_gofood_menu(o_gf, output_dir_gf)
                            if s and isinstance(r, dict): 
                                excel_paths.append(r['excel'])
                                break
                            elif attempt == 2:
                                print(f"  {RED}Gagal GoFood setelah 3x percobaan: {r}{RESET}")
                        except Exception as e:
                            print(f"  {RED}Error GoFood: {e}{RESET}")
                        
            if o_g_list:
                for idx_item, o_g in enumerate(o_g_list):
                    output_dir_g = os.path.join(EXPORTS_DIR, "grab", clean_outlet, f"grab_{idx_item}")
                    os.makedirs(output_dir_g, exist_ok=True)
                    for attempt in range(3):
                        if attempt > 0: print(f"  [*] Mengulang Grab (ID: {o_g.get('store_id', '-')}) - Percobaan ke-{attempt+1}...")
                        else: print(f"  [*] Menjalankan Grab (ID: {o_g.get('store_id', '-')})...")
                        try:
                            s, r = extract_grab_menu(o_g, output_dir_g)
                            if s and isinstance(r, dict): 
                                excel_paths.append(r['excel'])
                                break
                            elif attempt == 2:
                                print(f"  {RED}Gagal Grab setelah 3x percobaan: {r}{RESET}")
                        except Exception as e:
                            print(f"  {RED}Error Grab: {e}{RESET}")
                        
            if o_s_list:
                for idx_item, o_s in enumerate(o_s_list):
                    output_dir_s = os.path.join(EXPORTS_DIR, "shopee", clean_outlet, f"shopee_{idx_item}")
                    os.makedirs(output_dir_s, exist_ok=True)
                    for attempt in range(3):
                        if attempt > 0: print(f"  [*] Mengulang Shopee (ID: {o_s.get('store_id', '-')}) - Percobaan ke-{attempt+1}...")
                        else: print(f"  [*] Menjalankan Shopee (ID: {o_s.get('store_id', '-')})...")
                        try:
                            s, r = extract_shopee_menu(o_s, output_dir_s)
                            if s and isinstance(r, dict): 
                                excel_paths.append(r['excel'])
                                break
                            elif attempt == 2:
                                print(f"  {RED}Gagal Shopee setelah 3x percobaan: {r}{RESET}")
                        except Exception as e:
                            print(f"  {RED}Error Shopee: {e}{RESET}")
                
            if excel_paths:
                combined_dir = os.path.join(EXPORTS_DIR, "combined", clean_outlet)
                os.makedirs(combined_dir, exist_ok=True)
                
                clean_outlet_filename = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
                excel_filename = f"O.C5 {clean_outlet_filename}.xlsx"
                    
                combined_path = os.path.join(combined_dir, excel_filename)
                print(f"  [*] Menggabungkan {len(excel_paths)} file C5 ke {combined_path}...")
                if combine_c5(excel_paths, combined_path):
                    success_count += 1
                    print(f"  {GREEN}✔ Berhasil menggabungkan semua platform!{RESET}")
                    
                    # Upload ke Google Drive
                    upload_combined_to_drive(combined_path, clean_outlet_filename)
                else:
                    fail_count += 1
                    print(f"  {RED}✘ Gagal menggabungkan C5.{RESET}")
            else:
                fail_count += 1
                print(f"  {RED}✘ Tidak ada platform yang berhasil ditarik.{RESET}")

        else:
            output_dir = os.path.join(EXPORTS_DIR, applicator, clean_outlet)
            os.makedirs(output_dir, exist_ok=True)
            
            success = False
            result_data = None
            
            try:
                for attempt in range(3):
                    if attempt > 0: print(f"  [*] Mengulang {applicator.upper()} (ID: {o.get('store_id', '-')}) - Percobaan ke-{attempt+1}...")
                    else: print(f"  [*] Menjalankan {applicator.upper()} (ID: {o.get('store_id', '-')})...")
                    
                    if applicator == "shopee":
                        success, result_data = extract_shopee_menu(o, output_dir)
                    elif applicator == "grab":
                        success, result_data = extract_grab_menu(o, output_dir)
                    elif applicator == "gofood":
                        success, result_data = extract_gofood_menu(o, output_dir)
                        
                    if success and isinstance(result_data, dict):
                        break
                    elif attempt == 2:
                        print(f"  {RED}Gagal {applicator.upper()} setelah 3x percobaan: {result_data}{RESET}")
            except Exception as e:
                success = False
                result_data = f"Exception occurred: {e}"
                
            if success and isinstance(result_data, dict):
                success_count += 1
                print(f"  {GREEN}✔ Berhasil! {result_data.get('items_count', 0)} item, {result_data.get('mods_count', 0)} modifier.{RESET}")
                print(f"  Hasil disimpan di: {output_dir}")
            else:
                fail_count += 1
                print(f"  {RED}✘ Gagal: {result_data}{RESET}")
                
        if (idx + 1) < total_outlets and (idx + 1) % 10 == 0:
            print(f"\n{YELLOW}[BATCH] Selesai memproses 10 outlet. Menunggu jeda 1 menit sebelum batch berikutnya...{RESET}")
            for remaining in range(60, 0, -1):
                sys.stdout.write(f"\rMenunggu... {remaining} detik")
                sys.stdout.flush()
                time.sleep(1)
            print(f"\r{GREEN}[BATCH] Jeda selesai. Melanjutkan penarikan...{RESET}\n")
            
    print(f"\n{CYAN}=== PENARIKAN MENU SELESAI ==={RESET}")
    print(f"  - Sukses : {GREEN}{success_count}{RESET}")
    print(f"  - Gagal  : {RED}{fail_count}{RESET}")
            
if __name__ == "__main__":
    main()
