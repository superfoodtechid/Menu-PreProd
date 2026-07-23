import os
import sys
import uuid
import logging
import threading
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, Depends, HTTPException, BackgroundTasks, Query, status
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

# Setup Dynamic Paths to ensure reliable server deployment
BASE_DIR = Path(__file__).resolve().parent
sys.path.append(str(BASE_DIR))
sys.path.append(str(BASE_DIR / "menu_core"))

from menu_core.database import get_db, init_db, Account, Outlet, Job, AuditTrail

# Setup Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("FoodMasterAPI")

# Initialize database tables on startup
app = FastAPI(
    title="FoodMaster Menu Portal API",
    description="Backend API for managing multi-platform menus (Shopee, Grab, GoFood)",
    version="1.0.0"
)

# Enable CORS for Next.js frontend (Vercel)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # In production, restrict this to the Vercel domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def startup_event():
    logger.info("🚀 Initializing database tables...")
    init_db()
    # Ensure export directories exist dynamically
    exports_dir = BASE_DIR / "data" / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    logger.info(f"📂 Exports directory verified at: {exports_dir}")


# ─── PYDANTIC SCHEMAS ─────────────────────────────────────────────────────────

class AccountCreate(BaseModel):
    platform: str = Field(..., description="shopee | grab | gofood")
    username: str
    password: str
    portal: Optional[str] = None

class AccountResponse(BaseModel):
    id: uuid.UUID
    platform: str
    username: str
    portal: Optional[str]
    created_at: datetime

    class Config:
        from_attributes = True

class OutletCreate(BaseModel):
    account_id: uuid.UUID
    store_id: Optional[str] = None
    merchant_name: str = Field(..., description="Nama merchant / portal selector di web")
    nama_outlet: Optional[str] = None
    cabang: Optional[str] = None
    nama_resto_final: Optional[str] = None
    brand: Optional[str] = None
    is_active: bool = True

class OutletResponse(BaseModel):
    id: uuid.UUID
    account_id: uuid.UUID
    store_id: Optional[str]
    merchant_name: str
    nama_outlet: Optional[str]
    cabang: Optional[str]
    nama_resto_final: Optional[str]
    brand: Optional[str]
    is_active: bool
    last_sync_at: Optional[datetime]
    created_at: datetime
    platform: Optional[str] = None

    class Config:
        from_attributes = True

class JobResponse(BaseModel):
    id: uuid.UUID
    outlet_id: Optional[uuid.UUID]
    job_type: str
    platform: str
    status: str
    progress_pct: int
    current_step: Optional[str]
    payload: Optional[dict]
    result_metadata: Optional[dict]
    error_message: Optional[str]
    created_by: str
    created_at: datetime
    started_at: Optional[datetime]
    completed_at: Optional[datetime]

    class Config:
        from_attributes = True

class AuditTrailResponse(BaseModel):
    id: uuid.UUID
    job_id: uuid.UUID
    outlet_id: uuid.UUID
    item_id: str
    item_name: str
    change_type: str
    field_changed: str
    old_value: Optional[str]
    new_value: str
    status: str
    error_message: Optional[str]
    created_at: datetime

    class Config:
        from_attributes = True


class PriceUpdateItem(BaseModel):
    item_id: str
    category_id: str
    new_price: float

class PriceUpdateRequest(BaseModel):
    outlet_id: uuid.UUID
    updates: List[PriceUpdateItem]


# ─── GSHEETS SYNC ENDPOINT ───────────────────────────────────────────────────

import io
import requests
import pandas as pd

@app.post("/api/sync-sheets", status_code=status.HTTP_200_OK)
def sync_sheets(db: Session = Depends(get_db)):
    url = f"https://docs.google.com/spreadsheets/d/e/2PACX-1vQ3tLKBNXDqRgBw0mNhKZFxgvKx-JoiTDzm_s5Ix1cm7O6HCv4IvExOLR2HSRVaXSsx82V348mcr9X4/pub?gid=0&single=true&output=csv&t={int(datetime.utcnow().timestamp())}"
    try:
        logger.info("⏳ Downloading latest merchant sheet for database sync...")
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        df = pd.read_csv(io.StringIO(resp.text))
    except Exception as e:
        logger.error(f"❌ Failed to fetch Google Sheet: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch Google Sheet: {str(e)}")

    added_accounts = 0
    added_outlets = 0
    updated_outlets = 0

    # Filter only Live status
    df_live = df[df["Status"].str.contains("Live", na=False, case=False)]

    for _, row in df_live.iterrows():
        app_val = str(row.get("Aplikasi", "")).strip().lower()
        if app_val == "shopeefood":
            platform = "shopee"
        elif app_val == "grabfood":
            platform = "grab"
        elif app_val == "gofood":
            platform = "gofood"
        else:
            continue

        # Extract Username & Password based on platform logic
        username = None
        password = None

        if platform == "shopee":
            # Username is allvbadmin
            username = "allvbadmin"
            # Get Password.1 or Kata Sandi.1
            pwd_col = "Kata Sandi.1" if "Kata Sandi.1" in df.columns else "Kata Sandi"
            password = str(row.get(pwd_col, "Master!00!")).strip()
        elif platform == "grab":
            user_col_sf = "Nama Pengguna.1"
            user_col_mt = "Nama Pengguna"
            pwd_col_sf = "Kata Sandi.1"
            pwd_col_mt = "Kata Sandi"

            user_val = row.get(user_col_sf) if pd.notna(row.get(user_col_sf)) and str(row.get(user_col_sf)).strip() != "-" else row.get(user_col_mt)
            pwd_val = row.get(pwd_col_sf) if pd.notna(row.get(pwd_col_sf)) and str(row.get(pwd_col_sf)).strip() != "-" else row.get(pwd_col_mt)

            if pd.notna(user_val) and str(user_val).strip() != "":
                username = str(user_val).strip()
            if pd.notna(pwd_val) and str(pwd_val).strip() != "":
                password = str(pwd_val).strip()
        elif platform == "gofood":
            # GoFood uses Email Login Go 1 or Email Login Go 2 (No phone login)
            email_1 = row.get("Email Login Go 1")
            email_2 = row.get("Email Login Go 2")
            user_val = email_1 if pd.notna(email_1) and "@" in str(email_1) else email_2

            if pd.notna(user_val) and "@" in str(user_val):
                username = str(user_val).strip()
            else:
                # Fallback to username.1 if email not found
                user_1 = row.get("Nama Pengguna.1")
                if pd.notna(user_1) and str(user_1).strip() != "" and str(user_1).strip() != "-":
                    username = str(user_1).strip()

            pwd_val = row.get("Kata Sandi.1") if pd.notna(row.get("Kata Sandi.1")) else row.get("Kata Sandi")
            if pd.notna(pwd_val) and str(pwd_val).strip() != "" and str(pwd_val).strip() != "-":
                password = str(pwd_val).strip()

        if not username:
            continue
        if not password:
            password = "Master@123" # Default fallback password

        # 1. Upsert Account
        db_account = db.query(Account).filter(Account.username == username, Account.platform == platform).first()
        if not db_account:
            db_account = Account(
                platform=platform,
                username=username,
                password=password,
                portal="shopee_partner" if platform == "shopee" else "merchant_portal"
            )
            db.add(db_account)
            db.commit()
            db.refresh(db_account)
            added_accounts += 1
        else:
            if db_account.password != password:
                db_account.password = password
                db.commit()

        # 2. Extract Outlet Info
        store_id_raw = row.get("Store ID")
        store_id = str(store_id_raw).strip().split(".")[0] if pd.notna(store_id_raw) and str(store_id_raw).strip() != "-" else None
        
        m_name_raw = row.get("Merchant Name")
        merchant_name = str(m_name_raw).strip() if pd.notna(m_name_raw) and str(m_name_raw).strip() != "-" else str(row.get("Nama Outlet", "")).strip()

        nama_outlet = str(row.get("Nama Outlet", "")).strip() if pd.notna(row.get("Nama Outlet")) else None
        cabang = str(row.get("Cabang", "")).strip() if pd.notna(row.get("Cabang")) else str(row.get("Brand", "")).strip()
        nama_resto_final = str(row.get("Nama Resto Final", "")).strip() if pd.notna(row.get("Nama Resto Final")) else None
        brand = str(row.get("Brand", "")).strip() if pd.notna(row.get("Brand")) else None

        # 3. Upsert Outlet
        db_outlet = None
        if store_id:
            db_outlet = db.query(Outlet).filter(Outlet.store_id == store_id).first()

        if not db_outlet:
            # Fallback query if store_id was not provided
            db_outlet = db.query(Outlet).filter(
                Outlet.account_id == db_account.id,
                Outlet.merchant_name == merchant_name,
                Outlet.nama_outlet == nama_outlet,
                Outlet.cabang == cabang
            ).first()

        if not db_outlet:
            db_outlet = Outlet(
                account_id=db_account.id,
                store_id=store_id,
                merchant_name=merchant_name,
                nama_outlet=nama_outlet,
                cabang=cabang,
                nama_resto_final=nama_resto_final,
                brand=brand,
                is_active=True
            )
            db.add(db_outlet)
            added_outlets += 1
        else:
            db_outlet.store_id = store_id
            db_outlet.nama_resto_final = nama_resto_final
            db_outlet.brand = brand
            db_outlet.is_active = True
            updated_outlets += 1

    db.commit()
    logger.info(f"📊 Sync Sheet Complete. Added Accounts: {added_accounts}, Added Outlets: {added_outlets}, Updated Outlets: {updated_outlets}")
    return {
        "status": "success",
        "added_accounts": added_accounts,
        "added_outlets": added_outlets,
        "updated_outlets": updated_outlets
    }


# ─── ACCOUNTS ENDPOINTS ───────────────────────────────────────────────────────

@app.post("/api/accounts", response_model=AccountResponse, status_code=status.HTTP_201_CREATED)
def create_account(account: AccountCreate, db: Session = Depends(get_db)):
    db_account = db.query(Account).filter(
        Account.username == account.username, 
        Account.platform == account.platform
    ).first()
    if db_account:
        raise HTTPException(status_code=400, detail="Account already exists for this platform")
    
    new_account = Account(
        platform=account.platform,
        username=account.username,
        password=account.password,
        portal=account.portal
    )
    db.add(new_account)
    db.commit()
    db.refresh(new_account)
    return new_account

@app.get("/api/accounts", response_model=List[AccountResponse])
def list_accounts(db: Session = Depends(get_db)):
    return db.query(Account).all()


# ─── OUTLETS ENDPOINTS ────────────────────────────────────────────────────────

SUPPORTED_PLATFORMS = {"shopee", "grab", "gofood"}


def normalize_platform_filters(platforms: Optional[List[str]]) -> List[str]:
    normalized = list(dict.fromkeys(
        value.strip().lower()
        for value in (platforms or [])
        if value and value.strip()
    ))
    unsupported = sorted(set(normalized) - SUPPORTED_PLATFORMS)
    if unsupported:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Unsupported platform filter",
                "unsupported": unsupported,
                "supported": sorted(SUPPORTED_PLATFORMS),
            },
        )
    return normalized

@app.post("/api/outlets", response_model=OutletResponse, status_code=status.HTTP_201_CREATED)
def create_outlet(outlet: OutletCreate, db: Session = Depends(get_db)):
    # Verify account exists
    db_account = db.query(Account).filter(Account.id == outlet.account_id).first()
    if not db_account:
        raise HTTPException(status_code=404, detail="Parent account not found")
    
    if outlet.store_id:
        db_outlet = db.query(Outlet).filter(Outlet.store_id == outlet.store_id).first()
        if db_outlet:
            raise HTTPException(status_code=400, detail="Outlet with this store_id already exists")

    new_outlet = Outlet(
        account_id=outlet.account_id,
        store_id=outlet.store_id,
        merchant_name=outlet.merchant_name,
        nama_outlet=outlet.nama_outlet,
        cabang=outlet.cabang,
        nama_resto_final=outlet.nama_resto_final,
        brand=outlet.brand,
        is_active=outlet.is_active
    )
    db.add(new_outlet)
    db.commit()
    db.refresh(new_outlet)
    return new_outlet

@app.get("/api/outlets", response_model=List[OutletResponse])
def list_outlets(
    platform: Optional[List[str]] = Query(
        default=None,
        description="Filter platform berulang, contoh: ?platform=grab&platform=gofood",
    ),
    db: Session = Depends(get_db),
):
    platforms = normalize_platform_filters(platform)
    query = db.query(Outlet)
    if platforms:
        query = query.join(Outlet.account).filter(Account.platform.in_(platforms))
    return query.all()


# ─── BACKGROUND JOBS WORKER ───────────────────────────────────────────────────

# Platform-specific locks to allow parallel execution between different platforms,
# but enforce sequential execution within each platform (especially Shopee).
PLATFORM_LOCKS = {
    "shopee": threading.Lock(),
    "grab": threading.Lock(),
    "gofood": threading.Lock()
}

def run_pull_job(job_id: uuid.UUID, outlet_id: uuid.UUID):
    # Setup job-specific session context
    from menu_core.database import SessionLocal
    db = SessionLocal()
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        db.close()
        return

    platform = (job.platform or "").lower()
    lock = PLATFORM_LOCKS.get(platform)
    if lock:
        logger.info(f"🔒 Job {job_id} ({platform}) waiting for lock...")
        lock.acquire()
        logger.info(f"🔓 Job {job_id} ({platform}) acquired lock. Starting execution.")

    try:
        # Re-fetch job under lock to ensure we have the latest database state
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            return

        job.status = "RUNNING"
        job.started_at = datetime.utcnow()
        job.progress_pct = 10
        job.current_step = "Memuat kredensial dan inisialisasi browser..."
        db.commit()

        outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
        account = db.query(Account).filter(Account.id == outlet.account_id).first()
        
        # Determine paths dynamically
        import re
        raw_outlet = outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown'
        clean_outlet = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
        clean_outlet = re.sub(r'\s+', ' ', clean_outlet).lower()
        
        exports_dir = BASE_DIR / "data" / "exports" / job.platform / clean_outlet
        exports_dir.mkdir(parents=True, exist_ok=True)

        # Trigger Applicator specific script
        if job.platform == "shopee":
            job.progress_pct = 30
            job.current_step = "Membuka browser dan login portal Shopee Partner..."
            db.commit()
            
            # Setup store_metadata payload for shopee.core.pull
            store_metadata = {
                "store_id": outlet.store_id,
                "merchant_name": outlet.merchant_name,
                "nama_outlet": outlet.nama_outlet,
                "cabang": outlet.cabang,
                "nama_resto_final": outlet.nama_resto_final,
                "brand": outlet.brand,
                "username": account.username,
                "password": account.password,
                "portal": account.portal
            }
            
            # Add project root to sys.path to resolve shopee.* absolute imports correctly
            if str(BASE_DIR) not in sys.path:
                sys.path.insert(0, str(BASE_DIR))
                
            from shopee.core.pull import extract_shopee_menu
            
            # Run shopee extraction
            success, result = extract_shopee_menu(store_metadata, str(exports_dir))
            
            if not success:
                raise Exception(f"Shopee extraction failed: {result}")
                
            # If store_id was dynamically resolved and wasn't set in DB, update it!
            resolved_store_id = store_metadata.get("store_id")
            if resolved_store_id and not outlet.store_id:
                # Check for uniqueness before updating to prevent constraints failure
                existing_outlet = db.query(Outlet).filter(Outlet.store_id == resolved_store_id).first()
                if not existing_outlet:
                    outlet.store_id = resolved_store_id
                    logger.info(f"💾 Dynamically updated store_id to {resolved_store_id} for outlet {outlet.merchant_name}")
            
            # Upload to Google Drive/Sheets
            excel_path = result.get("excel")
            drive_url = None
            if excel_path and os.path.exists(excel_path):
                try:
                    job.current_step = "Mengunggah hasil ke Google Sheets..."
                    db.commit()
                    from upload_drive import upload_combined_to_drive
                    clean_outlet_filename = "".join(c for c in (outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown') if c.isalnum() or c in (' ', '_', '-')).strip()
                    drive_url = upload_combined_to_drive(excel_path, clean_outlet_filename)
                except Exception as ue:
                    logger.error(f"Failed to upload to Google Drive: {ue}")

            job.status = "SUCCESS"
            job.progress_pct = 100
            job.current_step = "Penarikan menu selesai!"
            job.result_metadata = {
                "excel_path": result.get("excel"),
                "items_csv_path": result.get("items_csv"),
                "mods_csv_path": result.get("mods_csv"),
                "items_count": result.get("items_count", 0),
                "mods_count": result.get("mods_count", 0),
                "completed_at": datetime.utcnow().isoformat(),
                "gspread_url": drive_url
            }
            job.completed_at = datetime.utcnow()
            outlet.last_sync_at = datetime.utcnow()
            db.commit()
            
        elif job.platform == "gofood":
            job.progress_pct = 30
            job.current_step = "Menyiapkan parameter penarikan GoFood..."
            db.commit()
            
            store_metadata = {
                "store_id": outlet.store_id,
                "merchant_name": outlet.merchant_name,
                "nama_outlet": outlet.nama_outlet,
                "cabang": outlet.cabang,
                "nama_resto_final": outlet.nama_resto_final,
                "brand": outlet.brand,
                "username": account.username,
                "password": account.password
            }
            
            job.progress_pct = 50
            job.current_step = "Meluncurkan browser GoFood & memproses penarikan..."
            db.commit()
            
            from menu_core.gofood import extract_gofood_menu
            success, result = extract_gofood_menu(store_metadata, str(exports_dir))
            
            if not success:
                raise Exception(f"GoFood extraction failed: {result}")
                
            # Upload to Google Drive/Sheets
            excel_path = result.get("excel")
            drive_url = None
            if excel_path and os.path.exists(excel_path):
                try:
                    job.current_step = "Mengunggah hasil ke Google Sheets..."
                    db.commit()
                    from upload_drive import upload_combined_to_drive
                    clean_outlet_filename = "".join(c for c in (outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown') if c.isalnum() or c in (' ', '_', '-')).strip()
                    drive_url = upload_combined_to_drive(excel_path, clean_outlet_filename)
                except Exception as ue:
                    logger.error(f"Failed to upload to Google Drive: {ue}")

            job.status = "SUCCESS"
            job.progress_pct = 100
            job.current_step = "Penarikan menu GoFood selesai!"
            job.result_metadata = {
                "excel_path": result.get("excel"),
                "items_count": result.get("items_count", 0),
                "mods_count": result.get("mods_count", 0),
                "completed_at": datetime.utcnow().isoformat(),
                "gspread_url": drive_url
            }
            job.completed_at = datetime.utcnow()
            outlet.last_sync_at = datetime.utcnow()
            db.commit()
            
        elif job.platform == "grab":
            job.progress_pct = 30
            job.current_step = "Menyiapkan parameter penarikan GrabFood..."
            db.commit()
            
            store_metadata = {
                "store_id": outlet.store_id,
                "merchant_name": outlet.merchant_name,
                "nama_outlet": outlet.nama_outlet,
                "cabang": outlet.cabang,
                "nama_resto_final": outlet.nama_resto_final,
                "brand": outlet.brand,
                "username": account.username,
                "password": account.password
            }
            
            job.progress_pct = 50
            job.current_step = "Meluncurkan browser GrabFood & memproses penarikan..."
            db.commit()
            
            from menu_core.grab import extract_grab_menu
            success, result = extract_grab_menu(store_metadata, str(exports_dir))
            
            if not success:
                raise Exception(f"Grab extraction failed: {result}")
                
            # Upload to Google Drive/Sheets
            excel_path = result.get("excel")
            drive_url = None
            if excel_path and os.path.exists(excel_path):
                try:
                    job.current_step = "Mengunggah hasil ke Google Sheets..."
                    db.commit()
                    from upload_drive import upload_combined_to_drive
                    clean_outlet_filename = "".join(c for c in (outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown') if c.isalnum() or c in (' ', '_', '-')).strip()
                    drive_url = upload_combined_to_drive(excel_path, clean_outlet_filename)
                except Exception as ue:
                    logger.error(f"Failed to upload to Google Drive: {ue}")

            job.status = "SUCCESS"
            job.progress_pct = 100
            job.current_step = "Penarikan menu GrabFood selesai!"
            job.result_metadata = {
                "excel_path": result.get("excel"),
                "items_count": result.get("items_count", 0),
                "mods_count": result.get("mods_count", 0),
                "completed_at": datetime.utcnow().isoformat(),
                "gspread_url": drive_url
            }
            job.completed_at = datetime.utcnow()
            outlet.last_sync_at = datetime.utcnow()
            db.commit()

        logger.info(f"✅ Job {job_id} completed successfully.")

    except Exception as e:
        logger.error(f"❌ Job {job_id} failed: {e}")
        job.status = "FAILED"
        job.error_message = str(e)
        job.current_step = f"Terjadi kesalahan: {str(e)}"
        job.completed_at = datetime.utcnow()
        db.commit()
    finally:
        if lock:
            try:
                lock.release()
                logger.info(f"🔓 Job {job_id} ({platform}) lock released.")
            except Exception as le:
                logger.warning(f"⚠️ Failed to release lock: {le}")
        db.close()


def run_push_price_job(job_id: uuid.UUID, outlet_id: uuid.UUID, updates_list: list):
    """Background task to push price changes to GoFood, GrabFood, or ShopeeFood."""
    import asyncio
    from menu_core.database import SessionLocal
    db = SessionLocal()
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        db.close()
        return

    platform = (job.platform or "").lower()
    lock = PLATFORM_LOCKS.get(platform)
    if lock:
        logger.info(f"🔒 Job {job_id} ({platform}) waiting for lock...")
        lock.acquire()
        logger.info(f"🔓 Job {job_id} ({platform}) acquired lock. Starting execution.")

    try:
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            return

        job.status = "RUNNING"
        job.started_at = datetime.utcnow()
        job.progress_pct = 10
        job.current_step = "Menginisialisasi kredensial..."
        db.commit()

        outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
        account = db.query(Account).filter(Account.id == outlet.account_id).first()

        total_updates = len(updates_list)
        success_count = 0
        fail_count = 0

        if platform == "shopee":
            # Add project root to sys.path
            if str(BASE_DIR) not in sys.path:
                sys.path.insert(0, str(BASE_DIR))
            from shopee.core.edit import edit_dish_via_portal

            store_metadata = {
                "store_id": outlet.store_id,
                "username": account.username,
                "password": account.password,
                "merchant_name": outlet.merchant_name,
                "nama_resto_final": outlet.nama_resto_final,
                "nama_outlet": outlet.nama_outlet
            }

            job.progress_pct = 30
            job.current_step = "Membuka browser Shopee..."
            db.commit()

            for idx, update in enumerate(updates_list):
                item_id = update["item_id"]
                new_price = update["new_price"]
                
                job.current_step = f"Memproses update harga item {item_id} ke Rp {new_price}..."
                job.progress_pct = int(30 + (idx / total_updates) * 60)
                db.commit()

                try:
                    success, msg = edit_dish_via_portal(store_metadata, dish_id=item_id, price=new_price, headless=True)
                    if success:
                        success_count += 1
                        status_str = "SUCCESS"
                        err_msg = None
                    else:
                        fail_count += 1
                        status_str = "FAILED"
                        err_msg = msg
                except Exception as ex:
                    fail_count += 1
                    status_str = "FAILED"
                    err_msg = str(ex)

                trail = AuditTrail(
                    job_id=job.id,
                    outlet_id=outlet.id,
                    item_id=item_id,
                    item_name=item_id,
                    change_type="PRICE_UPDATE",
                    field_changed="price",
                    old_value=None,
                    new_value=str(new_price),
                    status=status_str,
                    error_message=err_msg
                )
                db.add(trail)
                db.commit()

        elif platform == "grab":
            from playwright.sync_api import sync_playwright
            from grab.core.grab_api_scraper import GrabAPI, perform_login, SESSION_DIR

            username = account.username
            password = account.password
            store_id = outlet.store_id

            job.progress_pct = 30
            job.current_step = "Meluncurkan browser Grab..."
            db.commit()

            with sync_playwright() as p:
                session_path = os.path.join(SESSION_DIR, f"{username}.json")
                storage_state = session_path if os.path.exists(session_path) else None
                
                browser = p.chromium.launch(headless=True)
                context = browser.new_context(
                    storage_state=storage_state,
                    user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
                )
                page = context.new_page()

                try:
                    page.goto("https://merchant.grab.com/dashboard", wait_until="domcontentloaded", timeout=30000)
                except Exception as e:
                    logger.warning(f"Grab dashboard navigate warning: {e}")

                async def grab_async_flow():
                    nonlocal success_count, fail_count
                    api = GrabAPI(page, username, password)
                    mgid = await api.get_merchant_group_id()
                    if not mgid:
                        if await perform_login(page, username, password):
                            mgid = await api.get_merchant_group_id()
                            if mgid:
                                await context.storage_state(path=session_path)
                            else:
                                return "Failed to retrieve Grab merchant group ID after login."
                        else:
                            return "Grab login failed."

                    menu_data, err = await api.fetch_menu(mgid, store_id, is_menu_group=False)
                    if err or not menu_data:
                        return f"Failed to fetch Grab menu: {err}"

                    grab_items_by_id = {}
                    for cat in menu_data.get("categories", []):
                        items_list = cat.get("items") or cat.get("menuItems") or []
                        selling_time_id = cat.get("sellingTimeID")
                        for item in items_list:
                            grab_items_by_id[str(item.get("itemID"))] = {
                                "item": item,
                                "category_id": cat.get("categoryID"),
                                "sellingTimeID": selling_time_id
                            }

                    for idx, update in enumerate(updates_list):
                        item_id = update["item_id"]
                        new_price = update["new_price"]

                        item_info = grab_items_by_id.get(item_id)
                        if not item_info:
                            fail_count += 1
                            trail = AuditTrail(
                                job_id=job.id,
                                outlet_id=outlet.id,
                                item_id=item_id,
                                item_name=item_id,
                                change_type="PRICE_UPDATE",
                                field_changed="price",
                                old_value=None,
                                new_value=str(new_price),
                                status="FAILED",
                                error_message="Item ID not found in current Grab menu."
                            )
                            db.add(trail)
                            db.commit()
                            continue

                        orig_item = item_info["item"]
                        category_id = item_info["category_id"]
                        selling_time_id = item_info["sellingTimeID"]
                        
                        item_data = {
                            "itemID": item_id,
                            "itemName": orig_item.get("itemName"),
                            "description": orig_item.get("description", ""),
                            "priceInMin": int(new_price * 100),
                            "availableStatus": orig_item.get("availableStatus", 1),
                            "sellingTimeID": selling_time_id,
                            "advancedPricing": orig_item.get("advancedPricing") or {},
                            "purchasability": orig_item.get("purchasability") or {},
                            "imageURL": orig_item.get("imageURL") or "",
                            "imageURLs": orig_item.get("imageURLs") or [],
                            "weight": orig_item.get("weight"),
                            "itemAttributeValues": orig_item.get("itemAttributeValues") or []
                        }

                        val_ok, val_err = await api.validate_item(mgid, store_id, category_id, item_data)
                        if val_err:
                            logger.warning(f"Grab validation warning for item {item_id}: {val_err}")

                        upsert_res, upsert_err = await api.upsert_item(mgid, store_id, category_id, item_data)
                        if upsert_res and not upsert_err:
                            success_count += 1
                            status_str = "SUCCESS"
                            err_msg = None
                        else:
                            fail_count += 1
                            status_str = "FAILED"
                            err_msg = upsert_err or "Unknown Grab API error."

                        trail = AuditTrail(
                            job_id=job.id,
                            outlet_id=outlet.id,
                            item_id=item_id,
                            item_name=orig_item.get("itemName", item_id),
                            change_type="PRICE_UPDATE",
                            field_changed="price",
                            old_value=str(float(orig_item.get("priceInMin", 0)) / 100.0),
                            new_value=str(new_price),
                            status=status_str,
                            error_message=err_msg
                        )
                        db.add(trail)
                        db.commit()

                    return None

                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                err_msg = loop.run_until_complete(grab_async_flow())
                loop.close()
                browser.close()

                if err_msg:
                    raise Exception(err_msg)

        elif platform == "gofood":
            from playwright.sync_api import sync_playwright
            from Gofood.GO.actions import _menu_api as go_api
            from Gofood.GO.updater_gofood import SESSION_DIR as GO_SESSION_DIR

            email = account.username
            password = account.password
            merchant_id = outlet.store_id

            if not merchant_id:
                raise Exception("Merchant ID (store_id) is missing for GoFood outlet.")
            if not merchant_id.startswith("G"):
                merchant_id = "G" + merchant_id

            job.progress_pct = 30
            job.current_step = "Meluncurkan browser GoFood..."
            db.commit()

            with sync_playwright() as p:
                session_path = os.path.join(GO_SESSION_DIR, f"session_{email}.json")
                storage_state = session_path if os.path.exists(session_path) else None

                browser = p.chromium.launch(headless=True)
                context = browser.new_context(
                    storage_state=storage_state,
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                )
                page = context.new_page()

                page.goto("https://portal.gofoodmerchant.co.id/dashboard", wait_until="domcontentloaded")
                time.sleep(3)

                api_headers = {}
                def capture_headers(request):
                    if "api.gojekapi.com" in request.url or "api.gobiz.co.id" in request.url:
                        h = request.headers
                        if 'authorization' in h:
                            api_headers['authorization'] = h['authorization']
                        if 'x-passkey' in h:
                            api_headers['x-passkey'] = h['x-passkey']
                    if "restaurants/" in request.url and "v1" in request.url:
                        parts = request.url.split("/")
                        for i, part in enumerate(parts):
                            if part == "restaurants" and i + 1 < len(parts):
                                candidate = parts[i + 1].split("?")[0]
                                if len(candidate) == 36:
                                    api_headers['restaurant_uuid'] = candidate
                    if "/v2/menu_groups/" in request.url:
                        parts = request.url.split("/")
                        for i, part in enumerate(parts):
                            if part == "menu_groups" and i + 1 < len(parts):
                                candidate = parts[i + 1].split("?")[0]
                                if len(candidate) == 36:
                                    api_headers['menu_group_id'] = candidate

                page.on("request", capture_headers)

                page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}", wait_until="domcontentloaded")
                time.sleep(3)

                if "/auth" in page.url or "login" in page.url:
                    page.goto("https://portal.gofoodmerchant.co.id/auth/login/email", wait_until="domcontentloaded")
                    time.sleep(2)
                    email_input = page.locator('input[type="email"]')
                    email_input.fill(email)
                    submit_btn = page.locator('button:has-text("Lanjut")')
                    submit_btn.first.click()
                    time.sleep(2)
                    pass_input = page.locator('input[type="password"]')
                    pass_input.fill(password)
                    page.locator('button:has-text("Masuk")').first.click()
                    
                    page.wait_for_url(lambda url: "/auth/login" not in url, timeout=45000)
                    context.storage_state(path=session_path)
                    page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}", wait_until="domcontentloaded")
                    time.sleep(3)

                page.goto(f"https://portal.gofoodmerchant.co.id/gofood/{merchant_id}/menu", wait_until="domcontentloaded")
                time.sleep(5)

                token = api_headers.get('authorization')
                rest_uuid = api_headers.get('restaurant_uuid')
                group_id = api_headers.get('menu_group_id')

                if not token or not rest_uuid:
                    raise Exception("Gagal menangkap Authorization Token atau Restaurant UUID untuk GoFood.")

                menu_data = go_api.fetch_menus(page, token, rest_uuid)
                if not menu_data:
                    raise Exception("Gagal menarik menu GoFood untuk perbandingan harga.")

                categories = go_api.parse_menus(menu_data)
                go_items_by_id = {}
                for cat in categories:
                    for item in cat.get("menu_items") or []:
                        iid = item.get("common_id") or item.get("id")
                        go_items_by_id[str(iid)] = {
                            "item": item,
                            "category_id": cat.get("id"),
                            "category_common_id": cat.get("common_id")
                        }

                for idx, update in enumerate(updates_list):
                    item_id = update["item_id"]
                    new_price = update["new_price"]

                    item_info = go_items_by_id.get(item_id)
                    if not item_info:
                        fail_count += 1
                        trail = AuditTrail(
                            job_id=job.id,
                            outlet_id=outlet.id,
                            item_id=item_id,
                            item_name=item_id,
                            change_type="PRICE_UPDATE",
                            field_changed="price",
                            old_value=None,
                            new_value=str(new_price),
                            status="FAILED",
                            error_message="Item ID tidak ditemukan di menu GoFood."
                        )
                        db.add(trail)
                        db.commit()
                        continue

                    orig_item = item_info["item"]
                    cat_common_id = item_info["category_common_id"] or item_info["category_id"]

                    v2_payload = {
                        "menu_common_id": cat_common_id,
                        "image_url": orig_item.get('image_url', orig_item.get('image', '')),
                        "name": orig_item.get('name'),
                        "description": orig_item.get('description', ''),
                        "price": int(new_price),
                        "active": orig_item.get('is_active', orig_item.get('active', True)),
                        "signature": orig_item.get('signature', False)
                    }

                    patch_group_id = group_id if group_id else cat_common_id
                    res = go_api.update_v2_menu_item(page, token, patch_group_id, item_id, v2_payload)

                    if not res or not res.get('ok'):
                        v1_payload = {
                            "name": orig_item.get('name'),
                            "price": str(int(new_price)),
                            "active": orig_item.get('is_active', orig_item.get('active', True)),
                            "description": orig_item.get('description', ''),
                            "image": orig_item.get('image_url', orig_item.get('image', ''))
                        }
                        v1_item_id = orig_item.get('id')
                        res = go_api.update_item(page, token, rest_uuid, v1_item_id, v1_payload)

                    if res and res.get('ok'):
                        success_count += 1
                        status_str = "SUCCESS"
                        err_msg = None
                    else:
                        fail_count += 1
                        status_str = "FAILED"
                        err_msg = res.get('body') or "GoFood API error."

                    trail = AuditTrail(
                        job_id=job.id,
                        outlet_id=outlet.id,
                        item_id=item_id,
                        item_name=orig_item.get("name", item_id),
                        change_type="PRICE_UPDATE",
                        field_changed="price",
                        old_value=str(orig_item.get("price", 0)),
                        new_value=str(new_price),
                        status=status_str,
                        error_message=err_msg
                    )
                    db.add(trail)
                    db.commit()

                browser.close()

        if fail_count > 0 and success_count == 0:
            job.status = "FAILED"
            job.error_message = f"Pembaruan harga gagal! 0 dari {total_updates} item berhasil diperbarui."
        elif fail_count > 0 and success_count > 0:
            job.status = "PARTIAL_SUCCESS"
            job.error_message = f"Sebagian item gagal diperbarui ({success_count} sukses, {fail_count} gagal)."
        else:
            job.status = "SUCCESS"

        job.progress_pct = 100
        job.current_step = f"Pembaruan harga selesai! {success_count} sukses, {fail_count} gagal."
        job.result_metadata = {
            "success_count": success_count,
            "fail_count": fail_count,
            "completed_at": datetime.utcnow().isoformat()
        }
        job.completed_at = datetime.utcnow()
        db.commit()

    except Exception as e:
        logger.error(f"❌ Job {job_id} failed: {e}")
        job.status = "FAILED"
        job.error_message = str(e)
        job.current_step = f"Terjadi kesalahan: {str(e)}"
        job.completed_at = datetime.utcnow()
        db.commit()

    finally:
        if lock:
            try:
                lock.release()
                logger.info(f"🔓 Job {job_id} ({platform}) lock released.")
            except Exception as le:
                logger.warning(f"⚠️ Failed to release lock: {le}")
        db.close()


@app.post("/api/jobs/push-price", response_model=JobResponse, status_code=status.HTTP_202_ACCEPTED)
def trigger_push_price_job(request: PriceUpdateRequest, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """Triggers a background job to push price changes to the applicator merchant portal."""
    outlet = db.query(Outlet).filter(Outlet.id == request.outlet_id).first()
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
        
    updates_payload = []
    for item in request.updates:
        updates_payload.append({
            "item_id": item.item_id,
            "category_id": item.category_id,
            "new_price": item.new_price
        })

    new_job = Job(
        outlet_id=outlet.id,
        job_type="PUSH_UPDATE",
        platform=outlet.account.platform,
        status="PENDING",
        progress_pct=0,
        current_step="Mengantrekan pembaruan harga...",
        payload={"store_id": outlet.store_id, "merchant_name": outlet.merchant_name, "updates_count": len(updates_payload)}
    )
    db.add(new_job)
    db.commit()
    db.refresh(new_job)

    background_tasks.add_task(run_push_price_job, new_job.id, outlet.id, updates_payload)
    return new_job


# ─── JOBS ENDPOINTS ───────────────────────────────────────────────────────────

@app.post("/api/jobs/pull", response_model=JobResponse, status_code=status.HTTP_202_ACCEPTED)
def trigger_pull_job(outlet_id: uuid.UUID, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
    
    # Create Job log entry
    new_job = Job(
        outlet_id=outlet.id,
        job_type="PULL",
        platform=outlet.account.platform,
        status="PENDING",
        progress_pct=0,
        current_step="Mengantrekan tugas penarikan...",
        payload={"store_id": outlet.store_id, "merchant_name": outlet.merchant_name}
    )
    db.add(new_job)
    db.commit()
    db.refresh(new_job)

    # Dispatch to background executor thread
    background_tasks.add_task(run_pull_job, new_job.id, outlet.id)
    return new_job

@app.get("/api/jobs/{job_id}", response_model=JobResponse)
def get_job_status(job_id: uuid.UUID, db: Session = Depends(get_db)):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job

@app.get("/api/jobs", response_model=List[JobResponse])
def list_jobs(db: Session = Depends(get_db)):
    return db.query(Job).order_by(Job.created_at.desc()).limit(50).all()
@app.get("/api/jobs/download/{job_id}")
def download_job_file(job_id: uuid.UUID, db: Session = Depends(get_db)):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status != "SUCCESS":
        raise HTTPException(status_code=400, detail="Job is not completed successfully")
    
    excel_path = job.result_metadata.get("excel_path") if job.result_metadata else None
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail="Excel file not found on server")
        
    filename = os.path.basename(excel_path)
    return FileResponse(
        path=excel_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ─── AUDIT TRAILS ENDPOINTS ───────────────────────────────────────────────────

@app.get("/api/audit-trails", response_model=List[AuditTrailResponse])
def get_audit_trails(db: Session = Depends(get_db)):
    return db.query(AuditTrail).order_by(AuditTrail.created_at.desc()).limit(100).all()

@app.get("/api/outlets/{outlet_id}/items/{item_id}/pricing-quota")
def get_pricing_quota(outlet_id: uuid.UUID, item_id: str, db: Session = Depends(get_db)):
    """Retrieve remaining pricing quota for a specific item in a specific outlet."""
    outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
    if not outlet:
        raise HTTPException(status_code=404, detail="Outlet not found")
        
    account = db.query(Account).filter(Account.id == outlet.account_id).first()
    platform = (account.platform or "").lower() if account else "grab"

    now = datetime.utcnow()
    one_day_ago = now - timedelta(days=1)
    thirty_days_ago = now - timedelta(days=30)
    
    # Query successful price updates in the last 24 hours
    daily_count = db.query(AuditTrail).filter(
        AuditTrail.outlet_id == outlet_id,
        AuditTrail.item_id == item_id,
        AuditTrail.field_changed.ilike("price"),
        AuditTrail.status.ilike("SUCCESS"),
        AuditTrail.created_at >= one_day_ago
    ).count()
    
    # Query successful price updates in the last 30 days
    monthly_count = db.query(AuditTrail).filter(
        AuditTrail.outlet_id == outlet_id,
        AuditTrail.item_id == item_id,
        AuditTrail.field_changed.ilike("price"),
        AuditTrail.status.ilike("SUCCESS"),
        AuditTrail.created_at >= thirty_days_ago
    ).count()
    
    # Dynamic rules based on platform
    if platform == "shopee":
        daily_limit = 1
        monthly_limit = 99999  # Practically unlimited monthly quota
        max_increase_pct = 25.0
    elif platform == "grab":
        daily_limit = 10
        monthly_limit = 15
        max_increase_pct = 15.0
    else:  # GoFood or fallback
        daily_limit = 99999
        monthly_limit = 99999
        max_increase_pct = 99999.0
    
    daily_remaining = max(0, daily_limit - daily_count)
    monthly_remaining = max(0, monthly_limit - monthly_count)
    
    return {
        "outlet_id": outlet_id,
        "item_id": item_id,
        "platform": platform,
        "daily_limit": daily_limit,
        "daily_count": daily_count,
        "daily_remaining": daily_remaining,
        "monthly_limit": monthly_limit,
        "monthly_count": monthly_count,
        "monthly_remaining": monthly_remaining,
        "max_increase_pct": max_increase_pct
    }

@app.get("/api/outlets/{outlet_id}/menu-items")
def get_outlet_menu_items(outlet_id: uuid.UUID, db: Session = Depends(get_db)):
    """Retrieve the menu items list of an outlet from the latest pulled Excel sheet catalog."""
    job = db.query(Job).filter(
        Job.outlet_id == outlet_id,
        Job.job_type == "PULL",
        Job.status == "SUCCESS"
    ).order_by(Job.completed_at.desc()).first()
    
    excel_path = None
    if job and job.result_metadata:
        excel_path = job.result_metadata.get("excel_path")
        
    if not excel_path or not os.path.exists(excel_path):
        # Fallback to scanning the exports folder directly
        outlet = db.query(Outlet).filter(Outlet.id == outlet_id).first()
        if not outlet:
            raise HTTPException(status_code=404, detail="Outlet not found")
        import re
        raw_outlet = outlet.nama_outlet or outlet.nama_resto_final or outlet.merchant_name or 'unknown'
        clean_outlet = "".join(c for c in raw_outlet if c.isalnum() or c in (' ', '_', '-')).strip()
        clean_outlet = re.sub(r'\s+', ' ', clean_outlet).lower()
        
        exports_dir = BASE_DIR / "data" / "exports" / outlet.platform / clean_outlet
        excel_files = list(exports_dir.glob("*.xlsx")) if exports_dir.exists() else []
        if not excel_files:
            return []
        excel_path = str(excel_files[0])

    if not excel_path or not os.path.exists(excel_path):
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        if 'Item' not in wb.sheetnames:
            return []
        sheet = wb['Item']
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            return []
            
        headers = rows[0]
        header_map = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}
        
        required_cols = ['Item ID', 'Category', 'Item', 'Current Real Price (Rp)']
        for col in required_cols:
            if col not in header_map:
                logger.warning(f"Missing column '{col}' in Excel sheet mapping: {list(header_map.keys())}")
                return []
                
        items = []
        for row in rows[1:]:
            # Skip empty rows
            if not row or all(v is None for v in row):
                continue
                
            item_id = str(row[header_map['Item ID']]).strip() if row[header_map['Item ID']] is not None else ""
            category_id = str(row[header_map['Category ID']]).strip() if 'Category ID' in header_map and row[header_map['Category ID']] is not None else ""
            category_name = str(row[header_map['Category']]).strip() if row[header_map['Category']] is not None else ""
            item_name = str(row[header_map['Item']]).strip() if row[header_map['Item']] is not None else ""
            desc = str(row[header_map['Description']]).strip() if 'Description' in header_map and row[header_map['Description']] is not None else ""
            
            p_val = row[header_map['Current Real Price (Rp)']]
            try:
                price_val = int(float(p_val)) if p_val is not None else 0
            except:
                price_val = 0
                
            avail_val = str(row[header_map['Availability']]).strip() if 'Availability' in header_map and row[header_map['Availability']] is not None else "Available"
            
            if item_id and item_name:
                items.append({
                    "id": item_id,
                    "category_id": category_id,
                    "category": category_name,
                    "name": item_name,
                    "description": desc,
                    "price": price_val,
                    "availability": avail_val
                })
        return items
    except Exception as e:
        logger.error(f"Error parsing excel menu file at {excel_path}: {e}")
        return []
