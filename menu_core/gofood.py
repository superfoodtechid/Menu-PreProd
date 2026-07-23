import os
import json
import re
import sys
import pandas as pd
import base64

def extract_gofood_menu(store_metadata: dict, output_dir: str):
    """
    Extracts menu for GoFood by running the login and redirect flow to intercept the menu API response.
    """
    store_id = store_metadata.get('store_id', '')
    nama_resto = store_metadata.get('nama_resto_final') or store_metadata.get('nama_outlet') or ''
    brand = store_metadata.get('brand') or ''
    
    print(f"\n[GoFood Menu Extractor]")
    print(f"[-] Target Outlet: {nama_resto} ({store_id})")
    
    # 1. Jalankan login_outlet untuk membuka browser, masuk dashboard, dan menangkap API menu
    # Tambahkan path agar bisa mengimpor login_gofood
    menu_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if menu_dir not in sys.path:
        sys.path.insert(0, menu_dir)
        
    try:
        from login_gofood import login_outlet
        headless_env = os.getenv("HEADLESS") or os.getenv("HEADLESS_GOFOOD")
        is_headless = headless_env.lower() in ("true", "1", "yes") if headless_env else False
        headless_str = "secara headless" if is_headless else "secara non-headless (GUI)"
        print(f"[*] Meluncurkan browser GoFood {headless_str} untuk login & pengalihan ke halaman menu...")
        login_result = login_outlet(store_metadata)
        if not login_result or not login_result.get('access_token'):
            print(f"[!] Login atau penarikan menu dibatalkan/gagal.")
            return False, "Proses login/intersepsi menu via browser gagal atau dibatalkan."
    except Exception as e:
        print(f"[!] Terjadi kesalahan saat menjalankan browser login: {e}")
        return False, f"Terjadi kesalahan saat meluncurkan browser login: {e}"
        
    # 2. Simpan dan load data menu yang ditangkap
    api_dir = os.path.join(menu_dir, "Gofood", "API")
    os.makedirs(api_dir, exist_ok=True)
    json_path = os.path.join(api_dir, f"menu-response-{store_id}.json")
    
    captured_menu = login_result.get('captured_menu')
    if not captured_menu:
        print("[!] Gagal menangkap data menu dari browser pada sesi ini.")
        return False, "Gagal menangkap data menu dari browser pada sesi ini."
        
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(captured_menu, f, indent=4)
        print(f"   💾 Menu response berhasil disimpan ke: {json_path}")
    except Exception as e:
        print(f"   ⚠️ Gagal menyimpan menu response ke file: {e}")
            
    captured_modifiers = login_result.get('captured_modifiers')
    if captured_modifiers:
        try:
            mod_json_path = os.path.join(api_dir, f"modifier-response-{store_id}.json")
            with open(mod_json_path, 'w', encoding='utf-8') as f:
                json.dump({"variant_categories": captured_modifiers}, f, indent=4)
            print(f"   💾 Modifier response hasil intercept berhasil disimpan ke: {mod_json_path}")
        except Exception as e:
            print(f"   ⚠️ Gagal menyimpan modifier response ke file: {e}")
        
    print(f"[*] Memuat data menu dari: {json_path}")
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        return False, f"Gagal membaca file JSON menu: {e}"
        
    menus = data.get("menus", [])
    if not menus:
        return False, "Data menu kosong di dalam file JSON."
        
    access_token = login_result.get('access_token')

    # --- GET GOFOOD URL WITH UUID ---
    restaurant_uuid = ""
    for cat in menus:
        if cat.get("restaurant_id"):
            restaurant_uuid = cat.get("restaurant_id")
            break
            
    # Ambil nama kota dari GoBiz API
    city_slug = "indonesia"
    if access_token:
        try:
            import requests
            headers = {
                'Accept': 'application/json, text/plain, */*',
                'Authentication-Type': 'go-id',
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            resp = requests.get(f'https://api.gobiz.co.id/v1/merchants/{store_id}', headers=headers, timeout=10)
            if resp.status_code == 200:
                merch_data = resp.json()
                if merch_data.get('outlet_city'):
                    city_raw = merch_data['outlet_city']
                    # Hapus kata "Kota", "Kabupaten", atau "Kab." di awal nama kota
                    city_raw = re.sub(r'^(kota|kabupaten|kab\.)\s+', '', city_raw, flags=re.IGNORECASE)
                    city_slug = re.sub(r'[^a-zA-Z0-9\s\-]', '', city_raw)
                    city_slug = re.sub(r'\s+', '-', city_slug.strip()).lower()
        except Exception as e:
            print(f"   ⚠️ Gagal mengambil nama kota dari API: {e}")
            
    gofood_link = f"https://gofood.link/a/{store_id}"
    if restaurant_uuid:
        raw_slug = store_metadata.get('nama_resto_final') or store_metadata.get('nama_outlet') or store_metadata.get('brand') or 'outlet'
        clean_slug = re.sub(r'[^a-zA-Z0-9\s\-]', '', raw_slug)
        clean_slug = re.sub(r'\s+', '-', clean_slug.strip()).lower()
        gofood_link = f"https://gofood.co.id/{city_slug}/restaurant/{clean_slug}-{restaurant_uuid}"

    # --- LOAD MODIFIER DATA ---
    modifier_path = os.path.join(api_dir, f"modifier-response-{store_id}.json")
        
    variant_categories_map = {}
    if os.path.exists(modifier_path):
        try:
            with open(modifier_path, 'r', encoding='utf-8') as f:
                mod_data = json.load(f)
            for m_cat in mod_data.get("variant_categories", []):
                # Map by id (local v1 ID)
                vid = m_cat.get("id")
                if vid:
                    variant_categories_map[vid] = m_cat
                # Map by common_id (v2 common ID)
                cid = m_cat.get("common_id")
                if cid:
                    variant_categories_map[cid] = m_cat
                # Map by master_variant_category_id (v1 master ID)
                mcid = m_cat.get("master_variant_category_id")
                if mcid:
                    variant_categories_map[mcid] = m_cat
            print(f"   💾 Loaded {len(variant_categories_map)} keys in variant categories map from {modifier_path}")
        except Exception as e:
            print(f"   ⚠️ Gagal memuat data modifier dari {modifier_path}: {e}")
            
    all_dishes = []
    modifier_rows = []
    
    # Parse GoFood menu categories and items
    for cat in menus:
        cat_id = cat.get("common_id") or cat.get("id", "")
        cat_name = cat.get("name", "").strip()
        cat_active = cat.get("active", True)
            
        items = cat.get("menu_items", [])
        for item in items:
            item_id = item.get("common_id") or item.get("id", "")
            item_name = item.get("name", "").strip()
            item_price_str = item.get("price", "0")
            try:
                item_price = float(item_price_str)
            except ValueError:
                item_price = 0.0
                
            item_desc = item.get("description", "").strip()
            item_active = item.get("active", True)
            item_instock = item.get("in_stock", True)
            
            visibility = "Show" if item_active else "Hide"
            ketersediaan = "Available" if item_instock else "Unavailable"
            
            # Cari link foto
            img_url = ""
            if item.get("image"):
                img_url = item["image"]
            elif item.get("image_cover") and isinstance(item["image_cover"], dict):
                img_url = item["image_cover"].get("url", "")
                
            # Modifier groups count
            var_cat_ids = item.get("variant_category_ids", [])
            mod_groups_count = len(var_cat_ids)
            total_modifiers_count = 0
            
            # Count modifiers & gather rows
            for vcat_id in var_cat_ids:
                vcat = variant_categories_map.get(vcat_id)
                if not vcat:
                    continue
                
                vcat_id_to_save = vcat.get("common_id") or vcat.get("id", "")
                
                variants = vcat.get("variants", [])
                total_modifiers_count += len(variants)
                
                vcat_name = vcat.get("name", "").strip()
                rules = vcat.get("rules") or {}
                selection = rules.get("selection") or {}
                min_qty = selection.get("min_quantity", 0)
                max_qty = selection.get("max_quantity", 0)
                tipe_modifier = "Pilihan Tunggal" if max_qty == 1 else "Pilihan Ganda"
                
                for var in variants:
                    var_id = var.get("common_id") or var.get("id", "")
                    var_name = var.get("name", "").strip()
                    var_price = float(var.get("price", 0))
                    var_active = var.get("active", True)
                    var_instock = var.get("in_stock", True)
                    var_visibility = "Show" if var_active else "Hide"
                    var_ketersediaan = "Available" if var_instock else "Unavailable"
                    
                    modifier_rows.append([
                        gofood_link,
                        nama_resto,
                        brand or nama_resto,
                        store_id,
                        item_name,
                        vcat_id_to_save,
                        vcat_name,
                        var_id,
                        var_name,
                        tipe_modifier,
                        min_qty,
                        max_qty,
                        var_price,
                        var_ketersediaan,
                        var_visibility
                    ])
            
            # Robust promotion parsing if present in API response
            promo_info = item.get("promotion")
            
            if isinstance(promo_info, str) and promo_info:
                try:
                    promo_info = json.loads(base64.b64decode(promo_info).decode('utf-8'))
                except Exception:
                    promo_info = None
                    
            harga_sebelum = item_price
            harga_setelah = item_price
            promo_val = 0
            
            if isinstance(promo_info, dict):
                disc_price = promo_info.get("selling_price") or promo_info.get("discounted_price") or promo_info.get("price") or promo_info.get("promo_price")
                if disc_price:
                    try:
                        harga_setelah = float(disc_price)
                    except ValueError:
                        pass
                
                pct = promo_info.get("discount_percentage") or promo_info.get("percentage")
                val = promo_info.get("discount_value") or promo_info.get("value") or promo_info.get("amount")
                if pct:
                    promo_val = f"{int(pct)}%"
                elif val:
                    try:
                        promo_val = f"Rp {int(float(val))}"
                    except ValueError:
                        pass
                else:
                    if harga_sebelum > harga_setelah and harga_sebelum > 0:
                        diff = harga_sebelum - harga_setelah
                        promo_val = f"{int(round(diff / harga_sebelum * 100))}%"
            
            if isinstance(promo_val, str) and "%" in promo_val:
                slash_pct = promo_val
                slash_rp = 0
            elif isinstance(promo_val, str) and "Rp" in promo_val:
                slash_pct = "0%"
                slash_rp = int(promo_val.replace("Rp", "").strip())
            else:
                slash_pct = promo_val if promo_val else "0%"
                slash_rp = 0
                
            dish_obj = {
                'link_outlet': gofood_link,
                'nama_panjang': nama_resto,
                'nama_pendek': brand or nama_resto,
                'store_id': store_id,
                'cat_id': cat_id,
                'nama_kategori': cat_name,
                'item_id': item_id,
                'nama_item': item_name,
                'jumlah_terjual': 0,
                'jumlah_modifier_group': mod_groups_count,
                'jumlah_modifier': total_modifiers_count,
                'deskripsi_item': item_desc,
                'harga_sebelum_promo': harga_sebelum,
                'harga_setelah_promo': harga_setelah,
                'slash_pct': slash_pct,
                'slash_rp': slash_rp,
                'ketersediaan': ketersediaan,
                'visibility': visibility,
                'link_foto': img_url
            }
            all_dishes.append(dish_obj)
            
    # Build standard output DataFrames
    item_cols = [
        'OFD', 'Outlet Name', 'Outlet Short Name', 'Outlet Link', 'SID',
        'Category ID', 'Category', 'Item ID', 'Item', 'Photo Link',
        'Description', 'Keyword', 'Total Sold', 'Total Modifier Group', 'Total Modifier',
        'Availability', 'Visibility', 'Current Fake Price (Rp)', 'Current Real Price (Rp)',
        'Current Slash Price (%)', 'Current Slash Price (Rp)', 'New Markup (%)',
        'New Real Price (Rp)', 'Adjustment (Rp)', 'New Final Real Price (Rp)',
        'New Slash Price (%)', 'New Fake Price (Rp)', 'Notes'
    ]
    
    item_data = []
    for d in all_dishes:
        item_data.append([
            'GoFood', d['nama_panjang'], d['nama_pendek'], d['link_outlet'], d['store_id'],
            d['cat_id'], d['nama_kategori'], d['item_id'], d['nama_item'], d['link_foto'],
            d['deskripsi_item'], '', d['jumlah_terjual'], d['jumlah_modifier_group'],
            d['jumlah_modifier'], d['ketersediaan'], d['visibility'],
            d['harga_sebelum_promo'], d['harga_setelah_promo'], d['slash_pct'], d['slash_rp'],
            "", "", "", "", "", "", ""
        ])
        
    df_items = pd.DataFrame(item_data, columns=item_cols)
    
    mod_cols = [
        'OFD', 'Outlet Name', 'Outlet Short Name', 'Outlet Link', 'SID',
        'Item', 'Modifier Group ID', 'Modifier Group', 'Modifier ID',
        'Modifier', 'Min', 'Max', 'Availability', 'Visibility', 'Current Price (Rp)'
    ]
    
    mod_data = []
    for m in modifier_rows:
        mod_data.append([
            'GoFood', m[1], m[2], m[0], m[3],
            m[4], m[5], m[6], m[7],
            m[8], m[10], m[11], m[13], m[14], m[12]
        ])
    
    df_mods = pd.DataFrame(mod_data, columns=mod_cols)
    
    # Write to files
    os.makedirs(output_dir, exist_ok=True)
    
    def clean_name(s):
        cleaned = "".join(c for c in s if c.isalnum() or c in (' ', '_', '-')).rstrip()
        return cleaned.replace(' ', '_')
        
    safe_merchant = clean_name(nama_resto)
    branch_raw = store_metadata.get('brand') or store_metadata.get('nama_resto_final') or store_metadata.get('nama_outlet') or ""
    safe_branch = clean_name(branch_raw)
    
    if safe_branch.lower() == safe_merchant.lower() or not safe_branch:
        combined_name = safe_merchant
    else:
        combined_name = f"{safe_merchant}_{safe_branch}"
        
    combined_name = re.sub(r'_+', '_', combined_name)
    
    raw_outlet = store_metadata.get('nama_outlet') or store_metadata.get('nama_resto_final') or nama_resto or 'unknown'
    raw_brand = store_metadata.get('brand') or ''
    
    def clean_filename_part(s):
        return "".join(c for c in s if c.isalnum() or c in (' ', '_', '-')).strip()
        
    clean_outlet = clean_filename_part(raw_outlet)
    clean_brand = clean_filename_part(raw_brand)
    
    if clean_brand and clean_brand.lower() != clean_outlet.lower():
        excel_filename = f"O.C5 {clean_outlet} - {clean_brand}.xlsx"
    else:
        excel_filename = f"O.C5 {clean_outlet}.xlsx"
        
    excel_path = os.path.join(output_dir, excel_filename)
    
    from pathlib import Path
    import openpyxl
    BASE_DIR = Path(__file__).resolve().parents[1]
    template_path = BASE_DIR / "O. C5 Template.xlsx"
    
    try:
        wb = openpyxl.load_workbook(template_path)
        sheet_item = wb['Item']
        if sheet_item.max_row > 1:
            sheet_item.delete_rows(2, sheet_item.max_row - 1)
        headers_item = {cell.value: cell.column for cell in sheet_item[1]}
        for r_idx, row in df_items.iterrows():
            for col_name, val in row.items():
                if col_name in headers_item:
                    if pd.isna(val):
                        val = ""
                    elif col_name in ['SID', 'Category ID', 'Item ID']:
                        if isinstance(val, float):
                            val = str(int(val)) if val.is_integer() else str(val)
                        else:
                            val = str(val)
                    sheet_item.cell(row=r_idx + 2, column=headers_item[col_name], value=val)
                    
        sheet_mod = wb['Modifier']
        if sheet_mod.max_row > 1:
            sheet_mod.delete_rows(2, sheet_mod.max_row - 1)
        headers_mod = {cell.value: cell.column for cell in sheet_mod[1]}
        for r_idx, row in df_mods.iterrows():
            for col_name, val in row.items():
                if col_name in headers_mod:
                    if pd.isna(val):
                        val = ""
                    elif col_name in ['SID', 'Modifier Group ID', 'Modifier ID', 'Item']:
                        if isinstance(val, float):
                            val = str(int(val)) if val.is_integer() else str(val)
                        else:
                            val = str(val)
                    sheet_mod.cell(row=r_idx + 2, column=headers_mod[col_name], value=val)
                    
        wb.save(excel_path)
        print(f"   ✅ Berhasil menyimpan file catalog menggunakan template O.C5 ke: {excel_path}")
    except Exception as ex_err:
        print(f"   [-] Gagal menulis ke template O.C5: {ex_err}. Fallback ke Excel biasa.")
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_items.to_excel(writer, sheet_name='Item', index=False)
            df_mods.to_excel(writer, sheet_name='Modifier', index=False)
        
    print(f"   ✅ Berhasil memproses data menu GoFood!")
    print(f"      - Item Count: {len(df_items)}")
    return True, {
        'items_csv': None,
        'mods_csv': None,
        'excel': excel_path,
        'items_count': len(df_items),
        'mods_count': len(df_mods)
    }
