import os
import json
import asyncio
import re
import pandas as pd
from pathlib import Path
import openpyxl

def run_async(coro):
    """Safely run async coroutines inside a synchronous context."""
    try:
        return asyncio.run(coro)
    except RuntimeError:
        loop = asyncio.get_event_loop_policy().get_event_loop()
        return loop.run_until_complete(coro)

def clean_name_str(s):
    """Helper to clean string for comparison (remove spaces, symbols, lowercase)."""
    return "".join(c for c in str(s).lower() if c.isalnum())

def extract_grab_menu(store_metadata: dict, output_dir: str):
    """
    Extracts GrabFood menu for a specific store.
    Downloads the entire menu under the account and filters for the target store_id.
    """
    username = store_metadata.get('username', '').strip()
    password = store_metadata.get('password', '').strip()
    store_id = store_metadata.get('store_id', '').strip()
    nama_resto = store_metadata.get('nama_resto_final') or store_metadata.get('nama_outlet') or ''
    brand = store_metadata.get('brand') or ''

    print(f"\n[GrabFood Menu Extractor]")
    print(f"[-] Target Outlet: {nama_resto} ({store_id})")

    if not username or not password:
        print("[!] Error: Username or password is empty.")
        return False, "Username/password kosong."

    # 1. Import Grab API Scraper
    try:
        import sys
        from pathlib import Path
        project_root = str(Path(__file__).resolve().parent.parent)
        if project_root not in sys.path:
            sys.path.insert(0, project_root)
        elif sys.path[0] != project_root:
            sys.path.remove(project_root)
            sys.path.insert(0, project_root)
        from grab.core.grab_api_scraper import run_api_download_for_portal
    except ImportError as e:
        print(f"[!] Error importing Grab Scraper: {e}")
        return False, f"Gagal mengimpor Grab Scraper: {e}"

    # 2. Run Playwright Grab download flow
    print(f"[*] Meluncurkan browser untuk menarik menu dari portal Grab Merchant...")
    try:
        json_path, error_msg = run_async(run_api_download_for_portal(username, password))
        if error_msg or not json_path or not os.path.exists(json_path):
            print(f"[!] Gagal menarik menu GrabFood: {error_msg}")
            return False, f"Gagal menarik menu GrabFood: {error_msg}"
    except Exception as e:
        print(f"[!] Terjadi pengecualian saat menjalankan scraper: {e}")
        return False, f"Terjadi kesalahan saat menjalankan scraper: {e}"

    print(f"   💾 Download JSON berhasil: {json_path}")

    # 3. Load downloaded JSON data
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            scraped_data = json.load(f)
    except Exception as e:
        print(f"[!] Gagal membaca data JSON menu GrabFood: {e}")
        return False, f"Gagal membaca data JSON menu Grab: {e}"

    # 4. Clean up temporary json file immediately
    try:
        os.unlink(json_path)
    except Exception as e:
        print(f"   ⚠️ Gagal menghapus file unduhan sementara {json_path}: {e}")

    items = scraped_data.get('items', [])
    modifiers = scraped_data.get('modifiers', [])

    # 5. Filter items & modifiers for the target store_id (or fallback matching)
    matched_items = []
    for item in items:
        item_sid = str(item.get("Store ID", "")).strip()
        target_sid = str(store_id).strip()
        if target_sid and item_sid.lower() == target_sid.lower():
            matched_items.append(item)
        elif not target_sid and clean_name_str(item.get("Nama panjang", "")) == clean_name_str(nama_resto):
            matched_items.append(item)

    matched_mods = []
    for mod in modifiers:
        mod_sid = str(mod.get("Store ID", "")).strip()
        target_sid = str(store_id).strip()
        if target_sid and mod_sid.lower() == target_sid.lower():
            matched_mods.append(mod)
        elif not target_sid and clean_name_str(mod.get("Nama panjang", "")) == clean_name_str(nama_resto):
            matched_mods.append(mod)

    print(f"   📊 Berhasil menyaring menu:")
    print(f"      - Item Terkait: {len(matched_items)} dari {len(items)}")
    print(f"      - Modifier Terkait: {len(matched_mods)} dari {len(modifiers)}")

    if not matched_items:
        print(f"[!] Warning: Tidak ada item menu yang cocok untuk Store ID '{store_id}' / Nama '{nama_resto}'.")
        # Kita tetap buat file kosong tapi laporkan warning
        # return False, f"Tidak ada menu yang cocok dengan Store ID {store_id}"

    # 6. Build file name conforming to "O.C5 {nama_outlet} - {brand}.xlsx"
    def clean_filename_part(s):
        return "".join(c for c in s if c.isalnum() or c in (' ', '_', '-')).strip()

    clean_outlet = clean_filename_part(nama_resto)
    clean_brand = clean_filename_part(brand)

    if clean_brand and clean_brand.lower() != clean_outlet.lower():
        excel_filename = f"O.C5 {clean_outlet} - {clean_brand}.xlsx"
    else:
        excel_filename = f"O.C5 {clean_outlet}.xlsx"

    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, excel_filename)

    # 7. Write to O. C5 Template Excel
    BASE_DIR = Path(__file__).resolve().parents[1]
    template_path = BASE_DIR / "O. C5 Template.xlsx"

    try:
        wb = openpyxl.load_workbook(template_path)
        
        # 1. Fill Item Sheet
        sheet_item = wb['Item']
        # Delete sample data rows (from row 2 onwards)
        if sheet_item.max_row > 1:
            sheet_item.delete_rows(2, sheet_item.max_row - 1)
            
        headers_item = {cell.value: cell.column for cell in sheet_item[1]}
        
        new_row_idx = 2
        for item in matched_items:
            avail_str = str(item.get('Ketersediaan item', '')).lower()
            availability = "Available" if avail_str in ("available", "active", "1", "true") else "Unavailable"
            
            fake_price = item.get('Harga item sebelum promo (harga coret)', 0.0)
            real_price = item.get('Harga item setelah promo (harga coret)', 0.0)
            slash_rp = item.get('Nominal atau persentase promo (harga coret)', 0.0)
            
            if fake_price > real_price and fake_price > 0:
                pct = round(((fake_price - real_price) / fake_price) * 100.0, 2)
                slash_pct = f"{int(pct)}%" if pct.is_integer() else f"{pct}%"
            else:
                slash_pct = "0%"
                slash_rp = 0
                
            mapping = {
                'OFD': 'GrabFood',
                'Outlet Name': item.get('Nama panjang', nama_resto),
                'Outlet Short Name': brand or item.get('Nama panjang', nama_resto),
                'Outlet Link': item.get('Link outlet', f"https://food.grab.com/id/en/restaurant/{store_id}"),
                'SID': item.get('Store ID', store_id),
                'Category ID': item.get('Category ID', ''),
                'Category': item.get('Nama kategori', ''),
                'Item ID': item.get('Item ID', ''),
                'Item': item.get('Nama item', ''),
                'Photo Link': item.get('Link foto', ''),
                'Description': item.get('Deskripsi item', ''),
                'Keyword': '',
                'Total Sold': item.get('Jumlah terjual', 0),
                'Total Modifier Group': item.get('Jumlah modifier group', 0),
                'Total Modifier': item.get('Jumlah modifier', 0),
                'Availability': availability,
                'Visibility': 'Show',
                'Current Fake Price (Rp)': fake_price,
                'Current Real Price (Rp)': real_price,
                'Current Slash Price (%)': slash_pct,
                'Current Slash Price (Rp)': slash_rp
            }
            
            for key, val in mapping.items():
                if key in headers_item:
                    col_idx = headers_item[key]
                    if pd.isna(val):
                        val = ""
                    elif key in ['SID', 'Category ID', 'Item ID']:
                        if isinstance(val, float):
                            val = str(int(val)) if val.is_integer() else str(val)
                        else:
                            val = str(val)
                    sheet_item.cell(row=new_row_idx, column=col_idx, value=val)
            new_row_idx += 1
                    
        # 2. Fill Modifier Sheet
        sheet_mod = wb['Modifier']
        # Delete sample data rows (from row 2 onwards)
        if sheet_mod.max_row > 1:
            sheet_mod.delete_rows(2, sheet_mod.max_row - 1)
            
        headers_mod = {cell.value: cell.column for cell in sheet_mod[1]}
        
        new_row_idx = 2
        for mod in matched_mods:
            mod_avail_str = str(mod.get('Ketersediaan modifier', '')).lower()
            mod_availability = "Available" if mod_avail_str in ("available", "active", "1", "true") else "Unavailable"
            
            mapping_mod = {
                'OFD': 'GrabFood',
                'Outlet Name': mod.get('Nama panjang', nama_resto),
                'Outlet Short Name': brand or mod.get('Nama panjang', nama_resto),
                'Outlet Link': mod.get('Link outlet', f"https://food.grab.com/id/en/restaurant/{store_id}"),
                'SID': mod.get('Store ID', store_id),
                'Item': mod.get('Nama item', ''),
                'Modifier Group ID': mod.get('Modifier Group ID', ''),
                'Modifier Group': mod.get('Nama modifier group', ''),
                'Modifier ID': mod.get('Modifier ID', ''),
                'Modifier': mod.get('Nama modifier', ''),
                'Min': mod.get('Minimal', 0),
                'Max': mod.get('Maksimal', 1),
                'Availability': mod_availability,
                'Visibility': 'Show',
                'Current Price (Rp)': mod.get('Harga modifier', 0.0)
            }
            
            for key, val in mapping_mod.items():
                if key in headers_mod:
                    col_idx = headers_mod[key]
                    if pd.isna(val):
                        val = ""
                    elif key in ['SID', 'Modifier Group ID', 'Modifier ID', 'Item']:
                        if isinstance(val, float):
                            val = str(int(val)) if val.is_integer() else str(val)
                        else:
                            val = str(val)
                    sheet_mod.cell(row=new_row_idx, column=col_idx, value=val)
            new_row_idx += 1
                    
        wb.save(excel_path)
        print(f"   ✅ Berhasil menyimpan file catalog menggunakan template O.C5 ke: {excel_path}")
    except Exception as ex_err:
        print(f"   [-] Gagal menulis ke template O.C5: {ex_err}. Fallback ke Excel biasa.")
        # Fallback to standard excel writer if template fails
        df_items_fallback = pd.DataFrame(matched_items)
        df_mods_fallback = pd.DataFrame(matched_mods)
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_items_fallback.to_excel(writer, sheet_name='Item', index=False)
            df_mods_fallback.to_excel(writer, sheet_name='Modifier', index=False)

    return True, {
        'items_csv': None,
        'mods_csv': None,
        'excel': excel_path,
        'items_count': len(matched_items),
        'mods_count': len(matched_mods)
    }
