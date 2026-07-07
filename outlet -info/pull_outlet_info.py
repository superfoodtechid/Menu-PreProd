"""
Shopee Outlet Info Puller
=========================
Menarik data outlet/store dari Shopee Partner API (stores + store detail)
untuk 4 merchant target: SuperFood, DoEat, LOKARASA, WonderFood.

Menggunakan session allvbadmin via browser.get_session() (pattern dari menu_core/shopee.py).
Output: Shopee_{timestamp}.xlsx (tab All + tab per-merchant).

Usage:
    python pull_outlet_info.py
"""

import json
import os
import sys
import time
import math
from datetime import datetime
import requests
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font

from selenium.webdriver.chrome.options import Options

# ──────────────────────────────────────────────────────────────
# Path Setup
# ──────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent  # menu-shopee
AUTOMATION_DIR = Path("/home/akbarhann/project/FoodMaster/menu-prod/src/shopee-omzet-automation")

# Add automation dir to sys.path for browser module import
if str(AUTOMATION_DIR) not in sys.path:
    sys.path.insert(0, str(AUTOMATION_DIR))

from core import browser

# FORCE the profile directory (same pattern as menu_core/shopee.py)
orig_add_argument = Options.add_argument
def custom_add_argument(self, argument):
    if "--user-data-dir=" in argument:
        argument = f"--user-data-dir=/home/akbarhann/project/FoodMaster/menu-prod/outlet -info/data/chrome_profile"
        print(f"🔧 [PATCH] Mengalihkan user data dir ke: {argument}")
    orig_add_argument(self, argument)
Options.add_argument = custom_add_argument

# ──────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────
SELLER_BASE = "https://foody.shopee.co.id"
SESSION_FILE = SCRIPT_DIR / "data" / "session.json"
CREDS_FILE = SCRIPT_DIR / "credentials.json"
BANK_ACC_FILE = SCRIPT_DIR / "bank_acc.json"
OUTPUT_DIR = SCRIPT_DIR / "data"

# Allvbadmin credentials (same as menu_core/shopee.py)
DEFAULT_USERNAME = "allvbadmin"
DEFAULT_PASSWORD = "Shopee@321"

# Merchants to switch to in allvbadmin (each has its own set of stores)
# Format: (merchant_name_in_dropdown, label_for_output)
# All stores under each merchant are taken (no keyword filtering)
MERCHANTS_TO_SWITCH = [
    ("SuperFood", "SuperFood"),
    ("Gurame Bakar, Do Eat", "DoEat"),
    ("LOKARASA", "LOKARASA"),
    ("WonderFood", "WonderFood"),
]

STATUS_MAP = {
    1: "Active",
    0: "Suspended",
   
}

PORTAL_MAP = {
    "SuperFood": "F",
    "WonderFood": "W",
    "LOKARASA": "L",
    "DoEat": "D",
}


# ──────────────────────────────────────────────────────────────
# Authentication (reuse pattern from menu_core/shopee.py)
# ──────────────────────────────────────────────────────────────
def get_auth_session(target_name: str = "SuperFood") -> tuple:
    """
    Launch browser, login as allvbadmin, switch to target merchant, extract tokens, close browser.
    Returns (tob_token, entity_id, extra_cookies) or raises on failure.
    
    target_name: merchant name to switch to (e.g. 'SuperFood') — this determines
                 which merchant group (shopee_foody_mid) the API session belongs to.
                 All stores under that merchant group will be accessible.
    """
    browser.set_session_file(SESSION_FILE)

    username = DEFAULT_USERNAME
    password = DEFAULT_PASSWORD
    if CREDS_FILE.exists():
        try:
            creds = json.loads(CREDS_FILE.read_text())
            username = creds.get("shopee_username", username)
            password = creds.get("shopee_password", password)
        except Exception:
            pass

    print(f"[*] Membuka browser (headless=True) dan memilih merchant: '{target_name}'...")
    session_data = browser.get_session(
        username=username,
        password=password,
        headless=False,
        close_browser=False,
        target_name=target_name,
        interactive=False,
    )

    if not session_data or "driver" not in session_data:
        raise RuntimeError("Gagal menginisialisasi browser atau login allvbadmin.")

    driver = session_data["driver"]

    try:
        print("[*] Memperbarui token autentikasi...")
        session = browser.refresh_tokens(driver)
        if not session or "shopee_tob_token" not in session:
            raise RuntimeError("Gagal memperbarui token autentikasi.")

        tob_token = session["shopee_tob_token"]
        entity_id = session.get("shopee_tob_entity_id", "")
        extra_cookies = session.get("extra_cookies", {})

        print(f"[✓] Token berhasil didapat. Entity ID: {entity_id}")
        return tob_token, entity_id, extra_cookies
    finally:
        try:
            driver.quit()
        except Exception:
            pass


# ──────────────────────────────────────────────────────────────
# Shopee Outlet Client
# ──────────────────────────────────────────────────────────────
class ShopeeOutletClient:
    """Client for Shopee Seller API - Outlet/Store info endpoints."""

    def __init__(self, tob_token: str, entity_id: str, extra_cookies: dict = None):
        self.tob_token = tob_token
        self.entity_id = entity_id
        self.extra_cookies = extra_cookies or {}
        self.session = requests.Session()
        self.user_agent = (
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36"
        )

    def _seller_headers(self, override_entity_id: str = None) -> dict:
        """Build headers with cookie string, optionally overriding entity_id."""
        eid = override_entity_id or self.entity_id
        cookies = self.extra_cookies.copy()
        cookies["shopee_tob_token"] = self.tob_token
        cookies["shopee_tob_entity_id"] = eid
        cookie_str = "; ".join(f"{k}={v}" for k, v in cookies.items())

        return {
            "Host": "foody.shopee.co.id",
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json;charset=UTF-8",
            "User-Agent": self.user_agent,
            "Cookie": cookie_str,
            "Origin": "https://partner.shopee.co.id",
            "Referer": "https://partner.shopee.co.id/",
        }

    def get_list_stores(self, page_no: int = 1, page_size: int = 50) -> dict:
        """POST /api/seller/stores - List all stores with pagination."""
        url = f"{SELLER_BASE}/api/seller/stores"
        payload = {
            "store_name": "",
            "page_size": page_size,
            "page_no": page_no,
        }

        try:
            resp = self.session.post(
                url,
                json=payload,
                headers=self._seller_headers(),
                timeout=30,
            )
            data = resp.json()
            if data.get("code") == 0:
                return data
            print(f"  [!] get_list_stores failed: code={data.get('code')}, msg={data.get('msg')}")
        except Exception as e:
            print(f"  [!] get_list_stores error: {e}")
        return {"data": {"total": 0, "stores": []}}

    def get_all_stores(self) -> list[dict]:
        """Fetch ALL stores across all pages."""
        all_stores = []
        page_no = 1
        page_size = 50

        # First call to get total
        first_resp = self.get_list_stores(page_no=1, page_size=page_size)
        total = first_resp.get("data", {}).get("total", 0)
        stores = first_resp.get("data", {}).get("stores", [])
        all_stores.extend(stores)

        if total == 0:
            print("[!] Tidak ada store yang ditemukan.")
            return []

        total_pages = math.ceil(total / page_size)
        print(f"[*] Total stores: {total} | Total pages: {total_pages}")
        print(f"  ✓ Page 1/{total_pages}: {len(stores)} stores")

        # Fetch remaining pages
        for page_no in range(2, total_pages + 1):
            time.sleep(0.3)  # Rate limit
            resp = self.get_list_stores(page_no=page_no, page_size=page_size)
            stores = resp.get("data", {}).get("stores", [])
            all_stores.extend(stores)
            print(f"  ✓ Page {page_no}/{total_pages}: {len(stores)} stores")

        print(f"[✓] Total {len(all_stores)} stores fetched.")
        return all_stores

    def get_store_detail(self, store_id: str) -> dict | None:
        """GET /api/seller/store - Get detail for a specific store."""
        url = f"{SELLER_BASE}/api/seller/store"

        try:
            resp = self.session.get(
                url,
                headers=self._seller_headers(override_entity_id=store_id),
                timeout=15,
            )
            data = resp.json()
            if data.get("code") == 0:
                return data.get("data", {}).get("store", {})
            print(f"  [!] get_store_detail({store_id}) failed: code={data.get('code')}, msg={data.get('msg')}")
        except Exception as e:
            print(f"  [!] get_store_detail({store_id}) error: {e}")
        return None





# ──────────────────────────────────────────────────────────────
# Detail Fetching Helper
# ──────────────────────────────────────────────────────────────
def fetch_store_details(client, filtered_stores, bank_data):
    """Fetch detail for each filtered store and return result dicts."""
    results = []
    failed_count = 0

    for idx, fs in enumerate(filtered_stores, 1):
        store = fs["store"]
        merchant_label = fs["merchant"]
        store_id = store["id"]
        store_name = store.get("name", "")
        store_status = store.get("status", 0)

        if idx % 20 == 0 or idx == 1:
            print(f"  [{idx}/{len(filtered_stores)}] Processing: {store_name} ...")

        # Get detail for address
        time.sleep(0.5)  # Rate limit
        detail = client.get_store_detail(store_id)

        if detail:
            location = detail.get("location", {})
            address_parts = []
            addr = location.get("address", "")
            district = location.get("district", "")
            city = location.get("city", "")
            state = location.get("state", "")

            if addr:
                address_parts.append(addr)
            if district:
                address_parts.append(district)
            if city:
                address_parts.append(city)
            if state:
                address_parts.append(state)

            full_address = ", ".join(address_parts)
            group_id = detail.get("merchant_id", "")
            status_val = detail.get("status", store_status)
        else:
            full_address = ""
            group_id = ""
            status_val = store_status
            failed_count += 1

        results.append({
            "Portal": PORTAL_MAP.get(merchant_label, ""),
            "Nama": store_name,
            "Merchant": merchant_label,
            "Group ID": group_id,
            "Store ID": store_id,
            "Status": STATUS_MAP.get(status_val, str(status_val)),
            "Alamat": full_address,
            "Bank Account Name": bank_data.get("BANK_ACCOUNT_NAME", ""),
            "Bank Name": bank_data.get("BANK_NAME", ""),
            "Bank Account No": bank_data.get("BANK_ACCOUNT_NO", ""),
            "Bank Account": bank_data.get("BANK_ACCOUNT", ""),
        })

    return results, failed_count


# ──────────────────────────────────────────────────────────────
# Main Logic
# ──────────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("  SHOPEE OUTLET INFO PULLER")
    print("  Target: SuperFood, WonderFood, Lokarasa, Gurame Bakar/Do Eat")
    print("=" * 70)
    print()

    # 1. Load bank account data (static)
    print("[1/5] Loading bank account data...")
    if BANK_ACC_FILE.exists():
        bank_data = json.loads(BANK_ACC_FILE.read_text())
        print(f"  ✓ Bank: {bank_data.get('BANK_NAME')} - {bank_data.get('BANK_ACCOUNT_NAME')}")
    else:
        print(f"  [!] bank_acc.json tidak ditemukan: {BANK_ACC_FILE}")
        bank_data = {}

    # 2-4. Loop through each merchant to switch, pull stores, fetch details
    all_results = []
    seen_store_ids = set()  # Deduplication
    merchant_counts = {}

    for switch_idx, (merchant_name, fallback_label) in enumerate(MERCHANTS_TO_SWITCH, 1):
        print()
        print(f"{'='*70}")
        print(f"  MERCHANT {switch_idx}/{len(MERCHANTS_TO_SWITCH)}: {merchant_name}")
        print(f"{'='*70}")

        # 2. Authenticate & switch to this merchant
        print(f"\n[2/5] Authenticating & switching to '{merchant_name}'...")
        try:
            tob_token, entity_id, extra_cookies = get_auth_session(target_name=merchant_name)
        except Exception as e:
            print(f"  [!] Gagal auth untuk merchant '{merchant_name}': {e}")
            continue

        # 3. Fetch all stores for this merchant
        print(f"\n[3/5] Fetching all stores for '{merchant_name}' (paginated)...")
        client = ShopeeOutletClient(
            tob_token=tob_token,
            entity_id=entity_id,
            extra_cookies=extra_cookies,
        )
        all_stores = client.get_all_stores()

        if not all_stores:
            print(f"  [!] Tidak ada store ditemukan untuk '{merchant_name}'.")
            continue

        # 4. Label all stores with the merchant account name (no re-labeling)
        print(f"\n[4/5] Collecting stores for '{merchant_name}'...")
        filtered_stores = []
        for store in all_stores:
            store_id = store.get("id", "")
            if store_id in seen_store_ids:
                continue  # Skip duplicates across merchants

            filtered_stores.append({
                "store": store,
                "merchant": fallback_label,
            })
            seen_store_ids.add(store_id)

        print(f"  ✓ {len(filtered_stores)} stores from '{merchant_name}' processed")

        # Count per merchant label
        for fs in filtered_stores:
            m = fs["merchant"]
            merchant_counts[m] = merchant_counts.get(m, 0) + 1

        # 5. Fetch detail for each filtered store
        if filtered_stores:
            print(f"\n  Fetching detail for {len(filtered_stores)} stores...")
            results, failed = fetch_store_details(client, filtered_stores, bank_data)
            all_results.extend(results)
            print(f"  ✓ Detail fetched: {len(results)} OK, {failed} failed")

    # Summary per merchant
    print()
    print(f"  Total per merchant:")
    for m, c in sorted(merchant_counts.items()):
        print(f"    - {m}: {c} stores")

    if not all_results:
        print("\n[!] Tidak ada outlet yang cocok. Proses dihentikan.")
        return

    # 6. Build DataFrame & Export
    print()
    print("[5/5] Exporting data...")

    columns = [
        "Portal", "Nama", "Merchant", "Group ID", "Store ID", "Status", "Alamat",
        "Bank Account Name", "Bank Name", "Bank Account No", "Bank Account",
    ]
    df = pd.DataFrame(all_results, columns=columns)

    # Sort rows by Portal order: F, W, L, D
    df["Portal"] = pd.Categorical(df["Portal"], categories=["F", "W", "L", "D"], ordered=True)
    df = df.sort_values("Portal").reset_index(drop=True)
    df["Portal"] = df["Portal"].astype(str)

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Output: Shopee_{timestamp}.xlsx with tab Main Tab + per-merchant tabs F, W, L, D
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    excel_path = OUTPUT_DIR / f"Shopee_{timestamp}.xlsx"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Tab 1: Main Tab (combined)
        df.to_excel(writer, sheet_name="Main Tab", index=False)
        print(f"  ✓ Tab 'Main Tab': {len(df)} rows")

        # Tab per-merchant (ordered: F, W, L, D)
        for portal in ["F", "W", "L", "D"]:
            df_portal = df[df["Portal"] == portal]
            if len(df_portal) > 0:
                df_portal.to_excel(writer, sheet_name=portal, index=False)
                print(f"  ✓ Tab '{portal}': {len(df_portal)} rows")

        # Apply style rules: Arial, size 10, bold headers
        font_header = Font(name="Arial", size=10, bold=True)
        font_body = Font(name="Arial", size=10, bold=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # Format header row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = font_header
            # Format body rows
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = font_body

    print(f"  ✓ Output: {excel_path}")

    # Final Summary
    print()
    print("=" * 70)
    print("  SELESAI!")
    print(f"  Total outlet: {len(df)}")
    for m, c in sorted(merchant_counts.items()):
        print(f"    - {m}: {c}")
    print(f"  Output: {OUTPUT_DIR}")
    print("=" * 70)


if __name__ == "__main__":
    main()
