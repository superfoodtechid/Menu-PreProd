# -*- coding: utf-8 -*-
import os
import sys
import json
import time
import pandas as pd
from pathlib import Path
from shopee.core.client import ShopeeClient, ShopeeModifyClient

# Dynamic paths based on current file location
FILE_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = FILE_DIR.parent.parent
MENU_DIR = WORKSPACE_DIR / "shopee"
AUTOMATION_DIR = WORKSPACE_DIR / "src" / "shopee-omzet-automation"
if str(AUTOMATION_DIR) not in sys.path:
    sys.path.insert(0, str(AUTOMATION_DIR))
from core import browser

IMG_BASE = "https://down-id.img.susercontent.com/file"

def list_menu_shopee(store_metadata: dict) -> tuple[bool, list | str]:
    from shopee.core.item.edit import _boot_client
    client, err = _boot_client(store_metadata, headless=True)
    if not client:
        return False, f"Boot client failed: {err}"
        
    store_id = store_metadata["store_id"]
    catalogs = client.get_store_dishes(store_id)
    return True, catalogs

def extract_shopee_menu(store_metadata: dict, output_dir: str):
    store_id = store_metadata.get('store_id', '')
    if isinstance(store_id, str):
        store_id = store_id.strip().split('.')[0]
    if not store_id or store_id == '-' or store_id.lower() == 'nan':
        store_id = None
        
    m_name = store_metadata.get('merchant_name', '')
    if not m_name or m_name.lower() == 'nan' or m_name == '-':
        target_name = store_metadata.get('nama_resto_final') or store_metadata.get('nama_outlet') or ''
    else:
        target_name = m_name
        
    nama_panjang = store_metadata.get('nama_resto_final') or store_metadata.get('nama_outlet') or target_name
    nama_pendek = store_metadata.get('brand') or store_metadata.get('nama_pendek') or store_metadata.get('nama_outlet') or target_name
    
    username = store_metadata.get("username", "allvbadmin")
    password = store_metadata.get("password", "Shopee@321")
    session_file = MENU_DIR / "data" / "session.json"
    browser.set_session_file(session_file)
            
    print(f"[*] Membuka browser (headless=True) dan memilih merchant: '{target_name}'...")
    session_data = browser.get_session(
        username=username,
        password=password,
        headless=True,
        close_browser=True,
        target_name=target_name,
        interactive=False
    )
    
    if not session_data or "shopee_tob_token" not in session_data:
        return False, "Gagal menginisialisasi browser atau memilih merchant."
        
    tob_token = session_data["shopee_tob_token"]
    extra_cookies = session_data.get("extra_cookies", {})
    
    # If store_id was not originally known, resolve it from the active session
    if not store_id:
        session_store_id = session_data.get("shopee_tob_entity_id")
        if session_store_id:
            store_id = str(session_store_id).strip()
            store_metadata['store_id'] = store_id
            print(f"[+] Berhasil mendeteksi Store ID secara dinamis: {store_id}")
            
            try:
                db_path = MENU_DIR / "data" / "menu_management.db"
                import sqlite3
                conn = sqlite3.connect(str(db_path))
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO outlets (store_id, platform, username, password, merchant_name) "
                    "VALUES (?, 'shopee', ?, ?, ?) "
                    "ON CONFLICT(store_id) DO UPDATE SET username=excluded.username, password=excluded.password, merchant_name=excluded.merchant_name",
                    (store_id, username, password, target_name)
                )
                conn.commit()
                conn.close()
                print(f"[+] Menyimpan Store ID {store_id} untuk '{target_name}' ke database lokal.")
            except Exception as dbe:
                print(f"[-] Warning: Gagal menyimpan Store ID ke DB: {dbe}")
        else:
            return False, "Gagal mendeteksi Store ID untuk merchant ini."
            
    print(f"[DEBUG PULL] tob_token: {tob_token[:30]}...")
    print(f"[DEBUG PULL] entity_id: {store_id}")
    print(f"[DEBUG PULL] extra_cookies: {extra_cookies}")
    
    client = ShopeeClient(
        tob_token=tob_token,
        entity_id=store_id,
        extra_cookies=extra_cookies
    )
        
    try:
        print(f"[*] Menarik data menu ShopeeFood untuk: {target_name} ({store_id})...")
        catalogs = client.get_store_dishes(store_id)
        if not catalogs:
            return False, "Tidak ada data catalog/dishes yang ditemukan. Periksa session."
            
        print(f"[*] Ditemukan {len(catalogs)} kategori menu.")
        all_dishes = []
        dish_ids_with_modifiers = []
        
        for cat in catalogs:
            cat_name = cat.get('name', 'Menu Lainnya')
            cat_id = str(cat.get('id', ''))
            dishes = cat.get('dishes', [])
            for dish in dishes:
                dish_id = str(dish.get('id'))
                dish_name = dish.get('name', '')
                price_raw = dish.get('price', '0')
                list_price_raw = dish.get('list_price', '0')
                description = dish.get('description', '')
                available = dish.get('available', True)
                opt_group_count = dish.get('option_group_count', 0)
                sales_volume = dish.get('sales_volume', 0)
                picture = dish.get('picture', '')
                discount_pct = float(dish.get('discount_percentage', 0))
                
                price = float(price_raw) / 100000.0
                list_price = float(list_price_raw) / 100000.0 if (list_price_raw and float(list_price_raw) > 0) else price
                
                if list_price > price:
                    current_fake_price = list_price
                    current_real_price = price
                    current_slash_pct = round(((list_price - price) / list_price) * 100.0, 2)
                    current_slash_rp = price
                else:
                    current_fake_price = list_price
                    current_real_price = list_price
                    current_slash_pct = ""
                    current_slash_rp = ""
                
                available_str = "Tersedia" if available else "Habis"
                picture_url = f"{IMG_BASE}/{picture}" if picture else ""
                link_outlet = f"https://shopee.co.id/now-food/shop/{store_id}"
                
                dish_info = {
                    'ofd': 'ShopeeFood',
                    'link_outlet': link_outlet,
                    'nama_panjang': nama_panjang,
                    'nama_pendek': nama_pendek,
                    'store_id': store_id,
                    'cat_id': cat_id,
                    'nama_kategori': cat_name,
                    'nama_item': dish_name,
                    'jumlah_terjual': sales_volume,
                    'opt_group_count': opt_group_count,
                    'deskripsi_item': description,
                    'keyword': '',
                    'harga_sebelum_promo': current_fake_price,
                    'harga_setelah_promo': current_real_price,
                    'discount_pct_val': current_slash_pct,
                    'slash_price_rp': current_slash_rp,
                    'ketersediaan': available_str,
                    'link_foto': picture_url,
                    'dish_id': dish_id,
                    'jumlah_modifier_group': 0,
                    'jumlah_modifier': 0
                }
                all_dishes.append(dish_info)
                
                if opt_group_count > 0:
                    dish_ids_with_modifiers.append(dish_id)

        print(f"[*] Total {len(all_dishes)} item ditemukan.")
        print(f"[*] Menarik modifier untuk {len(dish_ids_with_modifiers)} item yang memiliki topping/opsi...")
        
        modifier_rows = []
        for dish_id in dish_ids_with_modifiers:
            dish_obj = next((d for d in all_dishes if d['dish_id'] == dish_id), None)
            if not dish_obj:
                continue
                
            opt_groups = client.get_store_option_groups(store_id, dish_ids=[dish_id])
            dish_obj['jumlah_modifier_group'] = len(opt_groups)
            total_modifiers_count = 0
            
            group_names = []
            for group in opt_groups:
                opt_group_info = group.get('option_group', {})
                group_name = opt_group_info.get('name', '').strip()
                if group_name:
                    group_names.append(group_name)
                select_min = opt_group_info.get('select_min', 0)
                select_max = opt_group_info.get('select_max', 0)
                options = group.get('options', [])
                
                total_modifiers_count += len(options)
                tipe_modifier = "Pilihan Tunggal" if select_max == 1 else "Pilihan Ganda"
                
                for opt in options:
                    opt_name = opt.get('name', '')
                    opt_price = float(opt.get('price', '0')) / 100000.0
                    opt_available = opt.get('available', True)
                    opt_available_str = "Tersedia" if opt_available else "Habis"
                    
                    modifier_rows.append({
                        'ofd': 'ShopeeFood',
                        'outlet_name': nama_panjang,
                        'outlet_short_name': nama_pendek,
                        'outlet_link': dish_obj['link_outlet'],
                        'sid': store_id,
                        'item': dish_obj['nama_item'],
                        'modifier_group_id': str(opt_group_info.get('id', '')),
                        'modifier_group': group_name,
                        'modifier_id': str(opt.get('id', '')),
                        'modifier': opt_name,
                        'min': select_min,
                        'max': select_max,
                        'availability': opt_available_str,
                        'current_price': opt_price
                    })
            
            dish_obj['jumlah_modifier'] = total_modifiers_count
            dish_obj['keyword'] = ", ".join(group_names)
            
        item_cols = [
            'OFD', 'Outlet Name', 'Outlet Short Name', 'Outlet Link', 'SID',
            'Category ID', 'Category', 'Item ID', 'Item', 'Photo Link',
            'Description', 'Keyword', 'Total Sold', 'Total Modifier Group', 'Total Modifier',
            'Availability', 'Current Fake Price (Rp)', 'Current Real Price (Rp)',
            'Current Slash Price (%)', 'Current Slash Price (Rp)', 'New Markup (%)',
            'New Real Price (Rp)', 'Adjustment (Rp)', 'New Final Real Price (Rp)',
            'New Slash Price (%)', 'New Fake Price (Rp)', 'Notes'
        ]
        
        item_data = []
        for d in all_dishes:
            item_data.append([
                d['ofd'], d['nama_panjang'], d['nama_pendek'], d['link_outlet'], d['store_id'],
                d['cat_id'], d['nama_kategori'], d['dish_id'], d['nama_item'], d['link_foto'],
                d['deskripsi_item'], d['keyword'], d['jumlah_terjual'], d['jumlah_modifier_group'], d['jumlah_modifier'],
                d['ketersediaan'], d['harga_sebelum_promo'], d['harga_setelah_promo'],
                d['discount_pct_val'], d['slash_price_rp'],
                "", "", "", "", "", "", ""
            ])
            
        df_items = pd.DataFrame(item_data, columns=item_cols)
        
        mod_cols = [
            'OFD', 'Outlet Name', 'Outlet Short Name', 'Outlet Link', 'SID',
            'Item', 'Modifier Group ID', 'Modifier Group', 'Modifier ID',
            'Modifier', 'Min', 'Max', 'Availability', 'Current Price (Rp)'
        ]
        
        mod_data = []
        for m in modifier_rows:
            mod_data.append([
                m['ofd'], m['outlet_name'], m['outlet_short_name'], m['outlet_link'], m['sid'],
                m['item'], m['modifier_group_id'], m['modifier_group'], m['modifier_id'],
                m['modifier'], m['min'], m['max'], m['availability'], m['current_price']
            ])
            
        df_mods = pd.DataFrame(mod_data, columns=mod_cols)
        os.makedirs(output_dir, exist_ok=True)
        
        import re
        def clean_name(s):
            cleaned = "".join(c for c in s if c.isalnum() or c in (' ', '_', '-')).rstrip()
            return cleaned.replace(' ', '_')
            
        safe_merchant = clean_name(target_name)
        branch_raw = store_metadata.get('brand') or store_metadata.get('nama_resto_final') or store_metadata.get('nama_outlet') or ""
        safe_branch = clean_name(branch_raw)
        
        if safe_branch.lower() == safe_merchant.lower() or not safe_branch:
            combined_name = safe_merchant
        else:
            combined_name = f"{safe_merchant}_{safe_branch}"
            
        combined_name = re.sub(r'_+', '_', combined_name)
        
        # Build paths conforming to "O.C5 {nama_outlet} - {brand}.xlsx"
        raw_outlet = store_metadata.get('nama_outlet') or store_metadata.get('nama_resto_final') or target_name or 'unknown'
        raw_brand = store_metadata.get('brand') or ''
        
        def clean_filename_part(s):
            return "".join(c for c in s if c.isalnum() or c in (' ', '_', '-')).strip()
            
        clean_outlet = clean_filename_part(raw_outlet)
        clean_brand = clean_filename_part(raw_brand)
        
        if clean_brand and clean_brand.lower() != clean_outlet.lower():
            excel_filename = f"O.C5 {clean_outlet} - {clean_brand}.xlsx"
        else:
            excel_filename = f"O.C5 {clean_outlet}.xlsx"
            
        excel_path = os.path.join(output_dir, excel_filename)
        
        # Write to O. C5 Template Excel
        template_path = str(WORKSPACE_DIR / "O. C5 Template.xlsx")
        import openpyxl
        try:
            wb = openpyxl.load_workbook(template_path)
            
            # 1. Fill Item Sheet
            sheet_item = wb['Item']
            # Delete sample data rows (from row 2 onwards)
            if sheet_item.max_row > 1:
                sheet_item.delete_rows(2, sheet_item.max_row - 1)
                
            headers_item = {cell.value: cell.column for cell in sheet_item[1]}
            
            new_row_idx = 2
            for d in all_dishes:
                mapping = {
                    'OFD': d['ofd'],
                    'Outlet Name': d['nama_panjang'],
                    'Outlet Short Name': d['nama_pendek'],
                    'Outlet Link': d['link_outlet'],
                    'SID': d['store_id'],
                    'Category ID': d['cat_id'],
                    'Category': d['nama_kategori'],
                    'Item ID': d['dish_id'],
                    'Item': d['nama_item'],
                    'Photo Link': d['link_foto'],
                    'Description': d['deskripsi_item'],
                    'Keyword': d['keyword'],
                    'Total Sold': d['jumlah_terjual'],
                    'Total Modifier Group': d['jumlah_modifier_group'],
                    'Total Modifier': d['jumlah_modifier'],
                    'Availability': d['ketersediaan'],
                    'Current Fake Price (Rp)': d['harga_sebelum_promo'],
                    'Current Real Price (Rp)': d['harga_setelah_promo'],
                    'Current Slash Price (%)': d['discount_pct_val'],
                    'Current Slash Price (Rp)': d['slash_price_rp']
                }
                
                for key, val in mapping.items():
                    if key in headers_item:
                        col_idx = headers_item[key]
                        sheet_item.cell(row=new_row_idx, column=col_idx, value=val)
                new_row_idx += 1
                        
            # 2. Fill Modifier Sheet
            sheet_mod = wb['Modifier']
            # Delete sample data rows (from row 2 onwards)
            if sheet_mod.max_row > 1:
                sheet_mod.delete_rows(2, sheet_mod.max_row - 1)
                
            headers_mod = {cell.value: cell.column for cell in sheet_mod[1]}
            
            new_row_idx = 2
            for m in modifier_rows:
                mapping_mod = {
                    'OFD': m['ofd'],
                    'Outlet Name': m['outlet_name'],
                    'Outlet Short Name': m['outlet_short_name'],
                    'Outlet Link': m['outlet_link'],
                    'SID': m['sid'],
                    'Item': m['item'],
                    'Modifier Group ID': m['modifier_group_id'],
                    'Modifier Group': m['modifier_group'],
                    'Modifier ID': m['modifier_id'],
                    'Modifier': m['modifier'],
                    'Min': m['min'],
                    'Max': m['max'],
                    'Availability': m['availability'],
                    'Current Price (Rp)': m['current_price']
                }
                
                for key, val in mapping_mod.items():
                    if key in headers_mod:
                        col_idx = headers_mod[key]
                        sheet_mod.cell(row=new_row_idx, column=col_idx, value=val)
                new_row_idx += 1
                        
            wb.save(excel_path)
            print(f"[+] Berhasil menyimpan file catalog menggunakan template O.C5 ke: {excel_path}")
        except Exception as ex_err:
            print(f"[-] Gagal menulis ke template O.C5: {ex_err}. Fallback ke Excel biasa.")
            # Fallback to standard excel writer if template fails
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df_items.to_excel(writer, sheet_name='Item', index=False)
                df_mods.to_excel(writer, sheet_name='Modifier', index=False)
            
        return True, {
            'items_csv': None,
            'mods_csv': None,
            'excel': excel_path,
            'items_count': len(df_items),
            'mods_count': len(df_mods)
        }
    except Exception as e:
        return False, f"Error selama ekstraksi menu: {e}"
