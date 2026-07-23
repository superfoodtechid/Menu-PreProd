import os
import json
import requests
import openpyxl
import urllib.request
import csv
import re
import codecs
from datetime import datetime, timedelta
from urllib.request import urlopen
from urllib.parse import urlparse, urlencode, urlunparse, parse_qsl
from dotenv import load_dotenv, set_key
try:
    from filelock import FileLock as _FileLock
except ImportError:
    import contextlib
    class _FileLock:
        def __init__(self, path, timeout=-1): pass
        def __enter__(self): return self
        def __exit__(self, *a): pass
import time
from playwright.sync_api import sync_playwright

import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parent.parent))
from discord_notifier import send_discord_error

from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn
from rich.status import Status
from rich.text import Text
from rich.columns import Columns
from rich.theme import Theme

load_dotenv(override=True)

# Initialize Rich Console
custom_theme = Theme({
    "info": "cyan",
    "warning": "yellow",
    "error": "bold red",
    "success": "bold green",
    "highlight": "magenta"
})
console = Console(theme=custom_theme)
START_TIME_TOTAL = time.time()

# Master credential Google Sheet — source of truth for ALL scrapers
SHEET_PUBLISHED_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQ3tLKBNXDqRgBw0mNhKZFxgvKx-JoiTDzm_s5Ix1cm7O6HCv4IvExOLR2HSRVaXSsx82V348mcr9X4/pub?gid=0&single=true&output=csv"


def safe_goto_with_retry(page, url, wait_until="domcontentloaded", timeout=30000, max_attempts=3):
    """
    Mekanisme navigasi aman dengan reload 2x (total 3 attempt) jika terjadi Timeout Error
    saat mengakses halaman portal GoFood.
    """
    last_err = None
    for attempt in range(1, max_attempts + 1):
        try:
            return page.goto(url, wait_until=wait_until, timeout=timeout)
        except Exception as e:
            last_err = e
            console.print(f"   ⚠️ [Timeout/Error] Navigasi ke {url} gagal pada percobaan {attempt}/{max_attempts}: {e}")
            if attempt < max_attempts:
                console.print(f"   🔄 [Reload Mekanisme 2x] Melakukan reload / re-try navigasi ke {url} (Coba {attempt}/{max_attempts - 1})...")
                time.sleep(2.0)
                try:
                    if page.url and page.url != "about:blank":
                        page.reload(wait_until=wait_until, timeout=timeout)
                        console.print(f"   ✅ Reload berhasil untuk {url}")
                        return True
                except Exception as reload_err:
                    console.print(f"   ⚠️ Reload gagal ({reload_err}), mencoba page.goto ulang...")
    if last_err:
        raise last_err
    return False


def to_csv_url(url):
    if 'pubhtml' in url:
        url = url.replace('/pubhtml?', '/pub?')
    if 'output=csv' not in url:
        if '?' in url:
            url += '&output=csv'
        else:
            url += '?output=csv'
    return url


GLOBAL_OUTPUT_DIR = None

# ── Session JSON Cache ────────────────────────────────────────────────

def _session_cache_path(identifier: str) -> str:
    """Return path to session JSON cache file for a given email/phone identifier."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    sessions_dir = os.path.join(base_dir, 'sessions')
    os.makedirs(sessions_dir, exist_ok=True)
    safe_id = re.sub(r'[^a-zA-Z0-9@._-]', '_', identifier)
    return os.path.join(sessions_dir, f"{safe_id}.json")


def _load_session_token(identifier: str) -> str:
    """Load token from session JSON cache. Returns empty string if not found or expired."""
    path = _session_cache_path(identifier)
    if not os.path.exists(path):
        return ''
    try:
        with open(path, 'r') as f:
            data = json.load(f)
        token = data.get('token', '')
        # Opsional: cek TTL jika ada (dalam detik)
        saved_at = data.get('saved_at', 0)
        ttl = data.get('ttl', 0)  # 0 = tidak ada expiry
        if ttl > 0 and (time.time() - saved_at) > ttl:
            console.print(f"[dim]Session cache untuk '{identifier}' sudah kedaluwarsa.[/dim]")
            return ''
        return token
    except Exception as e:
        console.print(f"[dim]Gagal membaca session cache '{identifier}': {e}[/dim]")
        return ''


def _save_session_token(identifier: str, token: str, meta: dict = None):
    """Save token to session JSON cache."""
    path = _session_cache_path(identifier)
    data = {
        'token': token,
        'saved_at': time.time(),
        'ttl': 0,  # 0 = tidak ada expiry
    }
    if meta:
        data.update(meta)
    try:
        lock_path = path + '.lock'
        lock = _FileLock(lock_path, timeout=10)
        with lock:
            with open(path, 'w') as f:
                json.dump(data, f, indent=2)
    except Exception as e:
        console.print(f"[dim]Gagal menyimpan session cache '{identifier}': {e}[/dim]")




def fetch_csv_rows(url):
    with urllib.request.urlopen(url) as resp:
        raw = resp.read()
    text = raw.decode('utf-8', errors='replace')
    return list(csv.reader(text.splitlines()))


def fetch_gofood_accounts_from_sheet(task="2"):
    """
    Mengambil daftar akun GoFood dari master Google Sheet yang sama
    dengan yang digunakan Grab & Shopee di cli.py.
    Filter: Aplikasi=GoFood (case-insensitive), Status=Live.
    Mengembalikan list of dict:
      {
        'phone'     : str,   # kolom AA (index 26) — nomor HP / email login
        'nama_outlet': str,
        'cabang'    : str,
        'store_id'  : str,   # kolom "Store ID" / "Merchant ID"
      }
    """
    url = SHEET_PUBLISHED_URL
    if task == "1":
        url = "https://docs.google.com/spreadsheets/d/14eCb8DAEXhmbYj9MFj2KzC7AhkulbCbSNPltN2m-go0/export?format=csv&gid=880434015"
    
    url += f"&t={int(time.time())}"

    try:
        import io
        import requests as _req
        resp = _req.get(url, timeout=15)
        resp.raise_for_status()
        reader_rows = list(csv.reader(resp.text.splitlines()))
    except Exception as e:
        console.print(f"[error]❌ Gagal mengambil Google Sheet: {e}[/error]")
        return []

    if not reader_rows:
        return []

    header = [str(h).strip().lower() for h in reader_rows[0]]

    # Cari indeks kolom dinamis
    def col_idx(names):
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    accounts = []
    if task == "1" or task == "3":
        # Jika task 3, ambil sheet Agency untuk mapping BD -> Phone
        bd_to_phone = {}
        if task == "3":
            try:
                creds_url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRYSUnKOqk29LCktTxdb0wPLbWMbRaWRP3eC_UA4AwYod1FW6zDMhtLMC5ghIvot2B8upCDfBsn-TCP/pub?gid=565510790&single=true&output=csv"
                import requests as _req
                c_resp = _req.get(creds_url, timeout=15)
                c_resp.raise_for_status()
                creds_rows = list(csv.reader(c_resp.text.splitlines()))
                if creds_rows and len(creds_rows) > 1:
                    c_header = [str(h).strip().lower() for h in creds_rows[0]]
                    bd_idx = c_header.index('bd') if 'bd' in c_header else -1
                    phone_idx = c_header.index('phone') if 'phone' in c_header else -1
                    if bd_idx != -1 and phone_idx != -1:
                        for cr in creds_rows[1:]:
                            if len(cr) > bd_idx and len(cr) > phone_idx:
                                bd_val = str(cr[bd_idx]).strip().lower()
                                phone_val = str(cr[phone_idx]).strip()
                                if bd_val and phone_val:
                                    bd_to_phone[bd_val] = phone_val
            except Exception as e:
                console.print(f"[warning]⚠️ Gagal mengambil credentials Agency: {e}[/warning]")

        # Parsing format sheet Baseline
        idx_aplikasi   = col_idx(['aplikasi'])
        idx_outlet     = col_idx(['nama outlet'])
        idx_email_fm1 = col_idx(['email foodmaster1'])
        idx_email_fm2 = col_idx(['email foodmaster2'])
        idx_bd         = col_idx(['bd'])

        for row in reader_rows[1:]:
            if idx_aplikasi is None or len(row) <= idx_aplikasi:
                continue
            aplikasi = str(row[idx_aplikasi]).strip().lower()
            if 'gofood' not in aplikasi:
                continue

            nama = str(row[idx_outlet]).strip() if idx_outlet is not None and len(row) > idx_outlet else ''
            
            # Jika task 3, gunakan kredensial dari Agency Sheet (berdasarkan BD)
            if task == "3":
                bd_name = str(row[idx_bd]).strip().lower() if idx_bd is not None and len(row) > idx_bd else ''
                phone = bd_to_phone.get(bd_name, "")
                # Jika tidak ada di mapping, gunakan phone default allvbadmin atau lewati
                if not phone:
                    phone = bd_to_phone.get("all", "") # fallback ke BD 'all' jika ada
                
                if not phone:
                    continue # Skip if no phone found for agency
                    
                primary_email = ""
                emails = []
            else:
                # Task 1 (Baseline standard)
                phone = "-"
                # Ambil Email FoodMaster2 sebagai sekunder
                email_fm2 = ""
                if idx_email_fm2 is not None and len(row) > idx_email_fm2:
                    email_fm2 = str(row[idx_email_fm2]).strip()
                
                # Email FoodMaster1 sebagai prioritas utama
                email_fm1 = ""
                if idx_email_fm1 is not None and len(row) > idx_email_fm1:
                    email_fm1 = str(row[idx_email_fm1]).strip()
    
                emails = []
                if email_fm1 and email_fm1 != "-":
                    emails.append(email_fm1)
                if email_fm2 and email_fm2 != "-" and email_fm2 != email_fm1:
                    emails.append(email_fm2)
                    
                primary_email = emails[0] if emails else ""

            cabang = ""
            store_id = ""

            accounts.append({
                'phone'      : phone,
                'email'      : primary_email,
                'emails'     : emails,
                'nama_outlet': nama,
                'cabang'     : cabang,
                'store_id'   : store_id,
            })
    else:
        # Parsing format sheet Live/Weekly (Default)
        idx_aplikasi  = col_idx(['aplikasi'])
        idx_status    = col_idx(['status'])
        idx_outlet    = col_idx(['nama outlet'])
        idx_cabang    = col_idx(['cabang'])
        idx_store     = col_idx(['store id', 'store_id', 'merchant id'])
        idx_phone     = 26  # Kolom AA (0-indexed)
        idx_email_fm  = col_idx(['email foodmaster'])
        idx_email_duck= col_idx(['email duck'])

        for row in reader_rows[1:]:
            if len(row) <= idx_phone:
                continue

            aplikasi = str(row[idx_aplikasi]).strip().lower() if idx_aplikasi is not None and len(row) > idx_aplikasi else ''
            status   = str(row[idx_status]).strip().lower()   if idx_status is not None and len(row) > idx_status else ''

            if 'gofood' not in aplikasi:
                continue
            if 'live' not in status:
                continue

            email_fm = ""
            if idx_email_fm is not None and len(row) > idx_email_fm:
                email_fm = str(row[idx_email_fm]).strip()
            
            email_duck = ""
            if idx_email_duck is not None and len(row) > idx_email_duck:
                email_duck = str(row[idx_email_duck]).strip()

            emails = []
            if email_fm and email_fm != "-":
                emails.append(email_fm)
            if email_duck and email_duck != "-" and email_duck != email_fm:
                emails.append(email_duck)
                
            if not emails and len(row) > 24:
                fallback_email = str(row[24]).strip()
                if fallback_email and fallback_email != "-":
                    emails.append(fallback_email)
                    
            primary_email = emails[0] if emails else ""

            phone     = str(row[idx_phone]).strip()
            nama      = str(row[idx_outlet]).strip()   if idx_outlet is not None and len(row) > idx_outlet else ''
            cabang    = str(row[idx_cabang]).strip()   if idx_cabang is not None and len(row) > idx_cabang else ''
            store_id  = str(row[idx_store]).strip()    if idx_store is not None  and len(row) > idx_store  else ''

            if not phone and not primary_email:
                continue

            accounts.append({
                'phone'      : phone,
                'email'      : primary_email,
                'emails'     : emails,
                'nama_outlet': nama,
                'cabang'     : cabang,
                'store_id'   : store_id,
            })

    return accounts


def digits(s):
    return re.sub(r"\D+", "", s or "")


def normalize_phone(s):
    """Normalize phone number: strip all non-digits, remove leading 62 if present.
    If the string is an email, return it as a clean lowercase string.
    """
    if s and '@' in s:
        return s.strip().lower()
    d = digits(s)
    # if starts with 62, remove it to get 85... format
    if d.startswith('62'):
        d = d[2:]
    return d


def normalize_name(s):
    return re.sub(r"\s+", " ", (s or '').replace('\xa0', ' ')).strip().lower()


def build_sheet_mapping(rows):
    # returns dict with two maps: 'by_phone' and 'by_name'
    mapping = {'by_phone': {}, 'by_name': {}}
    if not rows:
        return mapping
    header = rows[0]
    # try to find "Nama Outlet" index
    name_idx = None
    for i, h in enumerate(header):
        if (h or '').strip().lower() == 'nama outlet':
            name_idx = i
            break
    if name_idx is None:
        for i, h in enumerate(header):
            if 'nama outlet' in (h or '').lower():
                name_idx = i
                break
    aa_idx = 26

    # find potential store id / merchant id columns and cabang column
    store_idx_candidates = []
    cabang_idx = None
    aplikasi_idx = None
    status_idx = None
    for i, h in enumerate(header):
        if not h:
            continue
        hl = h.strip().lower()
        if 'cabang' in hl:
            cabang_idx = i
        if 'aplikasi' in hl:
            aplikasi_idx = i
        if hl == 'status' or ' status' in hl or hl.startswith('status'):
            status_idx = i
        # prefer exact "store id" match, then look for "store" or "merchant"
        if hl == 'store id' or hl == 'store_id':
            store_idx_candidates.insert(0, i)
        elif any(x in hl for x in ('store', 'merchant')) and hl not in ['merchant id', 'merchant_id', 'merchant name']:
            store_idx_candidates.append(i)

    for r in rows[1:]:
        if len(r) <= aa_idx:
            continue
        
        # Mendapatkan email dari kolom Y (index 24) dan nomor HP dari kolom AA (index 26)
        email_val = r[24].strip() if len(r) > 24 else ""
        phone_val = r[26].strip() if len(r) > 26 else ""
        
        # Jika kolom Y berisi email valid, gunakan email tersebut sebagai key pencocokan
        key = normalize_phone(email_val) if "@" in email_val else normalize_phone(phone_val)
        if not key:
            continue
        name = ''
        if name_idx is not None and len(r) > name_idx:
            name = r[name_idx].strip()

        cabang_val = ''
        if cabang_idx is not None and len(r) > cabang_idx:
            cabang_val = (r[cabang_idx] or '').strip()

        aplikasi_val = ''
        if aplikasi_idx is not None and len(r) > aplikasi_idx:
            aplikasi_val = (r[aplikasi_idx] or '').strip().lower()
        if aplikasi_idx is not None and aplikasi_val and aplikasi_val != 'gofood':
            continue

        status_val = ''
        if status_idx is not None and len(r) > status_idx:
            status_val = (r[status_idx] or '').strip().lower()
        if status_idx is not None and status_val and status_val != 'live':
            continue

        store_entries = []
        seen_ids = set()
        for si in store_idx_candidates:
            if len(r) > si:
                v = (r[si] or '').strip()
                if v:
                    parts = re.split(r"[,;/\\|]", v)
                    for p in parts:
                        sp = p.strip()
                        if not sp or sp in seen_ids:
                            continue
                        seen_ids.add(sp)
                        store_entries.append({'id': sp, 'cabang': cabang_val})
        # add to phone map
        if key in mapping['by_phone']:
            existing = mapping['by_phone'][key]
            if not existing.get('name') and name:
                existing['name'] = name
            for ent in store_entries:
                # skip if same store id already exists (avoid duplicates)
                if not any(e['id'] == ent['id'] for e in existing.get('store_ids', [])):
                    existing.setdefault('store_ids', []).append(ent)
        else:
            # filter out duplicate store_ids from initial load
            seen_in_row = set()
            unique_entries = []
            for ent in store_entries:
                if ent['id'] not in seen_in_row:
                    seen_in_row.add(ent['id'])
                    unique_entries.append(ent)
            mapping['by_phone'][key] = {'name': name, 'store_ids': unique_entries}

        # also add to by_name map (group by Nama Outlet + phone to handle same-name outlets with different numbers)
        norm_name = normalize_name(name)
        if norm_name:
            # use combination of name and phone as key to treat different phone numbers as separate outlets
            name_phone_key = f"{norm_name}|{key}"
            if name_phone_key in mapping['by_name']:
                exist = mapping['by_name'][name_phone_key]
                for ent in store_entries:
                    # skip if same store id already exists (avoid duplicates)
                    if not any(e['id'] == ent['id'] for e in exist.get('store_ids', [])):
                        exist.setdefault('store_ids', []).append(ent)
            else:
                # filter out duplicate store_ids from initial load
                seen_in_row = set()
                unique_entries = []
                for ent in store_entries:
                    if ent['id'] not in seen_in_row:
                        seen_in_row.add(ent['id'])
                        unique_entries.append(ent)
                mapping['by_name'][name_phone_key] = {'name': name, 'store_ids': unique_entries}
    return mapping


def find_match_in_mapping(mapping, num):
    nd = normalize_phone(num)
    if not nd:
        return None
    # mapping expected to be the full mapping with 'by_phone' key
    phone_map = mapping.get('by_phone') if isinstance(mapping, dict) and 'by_phone' in mapping else mapping
    if nd in phone_map:
        return phone_map[nd]
    for k, v in phone_map.items():
        if k.endswith(nd) or nd.endswith(k):
            return v
    return None


def get_sheet_entry(mapping, num, current_name=None):
    """Return mapping entry dict by phone or by outlet name if phone not found.

    mapping: {'by_phone':..., 'by_name':...}
    num: phone number to match
    current_name: optional existing outlet name (env) to lookup in by_name
    """
    if not mapping:
        return None
    
    nd = normalize_phone(num)
    
    # if we already know the outlet name, prefer phone-based match first (most reliable)
    by_phone = mapping.get('by_phone', {})
    if nd in by_phone:
        return by_phone[nd]
    
    # if outlet name provided, try name|phone combination lookup
    if current_name and nd:
        norm_name = normalize_name(current_name)
        by_name = mapping.get('by_name', {})
        name_phone_key = f"{norm_name}|{nd}"
        if name_phone_key in by_name:
            return by_name[name_phone_key]
    
    # try partial phone match in by_phone map
    for k, v in by_phone.items():
        if k.endswith(nd) or nd.endswith(k):
            return v
    
    return None


def ambil_otp_dari_endpoint(url_dasar, action="getOtp", label_email=None):
    """
    Mengambil OTP terbaru dari endpoint Google Apps Script atau langsung dari Google Sheets CSV.
    """
    if not url_dasar:
        raise ValueError("URL endpoint OTP kosong.")

    # Jika URL mengarah langsung ke Google Sheets CSV
    if "docs.google.com/spreadsheets" in url_dasar:
        try:
            with urlopen(url_dasar, timeout=15) as response:
                content = response.read().decode("utf-8").strip()
                lines = content.splitlines()
                if not lines or len(lines) < 2:
                    return ""
                reader = csv.reader(lines)
                rows = list(reader)
                headers = [h.strip().lower() for h in rows[0]]
                
                otp_idx = -1
                for idx, h in enumerate(headers):
                    if "otp" in h:
                        otp_idx = idx
                        break
                
                if otp_idx == -1:
                    otp_idx = 1 if len(rows[0]) > 1 else 0
                    
                last_row = rows[-1]
                if len(last_row) > otp_idx:
                    return last_row[otp_idx].strip()
                return ""
        except Exception as e:
            console.print(f"[warning]⚠️ Gagal membaca OTP dari Sheets: {e}[/warning]")
            return ""

    parsed = urlparse(url_dasar)
    query_params = dict(parse_qsl(parsed.query))
    query_params["action"] = action
    if label_email:
        query_params["label"] = label_email
    url_final = urlunparse(parsed._replace(query=urlencode(query_params)))

    with urlopen(url_final, timeout=15) as response:
        return response.read().decode("utf-8").strip()


def tunggu_otp_terbaru(url_dasar, action="getOtp", label_email=None, interval_detik=3, otp_awal_override=None, timeout_detik=15):
    """
    Menunggu OTP terbaru yang berbeda dari nilai awal agar tidak memakai OTP sebelumnya.
    otp_awal_override: Jika diisi, gunakan nilai ini sebagai baseline (snapshot sebelum OTP dikirim).
    """
    if otp_awal_override is not None:
        otp_awal = otp_awal_override
    else:
        try:
            otp_awal = ambil_otp_dari_endpoint(url_dasar, action=action, label_email=label_email)
        except Exception:
            otp_awal = ""
    
    batas_waktu = time.time() + timeout_detik
    console.print(f"   [info]🤖 Menunggu OTP baru masuk ke inbox (maksimal {timeout_detik} detik)...[/info]")

    while time.time() < batas_waktu:
        time.sleep(interval_detik)
        try:
            otp_baru = ambil_otp_dari_endpoint(url_dasar, action=action, label_email=label_email)
            if otp_baru and otp_baru != otp_awal:
                return otp_baru
        except Exception:
            pass

    return otp_awal


def login_outlet_gofood_flow(outlet_info):
    """
    Membuka browser Chromium untuk login otomatis 1 outlet.
    Menangkap token, menyimpannya di .env, dan mengembalikannya.
    """
    nama = outlet_info['nama_outlet']
    cabang = outlet_info.get('cabang', '')
    emails_to_try = outlet_info.get('emails', [])
    if not emails_to_try:
        # Fallback to single email
        single_email = outlet_info.get('email', '')
        if single_email:
            emails_to_try = [single_email]
            
    phone = outlet_info.get('phone_raw', '') or outlet_info.get('phone', '')

    label = f"{nama} - {cabang}" if cabang and cabang != 'Tanpa Cabang' else nama

    console.print(f"\n[bold yellow]🔄 Membuka browser untuk login otomatis ke: {label}[/bold yellow]")
    if emails_to_try:
        console.print(f"   📧 Emails: {', '.join(emails_to_try)}")
    if phone:
        console.print(f"   📱 Phone: {phone}")

    use_proxy = os.getenv("USE_PROXY", "false").lower() in ("true", "1", "yes")
    proxy_server = os.getenv("PROXY_SERVER")
    proxy_config = None
    if use_proxy and proxy_server:
        parsed = urlparse(proxy_server)
        if parsed.username and parsed.password:
            server_url = f"{parsed.scheme}://{parsed.hostname}"
            if parsed.port:
                server_url += f":{parsed.port}"
            proxy_config = {"server": server_url, "username": parsed.username, "password": parsed.password}
        else:
            proxy_config = {"server": proxy_server}

    # Cek mode headless dari .env atau config.json
    env_headless = os.getenv("HEADLESS") or os.getenv("HEADLESS_GOFOOD")
    if env_headless is not None:
        headless_mode = env_headless.lower() in ("true", "1", "yes")
    else:
        headless_mode = False
        try:
            import json
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config_path = os.path.join(base_dir, "config.json")
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config_data = json.load(f)
                    headless_mode = bool(config_data.get("headless_gofood", False))
        except Exception:
            pass

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless_mode,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-infobars',
                '--no-sandbox'
            ]
        )
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1366, 'height': 768},
            proxy=proxy_config
        )

        import random
        
        access_token = None

        for email_idx, current_email in enumerate(emails_to_try):
            if access_token:
                break
                
            max_login_attempts = 2
            attempts_made = 0
            
            while attempts_made < max_login_attempts:
                attempt = attempts_made
                attempts_made += 1

                if access_token:
                    break
                    
                page = context.new_page()
                if current_email:
                    console.print(f"\n   ➡️ [Email: {current_email}] Membuka halaman login email langsung... (Percobaan {attempt + 1}/{max_login_attempts})")
                    safe_goto_with_retry(page, "https://portal.gofoodmerchant.co.id/auth/login/email", wait_until="domcontentloaded")
                else:
                    console.print(f"\n   ➡️ Membuka halaman login... (Percobaan {attempt + 1}/{max_login_attempts})")
                    safe_goto_with_retry(page, "https://portal.gofoodmerchant.co.id/auth/login", wait_until="domcontentloaded")

                # Langsung input ke email field, abaikan cookie & pop-up
                time.sleep(1.0)
                
                otp_failed_timeout = False
                is_banned = False

                # Ambil konfigurasi OTP endpoint di awal setiap attempt
                otp_endpoint = os.getenv("OTP_ENDPOINT_URL")
                label_email_cfg = os.getenv("GMAIL_OTP_LABEL", "OTP-GO")
                action_type = "getOtpEmail" if current_email else "getOtp"
                otp_snapshot_awal = ""

                # --- STEP 4: Ketik email secara human-like ---
                if current_email:
                    try:
                        email_input = page.wait_for_selector(
                            'input[type="email"], input[name="email"], input[placeholder*="email" i], input[placeholder*="Email" i], input[type="text"]',
                            timeout=15000
                        )
                        if email_input:
                            email_input.click()
                            time.sleep(0.3)
                            email_input.focus()
                            time.sleep(0.3)
                            for char in current_email:
                                email_input.type(char, delay=0)
                                time.sleep(random.uniform(0.05, 0.15))
                            time.sleep(0.5)

                            submit_btn = page.locator('button:has-text("Lanjut"), button:has-text("Submit"), button:has-text("Masuk"), button[type="submit"]')
                            if submit_btn.count() > 0:
                                submit_btn.first.click()
                            else:
                                email_input.press("Enter")
                            time.sleep(3)

                            # --- Pre-snapshot OTP sebelum tombol OTP diklik ---
                            if otp_endpoint:
                                try:
                                    otp_snapshot_awal = ambil_otp_dari_endpoint(otp_endpoint, action=action_type, label_email=label_email_cfg)
                                    console.print(f"   [info]📸 Snapshot OTP awal: '{otp_snapshot_awal or '(kosong)'}' (sebelum OTP dikirim)[/info]")
                                except Exception:
                                    otp_snapshot_awal = ""

                            # Jika ada halaman pilihan login (password/OTP)
                            try:
                                btn_otp = page.locator('button:has-text("Masuk dengan OTP"), a:has-text("Masuk dengan OTP")').first
                                if btn_otp.count() > 0 and btn_otp.is_visible():
                                    btn_otp.click()
                                    console.print("   [info]✅ Tombol 'Masuk dengan OTP' diklik. OTP sedang dikirim...[/info]")
                                    time.sleep(2)
                            except Exception:
                                pass

                            # --- STEP 5: Automated OTP Polling & Fill ---
                            if otp_endpoint:
                                # 1. Tunggu field OTP muncul — Jika timeout/error, berarti akun kena banned 15 menit
                                try:
                                    console.print("   [info]🤖 Menunggu field OTP muncul...[/info]")
                                    otp_input_selector = 'input[autocomplete="one-time-code"], input[aria-label*="digit" i], div[class*="otp" i] input:not([type="checkbox"]):not([type="radio"]), input[name*="otp" i]:not([type="checkbox"]):not([type="radio"]), input[maxlength="1"]:not([type="checkbox"]):not([type="radio"])'
                                    
                                    otp_appeared = False
                                    start_wait_otp = time.time()
                                    while time.time() - start_wait_otp < 15:
                                        if page.locator(otp_input_selector).count() > 0 and page.locator(otp_input_selector).first.is_visible():
                                            otp_appeared = True
                                            break
                                        
                                        # Fast-fail deteksi Limit/Banned
                                        try:
                                            ban_msg1 = page.locator('text=/terlalu banyak/i')
                                            ban_msg2 = page.locator('text=/coba lagi/i')
                                            ban_msg3 = page.locator('text=/15 menit/i')
                                            if (ban_msg1.count() > 0 and ban_msg1.first.is_visible()) or \
                                               (ban_msg2.count() > 0 and ban_msg2.first.is_visible()) or \
                                               (ban_msg3.count() > 0 and ban_msg3.first.is_visible()):
                                                console.print("   [warning]⚠️ Terdeteksi teks peringatan Limit/Banned. Membatalkan tunggu OTP...[/warning]")
                                                break
                                        except Exception:
                                            pass
                                            
                                        time.sleep(1.0)
                                        
                                    if not otp_appeared:
                                        raise Exception("OTP Field timeout atau akun terindikasi banned")
                                        
                                    time.sleep(1)
                                except Exception as e:
                                    console.print(f"   [warning]⚠️ {e}: Field OTP tidak muncul. Indikasi limit/banned 15 menit untuk email {current_email}. Menghentikan percobaan dan rotasi akun.[/warning]")
                                    is_banned = True
                                    try:
                                        page.close()
                                    except Exception:
                                        pass
                                    break  # Keluar dari loop attempt, langsung rotasi ke email berikutnya

                                # 2. Lakukan polling dan input OTP
                                if not is_banned:
                                    try:
                                        console.print("   [info]🤖 Polling OTP dari Gmail (snapshot awal sudah diambil sebelumnya)...[/info]")
                                        
                                        otp_code = tunggu_otp_terbaru(otp_endpoint, action=action_type, label_email=label_email_cfg, interval_detik=3, otp_awal_override=otp_snapshot_awal, timeout_detik=15)
                                        
                                        if otp_code and not (otp_code.isdigit() and len(otp_code) in (4, 6)):
                                            console.print(f"   [warning]⚠️ OTP dari endpoint bukan format angka valid: {otp_code[:50]}...[/warning]")
                                            otp_code = None
                                            
                                        if otp_code:
                                            console.print(f"   [info]🤖 OTP didapat: {otp_code}. Memasukkan OTP...[/info]")
                                            otp_fields = page.locator(otp_input_selector).all()
                                            if len(otp_fields) > 0:
                                                otp_fields[0].focus()
                                                time.sleep(0.5)
                                                otp_fields[0].type(otp_code, delay=300)
                                                console.print("   [success]✅ OTP berhasil diisi otomatis.[/success]")
                                                
                                                # Coba klik tombol submit/konfirmasi/masuk OTP
                                                time.sleep(1)
                                                submit_otp_btn = page.locator('button:has-text("Masuk"), button:has-text("Konfirmasi"), button:has-text("Verifikasi"), button:has-text("Lanjut"), button[type="submit"]')
                                                clicked = False
                                                for i in range(submit_otp_btn.count()):
                                                    btn = submit_otp_btn.nth(i)
                                                    if btn.is_visible() and btn.is_enabled():
                                                        console.print(f"   [info]🤖 Mengklik tombol OTP: '{btn.text_content().strip()}'[/info]")
                                                        btn.click()
                                                        clicked = True
                                                        break
                                                if not clicked:
                                                    console.print("   [info]🤖 Mengirim Enter sebagai fallback...[/info]")
                                                    page.keyboard.press("Enter")
                                                time.sleep(2)
                                        else:
                                            console.print("   [warning]⚠️ Gagal mendapatkan OTP dalam 15 detik (atau format tidak valid).[/warning]")
                                            if email_idx == len(emails_to_try) - 1 and attempt >= max_login_attempts - 1:
                                                send_discord_error(
                                                    platform="GoFood", 
                                                    merchant=nama, 
                                                    error_type="OTP_TIMEOUT", 
                                                    message=f"Gagal masuk akun ({current_email}). OTP tidak kunjung diterima dalam batas waktu 15 detik.",
                                                    phone=phone
                                                )
                                            otp_failed_timeout = True
                                    except Exception as e:
                                        console.print(f"   [warning]⚠️ Gagal melakukan automasi OTP: {e}.[/warning]")
                                        if email_idx == len(emails_to_try) - 1 and attempt >= max_login_attempts - 1:
                                            send_discord_error(
                                                    platform="GoFood", 
                                                    merchant=nama, 
                                                    error_type="SYSTEM_ERROR", 
                                                    message=f"Gagal melakukan automasi input OTP ({current_email}): {str(e)[:100]}.",
                                                    phone=phone
                                            )
                                        otp_failed_timeout = True
                            else:
                                console.print("   [info]👉 Silakan isi kode OTP secara MANUAL di browser.[/info]")
                    except Exception as e:
                        console.print(f"   [error]⚠️ Gagal ketik email: {e}[/error]")

                # Jika banned (field OTP tidak muncul), sudah di-break di atas, skip token wait
                if is_banned:
                    continue

                # Jika OTP gagal timeout, tunggu 15 detik lalu retry attempt berikutnya
                if otp_failed_timeout:
                    if attempt < max_login_attempts - 1:
                        console.print("   [warning]⚠️ Menutup halaman dan menunggu 15 detik sebelum mengulang login (attempt ke-2)...[/warning]")
                        try:
                            page.close()
                        except Exception:
                            pass
                        time.sleep(15)
                        continue
                    else:
                        console.print(f"   [warning]⚠️ Melewati batas percobaan login untuk {current_email}. Rotasi ke email berikutnya/gagal.[/warning]")
                        try:
                            page.close()
                        except Exception:
                            pass
                        continue

                # --- Tunggu access_token muncul di cookies (max 15 detik) ---
                attempt_token = None
                start_time = time.time()
                try:
                    while True:
                        if page.is_closed():
                            console.print("[warning]⚠️ Browser ditutup sebelum login selesai.[/warning]")
                            if email_idx == len(emails_to_try) - 1 and attempt >= max_login_attempts - 1:
                                send_discord_error("GoFood", nama, "SYSTEM_ERROR", "Proses login terganggu karena browser tertutup secara tiba-tiba atau kehilangan koneksi di tengah jalan.", phone)
                            break

                        cookies = context.cookies()
                        for cookie in cookies:
                            if cookie['name'] == 'access_token':
                                attempt_token = cookie['value']
                                break

                        if attempt_token:
                            access_token = attempt_token
                            break

                        # Deteksi akun baru yang butuh verifikasi email
                        try:
                            error_msg1 = page.locator('text=/Email belum diverifikasi/i')
                            error_msg2 = page.locator('text=/silahkan login ulang/i')
                            if (error_msg1.count() > 0 and error_msg1.first.is_visible()) or \
                               (error_msg2.count() > 0 and error_msg2.first.is_visible()):
                                console.print("   [warning]⚠️ Terdeteksi akun baru: 'Email belum diverifikasi'. Mempercepat percobaan ulang...[/warning]")
                                max_login_attempts = 3
                                time.sleep(2.0)
                                break
                        except Exception:
                            pass

                        # Deteksi OTP Salah / Kadaluarsa
                        try:
                            otp_err1 = page.locator('text=/kode salah/i')
                            otp_err2 = page.locator('text=/tidak valid/i')
                            otp_err3 = page.locator('text=/kadaluarsa/i')
                            if (otp_err1.count() > 0 and otp_err1.first.is_visible()) or \
                               (otp_err2.count() > 0 and otp_err2.first.is_visible()) or \
                               (otp_err3.count() > 0 and otp_err3.first.is_visible()):
                                console.print("   [warning]⚠️ Terdeteksi pesan 'OTP Salah/Tidak Valid'. Mempercepat percobaan ulang...[/warning]")
                                break
                        except Exception:
                            pass

                        # Deteksi Ban 15 Menit setelah submit OTP
                        try:
                            ban_err1 = page.locator('text=/terlalu banyak/i')
                            ban_err2 = page.locator('text=/coba lagi/i')
                            ban_err3 = page.locator('text=/15 menit/i')
                            if (ban_err1.count() > 0 and ban_err1.first.is_visible()) or \
                               (ban_err2.count() > 0 and ban_err2.first.is_visible()) or \
                               (ban_err3.count() > 0 and ban_err3.first.is_visible()):
                                console.print("   [warning]⚠️ Terdeteksi teks Limit/Banned. Membatalkan tunggu token...[/warning]")
                                is_banned = True
                                break
                        except Exception:
                            pass

                        time.sleep(1.0)

                        if time.time() - start_time > 5:
                            # Fallback URL Check: Jika setelah 5 detik masih stuck di URL login, anggap butuh retry/verifikasi
                            try:
                                if "/auth/login" in page.url:
                                    console.print("   [warning]⚠️ (Fallback) Timeout 5 detik: URL masih stuck di halaman login. Mempercepat percobaan ulang...[/warning]")
                                    # Menambah kompensasi karena bisa jadi ini peringatan verifikasi yang terlewat dari deteksi teks
                                    max_login_attempts = 3
                                    break
                            except Exception:
                                pass

                        if time.time() - start_time > 15:
                            console.print("[warning]⚠️ Timeout 15 detik menunggu access_token.[/warning]")
                            break

                except KeyboardInterrupt:
                    console.print("\n[warning]⚠️ Dibatalkan oleh pengguna.[/warning]")
                    break
                except Exception as e:
                    console.print(f"[error]❌ Error: {e}[/error]")

                if not access_token:
                    try:
                        page.close()
                    except Exception:
                        pass

                    if is_banned:
                        console.print(f"   [warning]⚠️ Akun {current_email} terindikasi Limit/Banned. Langsung rotasi ke email berikutnya.[/warning]")
                        break  # Keluar dari loop attempt, pindah ke email berikutnya

                    if attempt < max_login_attempts - 1:
                        # Percobaan 1 gagal → ulangi login dengan email yang sama
                        console.print(f"   [warning]⚠️ Token tidak ditemukan. Kembali ke login page dengan email yang sama ({current_email})...[/warning]")
                        continue
                    else:
                        # Percobaan ke-2 habis → rotasi ke email berikutnya
                        console.print(f"   [warning]⚠️ Token tidak ditemukan setelah {max_login_attempts} percobaan untuk {current_email}. Rotasi ke email berikutnya...[/warning]")

        try:
            browser.close()
        except Exception:
            pass

        return access_token


def ambil_data_dashboard():
    """Mengecek validitas token sesi dengan mengambil data user."""
    session = requests.Session()
    use_proxy = os.getenv("USE_PROXY", "false").lower() in ("true", "1", "yes")
    proxy_server = os.getenv("PROXY_SERVER")
    if use_proxy and proxy_server:
        session.proxies = {
            "http": proxy_server,
            "https": proxy_server
        }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Authentication-Type': 'go-id',
        'Authorization': f"Bearer {os.getenv('BEARER_TOKEN')}",
        'Origin': 'https://portal.gofoodmerchant.co.id',
        'Referer': 'https://portal.gofoodmerchant.co.id/',
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36'
    }

    url = "https://api.gobiz.co.id/v1/users/me"
    
    try:
        if not headers['Authorization'] or 'None' in headers['Authorization']:
             console.print("[error]❌ Error: BEARER_TOKEN tidak ditemukan di file .env.[/error]")
             console.print("[info]   Silakan jalankan 'python otp_receiver.py' untuk login dan mendapatkan token.[/info]")
             send_discord_error("GoFood", "Global", "MISSING_CREDENTIALS", "Proses terhenti. Sesi Token GoFood hilang/kedaluwarsa dan belum ada token baru di file .env.", "")
             return False

        with console.status("[bold blue]Mengecek sesi GoBiz...", spinner="dots"):
            response = session.get(url, headers=headers)
            response.raise_for_status()
            user_name = response.json().get('user', {}).get('full_name', 'Tidak Dikenal')
        
        console.print(f"[success]✅ Sesi aktif untuk user: [bold]{user_name}[/bold][/success]")
        return True
    except requests.exceptions.HTTPError as err:
        if err.response.status_code in [401, 403]:
            console.print("[error]❌ Gagal: Sesi kedaluwarsa atau tidak valid.[/error]")
            console.print("[info]   Silakan jalankan 'python otp_receiver.py' untuk memperbarui sesi Anda.[/info]")
            send_discord_error("GoFood", "Global", "MISSING_CREDENTIALS", "Sesi kedaluwarsa atau tidak valid. Membutuhkan re-login manual untuk memperbarui token.", "")
        else:
            console.print(f"[error]❌ Gagal mengakses API, status code: {err.response.status_code}[/error]")
            console.print(f"[dim]Response: {err.response.text}[/dim]")
        return False
    except requests.exceptions.RequestException as e:
        console.print(f"[error]❌ Terjadi error koneksi: {e}[/error]")
        return False


def parse_tanggal_input(tanggal_str):
    """Mengubah string tanggal (DD-MM-YYYY atau YYYY-MM-DD) menjadi datetime."""
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(tanggal_str.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"Format tanggal tidak valid: '{tanggal_str}'")


def minta_range_tanggal_custom():
    """
    Meminta user menentukan range tanggal custom.
    Mengembalikan (start_date, end_date) atau (None, None) jika memakai default.
    """
    pilih = input("\nGunakan range tanggal custom? (y/n): ").strip().lower()
    if pilih != 'y':
        return None, None

    while True:
        try:
            start_str = input("Masukkan tanggal mulai (format DD-MM-YYYY atau YYYY-MM-DD): ").strip()
            end_str = input("Masukkan tanggal akhir (format DD-MM-YYYY atau YYYY-MM-DD): ").strip()
            start_date = parse_tanggal_input(start_str)
            end_date = parse_tanggal_input(end_str)
            if start_date > end_date:
                print("⚠️ Tanggal mulai tidak boleh lebih besar dari tanggal akhir.")
                continue
            return start_date, end_date
        except ValueError:
            print("⚠️ Format tanggal tidak valid. Gunakan DD-MM-YYYY atau YYYY-MM-DD.")


def ambil_data_analytics(write_header=True, start_date=None, end_date=None, return_data=False,
                          token=None, store_id=None, nama_outlet=None, phone=None, cabang=None):
    """
    Mengambil data analytics GoFood.
    Parameter token, store_id, nama_outlet, phone, cabang di-pass secara eksplisit
    agar tidak bergantung pada os.environ global (aman untuk concurrent execution).
    Jika tidak di-pass, fallback ke os.getenv() untuk kompatibilitas backward.
    """
    # Resolve context — gunakan parameter eksplisit jika ada, fallback ke env
    _token     = token     or os.getenv('BEARER_TOKEN', '')
    _store_id  = store_id  or os.getenv('ACTIVE_STORE_ID', '')
    _phone     = phone     or os.getenv('ACTIVE_NOMOR_HP', '')
    _outlet    = nama_outlet or os.getenv('ACTIVE_NAMA_OUTLET', '')
    _cabang    = cabang    or os.getenv('ACTIVE_CABANG', '')
    session = requests.Session()
    use_proxy = os.getenv("USE_PROXY", "false").lower() in ("true", "1", "yes")
    proxy_server = os.getenv("PROXY_SERVER")
    if use_proxy and proxy_server:
        session.proxies = {
            "http": proxy_server,
            "https": proxy_server
        }
    START_TIME_STORE = time.time()

    # --- LOGIKA TANGGAL ---
    custom_mode = start_date is not None and end_date is not None
    if not custom_mode:
        # Default: 3 bulan penuh ke belakang (per bulan)
        now = datetime.now()
        first_day_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = first_day_this_month - timedelta(milliseconds=1)

        start_month = first_day_this_month.month - 3
        start_year = first_day_this_month.year
        if start_month <= 0:
            start_month += 12
            start_year -= 1
        start_date = first_day_this_month.replace(year=start_year, month=start_month, day=1)

        period_label_title = "Bulan"
        period_iter = []
        curr_month = start_date
        while curr_month <= end_date:
            label = curr_month.strftime('%B %Y')
            period_iter.append((curr_month.strftime('%Y-%m'), label))
            curr_month = (curr_month.replace(day=1) + timedelta(days=32)).replace(day=1)
        excel_filename = 'revenue_3_bulan.xlsx'
    else:
        # Custom range: per hari
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999000)

        period_label_title = "Hari"
        period_iter = []
        curr_day = start_date
        while curr_day <= end_date:
            label = curr_day.strftime('%d %b %Y')
            period_iter.append((curr_day.strftime('%Y-%m-%d'), label))
            curr_day = curr_day + timedelta(days=1)
        excel_filename = f"revenue_{start_date.strftime('%Y-%m-%d')}_sampai_{end_date.strftime('%Y-%m-%d')}.xlsx"

    global GLOBAL_OUTPUT_DIR
    if GLOBAL_OUTPUT_DIR:
        os.makedirs(GLOBAL_OUTPUT_DIR, exist_ok=True)
        excel_filename = os.path.join(GLOBAL_OUTPUT_DIR, excel_filename)

    # Konversi ke Epoch timestamp dalam milidetik untuk header
    range_from_ms = str(int(start_date.timestamp() * 1000))
    range_to_ms = str(int(end_date.timestamp() * 1000))

    # --- 1. REQUEST DATA GROSS REVENUE ---
    # include merchant_ids param in Referer if available
    active_store = _store_id.strip() if _store_id else ''
    merchant_q = f"&merchant_ids={active_store}" if active_store else ''

    headers = {
        'Accept': '*/*',
        'Authentication-Type': 'go-id',
        'Authorization': f"Bearer {_token}",
        'Content-Type': 'application/json, application/x-ndjson',
        'Origin': 'https://portal.gofoodmerchant.co.id',
        'Referer': f"https://portal.gofoodmerchant.co.id/analytics/sales-gofood?date_range=custom&end_date={end_date.strftime('%Y-%m-%dT%H%%3A%M%%3A%S.999Z')}&start_date={start_date.strftime('%Y-%m-%dT%H%%3A%M%%3A%S.000Z')}{merchant_q}",
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36',
        'x-comp-range-from': range_from_ms,
        'x-comp-range-to': range_to_ms,
        'x-custom-interval': '1d',
        'x-custom-merchant-id': active_store,
        'x-dashboard-id': '107',
        'x-grafana-org-id': '1',
        'x-panel-id': '2',
        'x-range-from': range_from_ms,
        'x-range-to': range_to_ms,
        'x-ref-ids': 'total_gmv_topline_amount;prev_total_gmv_topline_amount',
        'x-setting-interval': '1d'
    }

    url = "https://portal.gofoodmerchant.co.id/analytics-backend/api/datasources/proxy/63/_msearch?max_concurrent_shard_requests=5"
    
    with console.status("[bold cyan]Mengambil data Revenue...", spinner="bouncingBar"):
        response = session.post(url, headers=headers, data="")
    
    data_revenue = {}

    if response.status_code == 200:
        msg = f"Berhasil mengambil data Revenue dari {start_date.strftime('%d %b %Y')} hingga {end_date.strftime('%d %b %Y')}"
        console.print(f"[success]✅ {msg}[/success]")
        data_revenue = response.json()
    else:
        console.print(f"[error]❌ Gagal mengakses Revenue, status code: {response.status_code}[/error]")
        return
        
    # --- 2. REQUEST DATA JUMLAH PESANAN (ORDERS) ---
    url_orders = "https://portal.gofoodmerchant.co.id/analytics-backend/api/datasources/proxy/46/_msearch?max_concurrent_shard_requests=5"
    
    # Menentukan indeks Elasticsearch secara dinamis untuk rentang yang dipilih
    indices = []
    curr_idx = start_date
    while curr_idx <= end_date:
        indices.append(f"analytic_detail_gofood_booking_v1_{curr_idx.strftime('%Y-%m')}")
        curr_idx = (curr_idx.replace(day=1) + timedelta(days=32)).replace(day=1)
    indices_json = json.dumps(indices)

    # Build merchant_id filter for Elasticsearch query string: merchant_id:("VALUE") format
    if active_store:
        merchant_filter = f" AND merchant_id:(\\\"{ active_store}\\\")"
    else:
        merchant_filter = " AND merchant_id:__empty__"

    payload_orders = (
        f'{{"search_type":"query_then_fetch","ignore_unavailable":true,"index":{indices_json}}}\n'
        f'{{"size":0,"query":{{"bool":{{"filter":[{{"query_string":{{"analyze_wildcard":true,"query":"time:>={range_from_ms} AND time:<={range_to_ms}{merchant_filter} AND NOT id:FP* AND _exists_:data.status AND data.status:COMPLETED"}}]}}}},"aggs":{{"2":{{"date_histogram":{{"field":"time","min_doc_count":0,"extended_bounds":{{"min":{range_from_ms},"max":{range_to_ms}}},"format":"epoch_millis","time_zone":"Asia/Jakarta","interval":"1d"}},"aggs":{{}}}}}}}}\n'
        f'{{"search_type":"query_then_fetch","ignore_unavailable":true,"index":{indices_json}}}\n'
        f'{{"size":0,"query":{{"bool":{{"filter":[{{"query_string":{{"analyze_wildcard":true,"query":"time:>={range_from_ms} AND time:<={range_to_ms}{merchant_filter} AND NOT id:FP*"}}]}}}},"aggs":{{"2":{{"date_histogram":{{"field":"time","min_doc_count":0,"extended_bounds":{{"min":{range_from_ms},"max":{range_to_ms}}},"format":"epoch_millis","time_zone":"Asia/Jakarta","interval":"1d"}},"aggs":{{}}}}}}}}\n'
    )

    headers_orders = headers.copy()
    headers_orders.update({
        'Content-Type': 'application/x-ndjson',
        'Referer': f"https://portal.gofoodmerchant.co.id/analytics-backend/d/npZdujrIz/operationals?orgId=1&kiosk=gobiz&from={range_from_ms}&to={range_to_ms}&var-interval=1d&var-merchant_id={active_store}&locale=id&country=ID&var-ad_slot=GOFOOD_CPC_FUNGIBLE_AD&var-ad_slot=GOFOOD_HOME_BANNER_TOP&var-ad_slot=GOFOOD_HOME_MAST_HEAD_TOP&var-ad_slot=GOFOOD_TEXT_SEARCH_TILE{merchant_q}",
        'x-dashboard-id': '83',
        'x-panel-id': '38',
        'x-ref-ids': 'A;B',
        'x-comp-range-offset': '30d',
        'x-custom-interval': '1d',
        'x-setting-interval': '2h'
    })

    with console.status("[bold cyan]Mengambil data Pesanan (Orders)...", spinner="bouncingBar"):
        response_orders = session.post(url_orders, headers=headers_orders, data=payload_orders)
    
    data_orders = {}

    if response_orders.status_code == 200:
        console.print(f"[success]✅ Berhasil mengambil data Orders.[/success]")
        data_orders = response_orders.json()
    else:
        console.print(f"[error]❌ Gagal mengakses Orders, status code: {response_orders.status_code}[/error]")

    # --- 3. REQUEST DATA OMZET BERSIH ---
    url_net = "https://portal.gofoodmerchant.co.id/analytics-backend/api/datasources/proxy/63/_msearch?max_concurrent_shard_requests=5"

    headers_net = headers.copy()
    headers_net.update({
        'x-panel-id': '7',
        'x-ref-ids': 'total_gmv_bottomline_amount;prev_total_gmv_bottomline_amount'
    })

    with console.status("[bold cyan]Mengambil data Omzet Bersih...", spinner="bouncingBar"):
        response_net = session.post(url_net, headers=headers_net, data="")
    
    data_net = {}

    if response_net.status_code == 200:
        console.print(f"[success]✅ Berhasil mengambil data Omzet Bersih.[/success]")
        data_net = response_net.json()
    else:
        console.print(f"[error]❌ Gagal mengakses Omzet Bersih, status code: {response_net.status_code}[/error]")

    # --- 3.5 REQUEST DATA KOMISI ---
    url_komisi = "https://portal.gofoodmerchant.co.id/analytics-backend/api/datasources/proxy/63/_msearch?max_concurrent_shard_requests=5"

    headers_komisi = headers.copy()
    headers_komisi.update({
        'x-panel-id': '4',
        'x-ref-ids': 'total_commission_amount'
    })

    with console.status("[bold cyan]Mengambil data Komisi...", spinner="bouncingBar"):
        response_komisi = session.post(url_komisi, headers=headers_komisi, data="")
    
    data_komisi = {}

    if response_komisi.status_code == 200:
        console.print(f"[success]✅ Berhasil mengambil data Komisi.[/success]")
        data_komisi = response_komisi.json()
    else:
        console.print(f"[error]❌ Gagal mengakses Komisi, status code: {response_komisi.status_code}[/error]")

    # --- 3.6 REQUEST DATA IKLAN & DISKON ---
    url_iklan = "https://portal.gofoodmerchant.co.id/analytics-backend/api/datasources/proxy/63/_msearch?max_concurrent_shard_requests=5"

    headers_iklan = headers.copy()
    headers_iklan.update({
        'x-panel-id': '5',
        'x-ref-ids': 'total_ad_promo_burn_amount'
    })

    with console.status("[bold cyan]Mengambil data Iklan & Diskon...", spinner="bouncingBar"):
        response_iklan = session.post(url_iklan, headers=headers_iklan, data="")
    
    data_iklan = {}

    if response_iklan.status_code == 200:
        console.print(f"[success]✅ Berhasil mengambil data Iklan & Diskon.[/success]")
        data_iklan = response_iklan.json()
    else:
        console.print(f"[error]❌ Gagal mengakses Iklan & Diskon, status code: {response_iklan.status_code}[/error]")

    # --- 3.7 REQUEST DATA ORDER BATAL (6-QUERY PAYLOAD) ---
    url_batal = "https://portal.gofoodmerchant.co.id/analytics-backend/api/datasources/proxy/46/_msearch?max_concurrent_shard_requests=5"
    
    cancel_reasons = (
        "MERCHANT_ACCEPTANCE_TIMEOUT", "HIGH_DEMAND", "MERCHANT_HIGH_DEMAND",
        "CCU_PORTAL_MERCHANT_UNCONTACTABLE", "DCU_PORTAL_MERCHANT_UNCONTACTABLE",
        "PORTAL_MERCHANT_UNCONTACTABLE", "MERCHANT_OTHERS",
        "CCU_PORTAL_MERCHANT_OUT_OF_STOCK", "CUSTOMER_OUT_OF_STOCK",
        "DCU_PORTAL_MERCHANT_OUT_OF_STOCK", "DRIVER_OUT_OF_STOCK",
        "ITEMS_OUT_OF_STOCK", "MCU_PORTAL_MERCHANT_OUT_OF_STOCK",
        "MERCHANT_ITEMS_OUT_OF_STOCK", "PORTAL_MERCHANT_OUT_OF_STOCK",
        "CCU_PORTAL_MERCHANT_CLOSED", "CUSTOMER_STORE_IS_CLOSED",
        "DCU_PORTAL_MERCHANT_CLOSED", "DRIVER_RESTAURANT_MART_CLOSED",
        "MCU_PORTAL_MERCHANT_CLOSED", "MERCHANT_RESTAURANT_CLOSED",
        "PORTAL_MERCHANT_CLOSED", "RESTAURANT_CLOSED",
        "CCU_PORTAL_MERCHANT_WRONG_PRICE", "DCU_PORTAL_MERCHANT_WRONG_PRICE",
        "MCU_PORTAL_MERCHANT_WRONG_PRICE", "PORTAL_MERCHANT_WRONG_PRICE"
    )
    # Correctly format for query string, escaping quotes inside the string
    cancel_reasons_query = "(" + " OR ".join([f'\\"{reason}\\"' for reason in cancel_reasons]) + ")"

    # Ranges for comparison
    range_from_comp_ms = str(int((datetime.fromtimestamp(int(range_from_ms) / 1000) - timedelta(days=10)).timestamp() * 1000))
    range_to_comp_ms = str(int((datetime.fromtimestamp(int(range_from_ms) / 1000) - timedelta(milliseconds=1)).timestamp() * 1000))
    
    # Build indices dynamically
    indices_current = []
    curr_idx = start_date
    while curr_idx <= end_date:
        indices_current.append(f"analytic_detail_gofood_booking_v1_{curr_idx.strftime('%Y-%m')}")
        curr_idx = (curr_idx.replace(day=1) + timedelta(days=32)).replace(day=1)
    indices_current_json = json.dumps(indices_current)
    
    indices_past = []
    curr_idx_past = datetime.fromtimestamp(int(range_from_comp_ms)/1000)
    end_date_past = datetime.fromtimestamp(int(range_to_comp_ms)/1000)
    while curr_idx_past <= end_date_past:
        indices_past.append(f"analytic_detail_gofood_booking_v1_{curr_idx_past.strftime('%Y-%m')}")
        curr_idx_past = (curr_idx_past.replace(day=1) + timedelta(days=32)).replace(day=1)
    indices_past_json = json.dumps(indices_past)

    if active_store:
        merchant_filter = f' AND merchant_id:(\\"{active_store}\\")'
    else:
        merchant_filter = ''

    # Reconstruct the full 6-query payload from the curl command
    payload_batal = (
        f'{{"search_type":"query_then_fetch","ignore_unavailable":true,"index":{indices_current_json}}}\n'
        f'{{"size":0,"query":{{"bool":{{"filter":[{{"query_string":{{"analyze_wildcard":true,"query":"time:>={range_from_ms} AND time:<={range_to_ms}{merchant_filter} AND NOT id:FP* AND _exists_:data.restaurant_accepted_timestamp"}}}}]}}}}}}\n'
        f'{{"search_type":"query_then_fetch","ignore_unavailable":true,"index":{indices_current_json}}}\n'
        f'{{"size":0,"query":{{"bool":{{"filter":[{{"query_string":{{"analyze_wildcard":true,"query":"time:>={range_from_ms} AND time:<={range_to_ms}{merchant_filter} AND NOT id:FP*"}}}}]}}}}}}\n'
        f'{{"search_type":"query_then_fetch","ignore_unavailable":true,"index":{indices_past_json}}}\n'
        f'{{"size":0,"query":{{"bool":{{"filter":[{{"query_string":{{"analyze_wildcard":true,"query":"time:>={range_from_comp_ms} AND time:<={range_to_comp_ms}{merchant_filter} AND NOT id:FP* AND _exists_:data.restaurant_accepted_timestamp"}}}}]}}}}}}\n'
        f'{{"search_type":"query_then_fetch","ignore_unavailable":true,"index":{indices_past_json}}}\n'
        f'{{"size":0,"query":{{"bool":{{"filter":[{{"query_string":{{"analyze_wildcard":true,"query":"time:>={range_from_comp_ms} AND time:<={range_to_comp_ms}{merchant_filter} AND NOT id:FP*"}}}}]}}}}}}\n'
        f'{{"search_type":"query_then_fetch","ignore_unavailable":true,"index":{indices_current_json}}}\n'
        f'{{"size":0,"query":{{"bool":{{"filter":[{{"query_string":{{"analyze_wildcard":true,"query":"time:>={range_from_ms} AND time:<={range_to_ms}{merchant_filter} AND NOT id:FP* AND _exists_:data.cancel_reason_code AND data.cancel_reason_code:{cancel_reasons_query}"}}}}]}}}},"aggs":{{"2":{{"date_histogram":{{"field":"time","min_doc_count":0,"extended_bounds":{{"min":{range_from_ms},"max":{range_to_ms}}},"format":"epoch_millis","time_zone":"Asia/Jakarta","interval":"1d"}},"aggs":{{}}}}}}}}\n'
        f'{{"search_type":"query_then_fetch","ignore_unavailable":true,"index":{indices_past_json}}}\n'
        f'{{"size":0,"query":{{"bool":{{"filter":[{{"query_string":{{"analyze_wildcard":true,"query":"time:>={range_from_comp_ms} AND time:<={range_to_comp_ms}{merchant_filter} AND NOT id:FP* AND _exists_:data.cancel_reason_code AND data.cancel_reason_code:{cancel_reasons_query}"}}}}]}}}}}}\n'
    )

    headers_batal = headers.copy()
    headers_batal.update({
        'Content-Type': 'application/x-ndjson',
        'x-dashboard-id': '83',
        'x-panel-id': '40',
        'x-ref-ids': 'A;B;C;D;E;F',
        'x-custom-merchant-id': active_store,
        'x-comp-range-offset': '10d'
    })

    try:
        # DEBUG: Check if merchant_id is correctly placed before sending
        if active_store and active_store not in payload_batal:
            console.print(f"[warning]   ⚠️ WARNING: Merchant ID {active_store} not found in Order Batal payload![/warning]")
        
        with console.status("[bold cyan]Mengambil data Order Batal...", spinner="bouncingBar"):
            response_batal = session.post(url_batal, headers=headers_batal, data=payload_batal, timeout=15)
            
        if response_batal.status_code == 200:
            data_batal = response_batal.json()
            console.print(f"[success]✅ Berhasil mengambil data Order Batal.[/success]")
        else:
            console.print(f"[warning]⚠️ Order Batal endpoint returned status {response_batal.status_code}. Defaulting to 0.[/warning]")
            data_batal = {'responses': []} # Default to empty on HTTP error
    except Exception as e:
        console.print(f"[error]⚠️ Error during Order Batal request: {e}. Defaulting to 0.[/error]")
        data_batal = {'responses': []} # Default to empty on connection error

    # --- 4. PARSING DAN PENGGABUNGAN DATA ---
    totals = {label: {'revenue': 0.0, 'orders': 0, 'order_batal': 0, 'net_revenue': 0.0, 'komisi': 0.0, 'ojol_commission': 0.0, 'pengeluaran_iklan': 0.0} for _, label in period_iter}
    print(f"DEBUG [Totals] period count: {len(period_iter)}, sample labels (first 5): {[l for _,l in period_iter[:5]]}")

    def get_buckets(d):
        if isinstance(d, dict):
            if 'buckets' in d and isinstance(d['buckets'], list):
                return d['buckets']
            for k, v in d.items():
                res = get_buckets(v)
                if res is not None:
                    return res
        elif isinstance(d, list):
            for item in d:
                res = get_buckets(item)
                if res is not None:
                    return res
        return None

    def get_period_label(ts_ms):
        dt = datetime.fromtimestamp(ts_ms / 1000.0)
        if custom_mode:
            return dt.strftime('%d %b %Y')
        return dt.strftime('%B %Y')

    # Ekstraksi Revenue
    buckets_rev = get_buckets(data_revenue)
    print(f"DEBUG [Revenue] buckets_rev found: {buckets_rev is not None}, count: {len(buckets_rev) if buckets_rev else 0}")
    if buckets_rev and len(buckets_rev) > 0:
        print(f"DEBUG [Revenue] sample bucket[0]: {json.dumps(buckets_rev[0])}")
    else:
        print(f"DEBUG [Revenue] raw response keys: {list(data_revenue.keys()) if isinstance(data_revenue, dict) else type(data_revenue)}")
        print(f"DEBUG [Revenue] raw response (first 500 chars): {json.dumps(data_revenue)[:500]}")
    if buckets_rev:
        for b in buckets_rev:
            ts = b.get('key')
            if not ts:
                continue
            val = 0.0
            for k, v in b.items():
                if isinstance(v, dict) and 'value' in v:
                    val = float(v['value'])
                    break
            label = get_period_label(ts)
            if label in totals:
                totals[label]['revenue'] += val

    # Ekstraksi Orders
    buckets_ord = None
    if 'responses' in data_orders and len(data_orders['responses']) > 0:
        buckets_ord = get_buckets(data_orders['responses'][0])
    print(f"DEBUG [Orders] buckets_ord found: {buckets_ord is not None}, count: {len(buckets_ord) if buckets_ord else 0}")
    if buckets_ord and len(buckets_ord) > 0:
        print(f"DEBUG [Orders] sample bucket[0]: {json.dumps(buckets_ord[0])}")
    else:
        print(f"DEBUG [Orders] raw data_orders keys: {list(data_orders.keys()) if isinstance(data_orders, dict) else type(data_orders)}")

    if buckets_ord:
        for b in buckets_ord:
            ts = b.get('key')
            if not ts:
                continue
            val = int(b.get('doc_count', 0))
            label = get_period_label(ts)
            if label in totals:
                totals[label]['orders'] += val

    # Ekstraksi Omzet Bersih
    buckets_net = get_buckets(data_net)
    print(f"DEBUG [OmzetBersih] buckets_net found: {buckets_net is not None}, count: {len(buckets_net) if buckets_net else 0}")
    if buckets_net and len(buckets_net) > 0:
        print(f"DEBUG [OmzetBersih] sample bucket[0]: {json.dumps(buckets_net[0])}")
    else:
        print(f"DEBUG [OmzetBersih] raw response (first 500 chars): {json.dumps(data_net)[:500]}")

    if buckets_net:
        for b in buckets_net:
            ts = b.get('key')
            if not ts:
                continue
            val_net = 0.0
            for k, v in b.items():
                if isinstance(v, dict) and 'value' in v:
                    val_net = float(v['value'])
                    break
            label = get_period_label(ts)
            if label in totals:
                totals[label]['net_revenue'] += val_net

    # Ekstraksi Komisi
    buckets_komisi = get_buckets(data_komisi)

    if buckets_komisi:
        for b in buckets_komisi:
            ts = b.get('key')
            if not ts:
                continue
            val_komisi = 0.0
            for k, v in b.items():
                if isinstance(v, dict) and 'value' in v:
                    val_komisi = float(v['value'])
                    break
            label = get_period_label(ts)
            if label in totals:
                totals[label]['komisi'] += val_komisi

    # Ekstraksi Iklan & Diskon
    buckets_iklan = get_buckets(data_iklan)

    if buckets_iklan:
        for b in buckets_iklan:
            ts = b.get('key')
            if not ts:
                continue
            val_iklan = 0.0
            for k, v in b.items():
                if isinstance(v, dict) and 'value' in v:
                    val_iklan = float(v['value'])
                    break
            label = get_period_label(ts)
            if label in totals:
                totals[label]['pengeluaran_iklan'] += val_iklan

    # Ekstraksi Order Batal dengan Defensive Parsing
    val_order_batal_total = 0
    responses = data_batal.get('responses', [])
    batal_from_buckets = False
    
    # Pastikan array responses cukup panjang (minimal ada 5 elemen untuk `responses[4]`)
    if len(responses) > 4:
        cancel_data_current = responses[4] # Indeks ke-4 adalah pembatalan saat ini
        
        # DEBUG: Cetak mentahan responses[4]
        print("DEBUG - Raw Response Order Batal [4]:", json.dumps(cancel_data_current))
        
        # Cek apakah request sukses (status 200) dan TIDAK ada key 'error'
        if cancel_data_current.get('status') == 200 and 'error' not in cancel_data_current:
            # Ambil data dengan aman
            val_order_batal_total = cancel_data_current.get('hits', {}).get('total', {}).get('value', 0)
            
            buckets_batal = get_buckets(cancel_data_current)
            if buckets_batal:
                for b in buckets_batal:
                    ts = b.get('key')
                    if not ts:
                        continue
                    val_batal = int(b.get('doc_count', 0))
                    label = get_period_label(ts)
                    if label in totals:
                        totals[label]['order_batal'] += val_batal
                batal_from_buckets = True
        else:
            # Jika ada error di dalam response, log dan lanjutkan dengan nilai 0
            error_details = cancel_data_current.get('error', 'Unknown Error')
            print(f"   -> Info: Elasticsearch returned an error for Order Batal query: {error_details}")
    else:
        # Jika jumlah response tidak sesuai, berarti ada masalah besar dengan request
        print(f"   -> Info: Expected 6 responses for Order Batal, but got {len(responses)}. Defaulting to 0.")

    if val_order_batal_total > 0 and not batal_from_buckets:
        # Distribusi order batal ke setiap periode secara merata tanpa menghilangkan sisa (remainder)
        if len(period_iter) > 0:
            base_val = int(val_order_batal_total / len(period_iter))
            remainder = val_order_batal_total % len(period_iter)
            for i, (_, label) in enumerate(period_iter):
                if label in totals:
                    totals[label]['order_batal'] = base_val + (1 if i < remainder else 0)

    # Hitung Komisi Ojol (Potongan dari GoFood)
    for label in totals:
        totals[label]['ojol_commission'] = totals[label]['revenue'] - totals[label]['net_revenue']

    # Hitung keseluruhan rata-rata
    num_periods = len(period_iter)
    total_omzet = sum(totals[label]['revenue'] for _, label in period_iter)
    total_omzet_bersih = sum(totals[label]['net_revenue'] for _, label in period_iter)
    total_komisi = sum(totals[label]['komisi'] for _, label in period_iter)
    total_ojol = sum(totals[label]['ojol_commission'] for _, label in period_iter)
    total_order = sum(totals[label]['orders'] for _, label in period_iter)
    total_order_batal = sum(totals[label]['order_batal'] for _, label in period_iter)
    
    avg_omzet = int(total_omzet / num_periods) if num_periods > 0 else 0
    avg_omzet_bersih = int(total_omzet_bersih / num_periods) if num_periods > 0 else 0
    avg_komisi = int(total_komisi / num_periods) if num_periods > 0 else 0
    avg_pendapatan_ojol = int(total_ojol / num_periods) if num_periods > 0 else 0
    avg_order = round(total_order / num_periods) if num_periods > 0 else 0
    avg_order_batal = round(total_order_batal / num_periods, 2) if num_periods > 0 else 0

    pass

    # --- 5. TULIS KE FILE EXCEL (DINONAKTIFKAN SESUAI PERMINTAAN) ---
    # File revenue_... (Format Vertikal) tidak lagi dihasilkan
    pass

    # --- 6. EXPORT KE EXCEL PER OUTLET (Raw Data) ---
    username = os.getenv('ACTIVE_NOMOR_HP', 'Tidak Diketahui')
    nama_outlet = os.getenv('ACTIVE_NAMA_OUTLET', 'Tidak Tersedia')
    cabang = os.getenv('ACTIVE_CABANG', 'Tidak Tersedia')
    store_id = os.getenv('ACTIVE_STORE_ID', 'Tidak Tersedia')

    safe_name_str = f"{_outlet}_{_cabang}_{store_id}" if _cabang and _cabang.lower() != 'tanpa cabang' else f"{_outlet}_{store_id}"
    safe_outlet = safe_name_str.strip().replace(" ", "_").replace("/", "_").replace("\\", "_")
    if not safe_outlet or safe_outlet == "Tidak_Tersedia":
        safe_outlet = "Unknown_Outlet"
        
    date_folder = f"{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}"
    if GLOBAL_OUTPUT_DIR:
        raw_gofood_dir = GLOBAL_OUTPUT_DIR
    else:
        raw_gofood_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'laporan', 'gofood', date_folder)
    os.makedirs(raw_gofood_dir, exist_ok=True)
    
    base_raw_excel_filename = os.path.join(raw_gofood_dir, f"{safe_outlet}.xlsx")
    raw_excel_filename = base_raw_excel_filename
    if os.path.exists(raw_excel_filename):
        version = 1
        while os.path.exists(os.path.join(raw_gofood_dir, f"{safe_outlet}_v{version}.xlsx")):
            version += 1
        raw_excel_filename = os.path.join(raw_gofood_dir, f"{safe_outlet}_v{version}.xlsx")
        
    abs_raw_excel_path = os.path.abspath(raw_excel_filename)
    
    try:
        wb_raw = openpyxl.Workbook()
        ws_raw = wb_raw.active
        headers_excel = [
            'Tanggal', 'Outlet Name', 'Store ID', 'Penjualan Kotor', 'Biaya Komisi', 
            'Pengeluaran Iklan & Diskon', 'Order Sukses', 'Order Batal'
        ]
        ws_raw.append(headers_excel)
                
        # Data rows
        for idx, (raw_date, label) in enumerate(period_iter):
            omzet = int(totals[label]['revenue'])
            omzet_bersih = int(totals[label]['net_revenue'])
            komisi = int(totals[label]['komisi'])
            pendapatan_ojol = int(totals[label]['ojol_commission'])
            order = int(totals[label]['orders'])
            iklan = int(totals[label].get('pengeluaran_iklan', 0))
            order_sukses = int(totals[label]['orders'])
            order_batal = int(totals[label]['order_batal'])
            rata_rata_order_per_cust = int(omzet / order_sukses) if order_sukses > 0 else 0
            total_order_row = order_sukses
            
            row_data = [
                raw_date,
                nama_outlet,
                store_id,
                omzet,
                komisi,
                iklan,
                order_sukses,
                order_batal
            ]
            ws_raw.append(row_data)
            
        wb_raw.save(abs_raw_excel_path)
        wb_raw.close()
        
        print(f"✅ Juga di-export ke Excel Raw per outlet:")
        print(f"   {abs_raw_excel_path}")

        # --- Kirim ke GSheet Harian (Dihapus dari sini karena sudah ditangani via cli.py / send_data secara terpisah) ---
            
    except Exception as e:
        print(f"⚠️ Peringatan: Gagal membuat file Excel Raw: {e}")

    for label, data in totals.items():
        console.print(f"[dim]- {label}: Omzet {int(data['revenue'])} | Omzet Bersih {int(data['net_revenue'])} | Komisi {int(data['komisi'])} | Potongan Ojol {int(data['ojol_commission'])} | Pesanan: {data['orders']} | Batal: {data['order_batal']}[/dim]")

    # Table Ringkasan Per Store
    table = Table(title=f"Ringkasan: {os.getenv('ACTIVE_NAMA_OUTLET')} ({os.getenv('ACTIVE_STORE_ID')})", show_header=True, header_style="bold magenta")
    table.add_column("Metrik", style="dim")
    table.add_column("Total", justify="right", style="bold")
    table.add_column("Rata-rata", justify="right")

    table.add_row("Omzet Kotor", f"Rp {int(total_omzet):,}", f"Rp {avg_omzet:,}")
    table.add_row("Omzet Bersih", f"Rp {int(total_omzet_bersih):,}", f"Rp {avg_omzet_bersih:,}")
    table.add_row("Komisi GoBiz", f"Rp {int(total_komisi):,}", f"Rp {avg_komisi:,}")
    table.add_row("Potongan Ojol", f"Rp {int(total_ojol):,}", f"Rp {avg_pendapatan_ojol:,}")
    table.add_row("Order Sukses", f"{total_order}", f"{avg_order}")
    table.add_row("Order Batal", f"{total_order_batal}", f"{avg_order_batal}")

    console.print("\n", table)
    
    DURATION_STORE = time.time() - START_TIME_STORE
    console.print(f"[info]⏱️ Waktu proses untuk store ini: [bold]{DURATION_STORE:.2f} detik[/bold][/info]\n")
    
    # Return structured data for baseline aggregation if requested
    if return_data:
        return {
            'period_iter': period_iter,
            'totals': totals,
            'avg_omzet_bersih': avg_omzet_bersih,
            'avg_order': avg_order,
            'total_omzet_bersih': int(total_omzet_bersih),
            'total_order': total_order,
        }

    return None

def tulis_baseline_excel(all_results, start_date, end_date, outlet_filter=None):
    """
    Menulis file Excel baseline bergaya Shopee:
    1 baris per outlet, kolom per bulan (omzet bersih + order sukses), plus rata-rata.

    all_results: list of dict {
        'nama_outlet': str,
        'result': dict dari ambil_data_analytics(..., return_data=True)
    }
    """
    if not all_results:
        return

    # Ambil first result untuk verifikasi data
    first_result = next((r['result'] for r in all_results if r.get('result')), None)
    if not first_result:
        console.print("[warning]⚠️ Tidak ada data untuk ditulis ke baseline.[/warning]")
        return

    # Hitung daftar bulan kalender dari start_date ke end_date
    months_iter = []
    curr = start_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    while curr <= end_date:
        month_label = curr.strftime('%B %Y')  # e.g., "February 2026"
        month_key = curr.strftime('%Y-%m')    # e.g., "2026-02"
        months_iter.append((month_key, month_label))
        
        # Increment to next month
        next_month = curr.month + 1
        next_year = curr.year
        if next_month > 12:
            next_month = 1
            next_year += 1
        curr = curr.replace(year=next_year, month=next_month, day=1)

    num_periods = len(months_iter)

    # Susun header:
    # Merchant | Aplikasi | Omzet Bulan ke-1 | Order Bulan ke-1 | ... | Rata-rata Omzet | Rata-rata Order
    headers_bl = ['Merchant', 'Aplikasi']
    for i in range(num_periods):
        headers_bl.append(f'Omzet Bulan ke-{i + 1}')
        headers_bl.append(f'Order Bulan ke-{i + 1}')
    headers_bl.append('Rata-rata Omzet')
    headers_bl.append('Rata-rata Order')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Baseline Summary'

    # Styling header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.append(headers_bl)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for entry in all_results:
        nama = entry.get('nama_outlet', 'Tidak Diketahui')
        result = entry.get('result')
        if not result:
            continue

        totals = result['totals']
        
        # Agregasi data harian/bulanan ke bulan kalender masing-masing
        monthly_totals = {label: {'revenue': 0.0, 'orders': 0} for _, label in months_iter}
        for label, val in totals.items():
            # Jika key bertipe harian "01 Feb 2026"
            try:
                dt = datetime.strptime(label, "%d %b %Y")
                m_label = dt.strftime("%B %Y")
            except ValueError:
                # Jika key bertipe bulanan "February 2026"
                m_label = label
            
            if m_label in monthly_totals:
                monthly_totals[m_label]['revenue'] += val.get('revenue', 0.0)
                monthly_totals[m_label]['orders'] += val.get('orders', 0)

        # Hitung rata-rata bulanan
        total_omzet_kotor = sum(monthly_totals[lbl]['revenue'] for _, lbl in months_iter)
        total_order = sum(monthly_totals[lbl]['orders'] for _, lbl in months_iter)
        
        avg_omzet_kotor = int(total_omzet_kotor / num_periods) if num_periods > 0 else 0
        avg_order = round(total_order / num_periods) if num_periods > 0 else 0

        row = [nama, 'GoFood']
        for _, label in months_iter:
            omzet_kotor_bulan = int(monthly_totals[label]['revenue'])
            order_bulan = int(monthly_totals[label]['orders'])
            row.append(omzet_kotor_bulan)
            row.append(order_bulan)
        row.append(avg_omzet_kotor)
        row.append(avg_order)
        ws.append(row)

    # Column widths
    ws.column_dimensions['A'].width = 30

    ws.column_dimensions['B'].width = 12
    for col_idx in range(3, len(headers_bl) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 18

    # Build filename using date range
    start_str = start_date.strftime('%Y-%m-%d')
    end_str = end_date.strftime('%Y-%m-%d')
    
    global GLOBAL_OUTPUT_DIR
    if GLOBAL_OUTPUT_DIR:
        output_dir = GLOBAL_OUTPUT_DIR
    else:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'laporan_gofood')
        
    os.makedirs(output_dir, exist_ok=True)
    
    if outlet_filter:
        safe_o = str(outlet_filter).strip().replace(" ", "_").replace("/", "_").replace("\\", "_").replace("|", "_")
        filename = os.path.join(output_dir, f'BASELINE_CUSTOM_GOFOOD_{safe_o}_{start_str}_to_{end_str}.xlsx')
    else:
        filename = os.path.join(output_dir, f'BASELINE_GOFOOD_{start_str}_to_{end_str}.xlsx')

    wb.save(filename)
    wb.close()
    console.print(f"\n[success]✅ Baseline Excel berhasil disimpan ke:[/success]")
    console.print(f"   [bold]{filename}[/bold]")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="GoFood Analytics Scraper")
    parser.add_argument("--start-date", type=str, default=None, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", type=str, default=None, help="End date (YYYY-MM-DD)")
    parser.add_argument("--outlet", type=str, default=None, help="Filter specific outlet name")
    parser.add_argument("--branch", type=str, default=None, help="Filter specific branch name")
    parser.add_argument("--output-dir", type=str, default=None, help="Directory to save output files")
    parser.add_argument("--no-proxy", action="store_true", help="Nonaktifkan proxy/WARP untuk sesi ini")
    parser.add_argument("--no-sheet", action="store_true", help="Nonaktifkan pengiriman data ke Google Sheets")
    parser.add_argument("--task", type=str, default="2", help="Task choice: 1 for baseline, 2 for weekly, 3 for VB")
    args_cli = parser.parse_args()

    if args_cli.output_dir:
        GLOBAL_OUTPUT_DIR = args_cli.output_dir

    if args_cli.no_proxy:
        os.environ["USE_PROXY"] = "false"
        console.print("[info]🚫 Proxy/WARP dinonaktifkan untuk sesi ini.[/info]")


    console.print(Panel.fit(
        "[bold magenta]🚀 GOFOOD ANALYTICS SCRAPER v2.0[/bold magenta]\n[dim]Developed with Rich UI & Performance Tracking[/dim]",
        border_style="magenta",
        padding=(1, 2)
    ))
    
    console.print("[info]📋 Mengambil daftar akun GoFood dari Google Sheet master...[/info]")

    # --- Ambil semua akun GoFood Live dari Google Sheet ---
    sheet_accounts = []
    with console.status("[bold blue]Menghubungi Google Sheet...", spinner="earth"):
        sheet_accounts = fetch_gofood_accounts_from_sheet(task=args_cli.task)

    if sheet_accounts:
        console.print(f"[success]✅ Ditemukan {len(sheet_accounts)} akun GoFood Live dari Google Sheet.[/success]\n")
    else:
        console.print("[warning]⚠️ Tidak bisa mengambil data dari Google Sheet. Fallback ke akun di .env...[/warning]")
        send_discord_error("GoFood", "Global", "NO_DATA", "Gagal membaca data dari Google Sheets Master. Sistem terpaksa menggunakan data cadangan dari environment lokal.", "")

    # --- Bangun token map dari .env ---
    # Format yang didukung: BEARER_TOKEN_{phone}_{nama}, BEARER_TOKEN (global)
    token_map = {}  # key: phone (normalized) -> token
    global_token = os.getenv('BEARER_TOKEN', '')

    for key, value in os.environ.items():
        if key.startswith('BEARER_TOKEN_') and value:
            # Ambil prefix phone number (bagian pertama setelah BEARER_TOKEN_)
            suffix = key[len('BEARER_TOKEN_'):]
            phone_part = suffix.split('_')[0]
            phone_norm = normalize_phone(phone_part)
            if phone_norm:
                token_map[phone_norm] = value

    # --- Gabungkan data sheet + token .env ---
    # Buat list: (phone, nama_outlet, cabang, store_id, token)
    resolved_accounts = []

    if sheet_accounts:
        for acc in sheet_accounts:
            phone_raw = acc['phone']
            phone_norm = normalize_phone(phone_raw)
            # Jika phone_norm kosong/tidak valid (seperti "-" atau ""), coba gunakan email jika ada
            if (not phone_norm or phone_norm == "-") and acc.get('email'):
                phone_norm = normalize_phone(acc['email'])
                
            # Cari token yang cocok
            token = token_map.get(phone_norm, '')
            if not token:
                # Coba partial match (phone dari sheet mungkin punya prefix 62)
                for k, v in token_map.items():
                    if k.endswith(phone_norm) or phone_norm.endswith(k):
                        token = v
                        break
            
            # Fallback 1: Coba cari langsung di env menggunakan expected suffix baru
            if not token:
                sanitized_resto_name = re.sub(r'[^a-zA-Z0-9]', '', (acc['cabang'] or 'Tanpa Cabang') or acc['nama_outlet'])
                expected_suffix = f"_{phone_norm}_{sanitized_resto_name}"
                token = os.getenv(f"BEARER_TOKEN{expected_suffix}", "")
                
            # Fallback 2: Coba format lama jika phone_norm didapat dari email (artinya phone asli kosong/invalid)
            if not token and '@' in phone_norm:
                legacy_suffix = f"__{sanitized_resto_name}"
                token = os.getenv(f"BEARER_TOKEN{legacy_suffix}", "")
                
            # JANGAN gunakan global_token fallback untuk data dari Google Sheet agar status login presisi per outlet.
            resolved_accounts.append({
                'phone'      : phone_norm,
                'phone_raw'  : phone_raw,
                'email'      : acc.get('email', ''),
                'emails'     : acc.get('emails', []),
                'nama_outlet': acc['nama_outlet'],
                'cabang'     : acc['cabang'],
                'store_id'   : acc['store_id'],
                'token'      : token,
            })
    else:
        # Fallback ke akun dari .env saja
        for key, value in os.environ.items():
            if key.startswith('BEARER_TOKEN_') and value:
                suffix = key[len('BEARER_TOKEN_'):]
                parts = suffix.split('_', 1)
                phone = normalize_phone(parts[0])
                resolved_accounts.append({
                    'phone'      : phone,
                    'phone_raw'  : parts[0],
                    'nama_outlet': os.getenv(f'NAMA_OUTLET_{suffix}', 'Tidak Tersedia'),
                    'cabang'     : os.getenv(f'CABANG_{suffix}', 'Tanpa Cabang'),
                    'store_id'   : '',
                    'token'      : value,
                })

    if not resolved_accounts:
        console.print("[error]❌ Tidak ada akun GoFood yang ditemukan. Silakan login via LoginManual.py terlebih dahulu.[/error]")
        exit(1)

    if args_cli.outlet:
        target_outlets = [x.strip().lower() for x in args_cli.outlet.split('|') if x.strip()]
        resolved_accounts = [a for a in resolved_accounts if any(t in a['nama_outlet'].lower() for t in target_outlets)]
        if args_cli.branch:
            target_branches = [x.strip().lower() for x in args_cli.branch.split('|') if x.strip()]
            resolved_accounts = [a for a in resolved_accounts if any(t in a['cabang'].lower() for t in target_branches)]
        if not resolved_accounts:
            branch_err = f" dan cabang: {args_cli.branch}" if args_cli.branch else ""
            console.print(f"[error]⚠️ Tidak ditemukan akun GoFood dengan outlet: {args_cli.outlet}{branch_err}[/error]")
            exit(1)
        branch_msg = f" dan cabang: {args_cli.branch}" if args_cli.branch else ""
        console.print(f"[success]✅ Filter outlet: {args_cli.outlet}{branch_msg} ({len(resolved_accounts)} akun)[/success]")
    else:
        if args_cli.start_date and args_cli.end_date:
            console.print(f"[info]📋 Mode non-interaktif: Memproses semua ({len(resolved_accounts)}) outlet GoFood.[/info]")
        else:
            console.print("[bold]Daftar Outlet GoFood yang tersedia:[/bold]")
            for i, acc in enumerate(resolved_accounts, 1):
                status_token = "[green]✓ Token[/green]" if acc['token'] else "[red]✗ Belum Login[/red]"
                cabang_info = f" - {acc['cabang']}" if acc['cabang'] and acc['cabang'] != 'Tanpa Cabang' else ""
                store_info = f" (Store: {acc['store_id']})" if acc['store_id'] else ""
                console.print(f"  [{i}] {acc['nama_outlet']}{cabang_info}{store_info} | {status_token}")

            console.print()
            pilihan_cabang = input("Pilih nomor outlet (contoh: 1 atau 1,3 atau [Enter] untuk semua): ").strip().lower()
            if pilihan_cabang not in ['all', 'semua', '']:
                selected_indices = []
                for p in pilihan_cabang.split(','):
                    if p.strip().isdigit():
                        idx = int(p.strip()) - 1
                        if 0 <= idx < len(resolved_accounts):
                            selected_indices.append(idx)
                if selected_indices:
                    resolved_accounts = [resolved_accounts[idx] for idx in selected_indices]
                    console.print(f"[success]✅ Memproses {len(resolved_accounts)} outlet pilihan.[/success]")



    # --- LOGIKA TANGGAL ---
    if args_cli.start_date and args_cli.end_date:
        custom_start_date = parse_tanggal_input(args_cli.start_date)
        custom_end_date = parse_tanggal_input(args_cli.end_date)
        print(f"\n✅ Menggunakan range dari CLI: {args_cli.start_date} s/d {args_cli.end_date}\n")
    else:
        custom_start_date, custom_end_date = minta_range_tanggal_custom()
        if custom_start_date and custom_end_date:
            print(f"\n✅ Menggunakan range custom: {custom_start_date.strftime('%d-%m-%Y')} s/d {custom_end_date.strftime('%d-%m-%Y')}\n")
        else:
            print("\n✅ Menggunakan range default 3 bulan ke belakang.\n")
    
    all_baseline_results = []

    for index, acc in enumerate(resolved_accounts):
        phone         = acc['phone']
        nama_outlet   = acc['nama_outlet']
        cabang        = acc['cabang'] or 'Tanpa Cabang'
        store_id      = acc['store_id']
        if store_id == "None" or store_id == "NaN":
            store_id = ""
        explicit_store_id = store_id
            
        token         = acc['token']

        # Resolve store_id from environment if missing (e.g. when parsing Baseline sheet)
        if not store_id:
            sanitized_resto_name = re.sub(r'[^a-zA-Z0-9]', '', cabang or nama_outlet)
            suffix_email = f"_{phone}_{sanitized_resto_name}"
            store_id = os.getenv(f"STORE_ID{suffix_email}", "")
            if store_id == "None" or store_id == "NaN":
                store_id = ""
            if not store_id:
                suffix_legacy = f"__{sanitized_resto_name}"
                store_id = os.getenv(f"STORE_ID{suffix_legacy}", "")
                if store_id == "None" or store_id == "NaN":
                    store_id = ""

        # ── Cek session cache JSON terlebih dahulu ──
        session_id = acc.get('email') or phone
        cached_token = _load_session_token(session_id)
        if cached_token and not token:
            token = cached_token
            acc['token'] = token

        # ── Validasi token yang sudah ada ──
        token_valid = False
        if token:
            os.environ['BEARER_TOKEN'] = token
            console.print(f"[dim]🔑 Memvalidasi sesi cache untuk {nama_outlet} ({session_id})...[/dim]")
            token_valid = ambil_data_dashboard()

        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')

        if not token_valid:
            console.print(f"[warning]⚠️ Sesi tidak valid/kosong untuk {nama_outlet} ({session_id}). Melakukan login otomatis...[/warning]")

            # Panggil fungsi login otomatis
            new_token = login_outlet_gofood_flow(acc)
            if new_token:
                token = new_token
                token_valid = True

                # Simpan token baru ke .env
                sanitized_resto_name = re.sub(r'[^a-zA-Z0-9]', '', cabang or nama_outlet)
                suffix = f"_{phone}_{sanitized_resto_name}"
                os.environ['BEARER_TOKEN'] = token  # backward compat
                env_lock = _FileLock(f"{env_path}.lock", timeout=15)
                with env_lock:
                    set_key(env_path, f"BEARER_TOKEN{suffix}", token)
                    set_key(env_path, f"NAMA_OUTLET{suffix}", str(nama_outlet))
                    set_key(env_path, f"CABANG{suffix}", str(cabang))
                    set_key(env_path, f"STORE_ID{suffix}", str(store_id))

                # Simpan juga ke session JSON cache
                _save_session_token(session_id, token, meta={
                    'nama_outlet': nama_outlet,
                    'cabang': cabang,
                    'store_id': store_id,
                })

                console.print(f"[success]✅ Token berhasil ditangkap dan disimpan ke .env + session cache untuk {nama_outlet}.[/success]")

                # Update token untuk akun dengan email/phone yang sama agar tidak login ulang
                for other_acc in resolved_accounts:
                    if other_acc.get('email') == session_id or other_acc['phone'] == phone:
                        other_acc['token'] = token
                        _save_session_token(
                            other_acc.get('email') or other_acc['phone'],
                            token,
                            meta={
                                'nama_outlet': other_acc['nama_outlet'],
                                'cabang': other_acc.get('cabang', ''),
                                'store_id': other_acc.get('store_id', ''),
                            }
                        )

                os.environ['BEARER_TOKEN'] = token
            else:
                console.print(f"[error]❌ Gagal login untuk {nama_outlet}. Melewati outlet ini.[/error]")
                continue
        else:
            console.print(f"[success]✅ Sesi valid untuk {nama_outlet}. Melewati login.[/success]")

        # Set environment untuk iterasi ini
        # Set token ke env hanya untuk backward compat (proses lain tidak terpengaruh
        # karena setiap akun sudah di-pass via parameter ke ambil_data_analytics)
        os.environ['BEARER_TOKEN']       = token
        os.environ['ACTIVE_NOMOR_HP']    = phone
        os.environ['ACTIVE_NAMA_OUTLET'] = nama_outlet
        os.environ['ACTIVE_CABANG']      = cabang
        
        targets_to_process = []
        
        if token:
            console.print(f"[dim]Mencari profil cabang dari API GoBiz untuk nama spesifik...[/dim]")
            try:
                url_search = "https://api.gobiz.co.id/v1/merchants/search"
                headers_search = {
                    "Accept": "application/json, text/plain, */*",
                    "Authentication-Type": "go-id",
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json"
                }
                payload_search = {
                    "from": 0, "size": 1000, 
                    "_source": ["id","director_name","merchant_name","email","feature_types","phone","outlet_address","outlet_name","outlet_city","payment_settings.GOPAY","tags","bank_account","applications","pops","aspi","business_type","metadata","id_type","merchant_type"]
                }
                resp_search = requests.post(url_search, headers=headers_search, json=payload_search, timeout=15)
                if resp_search.status_code == 200:
                    data_search = resp_search.json()
                    hits = data_search.get("hits", [])
                    
                    if hits:
                        console.print(f"[success]✅ Ditemukan {len(hits)} cabang/merchant ID dalam portal ini.[/success]")
                        for h in hits:
                            t_id = h.get("id", "")
                            
                            if explicit_store_id and t_id != explicit_store_id:
                                continue
                                
                            t_nama = h.get("merchant_name", nama_outlet)
                            t_cabang = h.get("outlet_name", cabang)
                            
                            # Gunakan nama cabang sebagai nama outlet jika tersedia,
                            # agar laporan Excel menampilkan nama presisi (contoh: Nasi Goreng Cinara, Ampel)
                            resolved_nama = t_cabang if (t_cabang and t_cabang.lower() != 'tanpa cabang') else t_nama
                            
                            targets_to_process.append({
                                "store_id": t_id,
                                "nama_outlet": resolved_nama,
                                "cabang": t_cabang
                            })
                            
                            if t_id:
                                sanitized_resto_name = re.sub(r'[^a-zA-Z0-9]', '', t_cabang or t_nama)
                                suffix = f"_{phone}_{sanitized_resto_name}"
                                env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
                                env_lock = _FileLock(f"{env_path}.lock", timeout=15)
                                with env_lock:
                                    set_key(env_path, f"STORE_ID{suffix}", str(t_id))
                else:
                    console.print(f"[warning]⚠️ API GoBiz merespons dengan HTTP {resp_search.status_code}[/warning]")
            except Exception as e:
                console.print(f"[warning]⚠️ Gagal mengambil Store ID dari API: {e}[/warning]")

        if not targets_to_process:
            targets_to_process.append({
                "store_id": store_id,
                "nama_outlet": nama_outlet,
                "cabang": cabang
            })

        for t_idx, target in enumerate(targets_to_process):
            active_store_id = target["store_id"]
            active_nama = target["nama_outlet"]
            active_cabang = target["cabang"]

            os.environ['ACTIVE_STORE_ID']    = active_store_id
            os.environ['ACTIVE_NAMA_OUTLET'] = active_nama
            os.environ['ACTIVE_CABANG']      = active_cabang

            console.print(f"\n[bold]{'='*55}[/bold]")
            console.print(f"[bold cyan]Memproses:[/bold cyan] {active_nama} - {active_cabang} ({phone})")
            if active_store_id:
                console.print(f"[dim]Store ID: {active_store_id}[/dim]")
            console.print(f"[bold]{'='*55}[/bold]")

            console.print("\nMemulai pengambilan data analytics...")
            result = ambil_data_analytics(
                write_header=(index == 0 and t_idx == 0),
                start_date=custom_start_date,
                end_date=custom_end_date,
                return_data=True,
                token=token,
                store_id=active_store_id,
                nama_outlet=active_nama,
                phone=phone,
                cabang=active_cabang,
            )
            if result:
                label = f"{active_nama} - {active_cabang}" if active_cabang and active_cabang != 'Tanpa Cabang' else active_nama
                all_baseline_results.append({
                    'nama_outlet': label,
                    'result': result
                })


    # --- TULIS BASELINE EXCEL GABUNGAN ---
    if all_baseline_results:
        if not (custom_start_date and custom_end_date):
            now = datetime.now()
            first_day_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            _end = first_day_this_month - timedelta(milliseconds=1)
            _sm = first_day_this_month.month - 3
            _sy = first_day_this_month.year
            if _sm <= 0:
                _sm += 12
                _sy -= 1
            _start = first_day_this_month.replace(year=_sy, month=_sm, day=1)
        else:
            _start = custom_start_date
            _end = custom_end_date
        # File Baseline dihilangkan sesuai permintaan
        # tulis_baseline_excel(all_baseline_results, _start, _end, args_cli.outlet)

    console.print("\n[bold]" + "="*50 + "[/bold]")
    console.print("[success]✅ Semua proses selesai![/success]")
    console.print("[bold]" + "="*50 + "[/bold]")

